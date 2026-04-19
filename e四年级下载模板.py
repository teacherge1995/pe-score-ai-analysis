from openpyxl import Workbook
from openpyxl.styles import Alignment, Font,Border,Side


def generate_4_grade_template(school_name, header_name, selected_items, boy_count, girl_count, save_path,file_name):
    """
    生成一年级体育成绩模板
    :param school_name: 学校名称
    :param header_name: 表头名称
    :param selected_items: 选中的自动填充成绩项目
    :param boy_count: 男生人数
    :param girl_count: 女生人数
    :param save_path: 保存路径
    """
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{file_name}体育成绩"

    # 填充学校名称和表头===========男生==================
    ws.merge_cells('A1:K1')
    ws['A1'] = header_name
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


    ws.merge_cells('A2:B2')
    ws['A2'] = '学校名称：'
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C2:E2')
    ws['C2'] = school_name
    ws['C2'].font = Font(size=12, bold=True)
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F2'] = '班级：'
    ws['F2'].font = Font(size=12, bold=True)
    ws['F2'].alignment = Alignment(horizontal='center', vertical='center')
    # 修复后
    ws.merge_cells('G2:I2')
    # 文件名是字符串（如"一年级1班"），无需转float，且仅操作合并区域左上角
    ws['G2'] = file_name  # 去掉float转换，直接赋值字符串
    ws['G2'].font = Font(size=12, bold=True)
    ws['G2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:B3')
    ws['A3'] = '教师姓名：'
    ws['A3'].font = Font(size=12, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C3:E3')
    ws['F3'] = '测试员：'
    ws['F3'].font = Font(size=12, bold=True)
    ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('G3:I3')

    # 填充学校名称和表头===========女生==================
    ws.merge_cells('A33:K33')
    ws['A33'] = header_name
    ws['A33'].font = Font(size=14, bold=True)
    ws['A33'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A34:B34')
    ws['A34'] = '学校名称：'
    ws['A34'].font = Font(size=12, bold=True)
    ws['A34'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C34:E34')
    ws['C34'] = school_name
    ws['C34'].font = Font(size=12, bold=True)
    ws['C34'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F34'] = '班级：'
    ws['F34'].font = Font(size=12, bold=True)
    ws['F34'].alignment = Alignment(horizontal='center', vertical='center')
    # 女生班级同理修复
    ws.merge_cells('G34:I34')
    ws['G34'].value = file_name
    ws['G34'].font = Font(size=12, bold=True)
    ws['G34'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A35:B35')
    ws['A35'] = '教师姓名：'
    ws['A35'].font = Font(size=12, bold=True)
    ws['A35'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('C35:E35')
    ws['F35'] = '测试员：'
    ws['F35'].font = Font(size=12, bold=True)
    ws['F35'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('G35:I35')




    # 填充男生数据（第5行开始）
    boy_start_row = 5
    for i in range(int(boy_count)):
        row = boy_start_row + i
        # 序号
        ws.cell(row=row, column=1, value=i + 1)
        # 性别
        ws.cell(row=row, column=3, value='男')

        # 自动填充选中项目的成绩
        if '50米' in selected_items:
            ws.cell(row=row, column=4, value=8.5)
        if '跳绳' in selected_items:
            ws.cell(row=row, column=5, value=190)
        if '坐位体前屈' in selected_items:
            ws.cell(row=row, column=6, value=20)
        if '仰卧起坐' in selected_items:
            ws.cell(row=row, column=7, value=50)
        if '立定跳远' in selected_items:
            ws.cell(row=row, column=8, value=190)
        if '肺活量' in selected_items:
            ws.cell(row=row, column=9, value=2700)
        if '身高' in selected_items:
            ws.cell(row=row, column=10, value=150)
        if '体重' in selected_items:
            ws.cell(row=row, column=11, value=34)

    # 填充女生数据（第37行开始）
    girl_start_row = 37
    for i in range(int(girl_count)):
        row = girl_start_row + i
        # 序号
        ws.cell(row=row, column=1, value=i + 1)
        # 性别
        ws.cell(row=row, column=3, value='女')

        # 自动填充选中项目的成绩
        if '50米' in selected_items:
            ws.cell(row=row, column=4, value=8.5)
        if '跳绳' in selected_items:
            ws.cell(row=row, column=5, value=190)
        if '坐位体前屈' in selected_items:
            ws.cell(row=row, column=6, value=20)
        if '仰卧起坐' in selected_items:
            ws.cell(row=row, column=7, value=50)
        if '立定跳远' in selected_items:
            ws.cell(row=row, column=8, value=190)
        if '肺活量' in selected_items:
            ws.cell(row=row, column=9, value=2700)
        if '身高' in selected_items:
            ws.cell(row=row, column=10, value=150)
        if '体重' in selected_items:
            ws.cell(row=row, column=11, value=34)

    # 调整列宽
    for col in range(1, 12):
        ws.column_dimensions[chr(64 + col)].width = 8
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["F"].width = 11

    # 调整行高
    for row in range(1, 65):  # 遍历第1行到第9行
        ws.row_dimensions[row].height = 20  # 设置行高为25磅（可根据需要调整数值）
    # 设置第一行的行高（数值单位是磅，默认行高约15磅）
    ws.row_dimensions[1].height = 42  # 1代表第一行，30是行高值
    ws.row_dimensions[33].height = 42  # 1代表第一行，30是行高值
    ws.row_dimensions[4].height = 25  # 1代表第一行，30是行高值
    ws.row_dimensions[36].height = 25  # 1代表第一行，30是行高值




    # 3. 定义边框样式
    # 双线样式（用于第4行、第36行）
    double_side = Side(border_style="double", color="000000")
    # 普通细边框（用于表格其他边框）
    thin_side = Side(border_style="thin", color="000000")

    # 4. 处理第一个表格区域：A4:I32
    ## 4.1 为第4行（表格标题行）设置双线边框
    for col in range(1, 12):  # A-I列
        cell = ws.cell(row=4, column=col)
        # 第4行单元格：上下左右都为双线
        cell.border = Border(
            top=double_side,
            bottom=thin_side,
            left=double_side if col == 1 else thin_side,  # 第一列左边框为双线
            right=double_side if col == 11 else thin_side  # 最后一列右边框为双线
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")  # 内容居中

    ## 4.2 为A5:I32设置普通细边框
    for row in range(5, 33):  # 5-32行
        for col in range(1, 12):  # A-I列
            cell = ws.cell(row=row, column=col)
            # 构建细边框（边缘列/行补充边框）
            left = double_side if col == 1 else thin_side
            right = double_side if col == 11 else thin_side
            bottom = double_side if row == 32 else thin_side
            cell.border = Border(
                top=thin_side,
                bottom=bottom,
                left=left,
                right=right
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # 5. 处理第二个表格区域：A36:I64
    ## 5.1 为第36行（表格标题行）设置双线边框
    for col in range(1, 12):  # A-I列
        cell = ws.cell(row=36, column=col)
        cell.border = Border(
            top=double_side,
            bottom=thin_side,
            left=double_side if col == 1 else thin_side,
            right=double_side if col == 11 else thin_side
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ## 5.2 为A37:I64设置普通细边框
    for row in range(37, 65):  # 37-64行
        for col in range(1, 12):  # A-I列
            cell = ws.cell(row=row, column=col)
            left = double_side if col == 1 else thin_side
            right = double_side if col == 11 else thin_side
            bottom = double_side if row == 64 else thin_side
            cell.border = Border(
                top=thin_side,
                bottom=bottom,
                left=left,
                right=right
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 男生表头（第4行）
        boy_headers = ['序号', '姓名', '性别', '50米\n(s)', '跳绳\n(个)', '坐位体前屈\n(cm)', '仰卧起坐\n(个)','立定跳远\n(cm)',
                       '肺活量\n(ml)', '身高\n(cm)', '体重\n(kg)']
        for col, header in enumerate(boy_headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(size=8, bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center',
                                       shrink_to_fit=False)  # # 必须开启这个属性，\n才会生效   # 明确关闭“收缩至适合”，避免冲突

        # 女生表头（第36行）
        for col, header in enumerate(boy_headers, 1):
            cell = ws.cell(row=36, column=col, value=header)
            cell.font = Font(size=8, bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center', shrink_to_fit=False)



    # 3. ✅ 核心：把 Sheet1 完整复制成 Sheet2（格式字体全部保留）
    ws2 = wb.copy_worksheet(ws)  # 复制整个表（内容+字体+颜色+边框）
    ws2.title = "打印模板"  # 改名

    # 删除最后 3 列：下面这句
    ws2.delete_cols(ws2.max_column - 2, 3)
    # 定义需要合并的【起始列】（D=4, F=6, H=8）
    merge_cols = [6, 8, 10]

    # 定义需要合并的【行范围】
    merge_row_ranges = [(4, 32), (36, 64)]

    # 批量执行合并
    for col in merge_cols:
        for start_row, end_row in merge_row_ranges:
            for row in range(start_row, end_row + 1):
                # 合并：当前列 + 右边一列（2个单元格）
                ws2.merge_cells(start_row=row, start_column=col,
                                end_row=row, end_column=col + 1)

    # 文件名是字符串（如"一年级1班"），无需转float，且仅操作合并区域左上角
    ws2['A4'] = '序号'
    ws2['A4'].font = Font(size=11, bold=True)
    ws2['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['A36'] = '序号'
    ws2['A36'].font = Font(size=11, bold=True)
    ws2['A36'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['B4'] = '姓名'
    ws2['B4'].font = Font(size=11, bold=True)
    ws2['B4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['B36'] = '姓名'
    ws2['B36'].font = Font(size=11, bold=True)
    ws2['B36'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['C4'] = '性别'
    ws2['C4'].font = Font(size=11, bold=True)
    ws2['C4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['C36'] = '性别'
    ws2['C36'].font = Font(size=11, bold=True)
    ws2['C36'].alignment = Alignment(horizontal='center', vertical='center')

    ws2['D4'] = '50米(s)'
    ws2['D4'].font = Font(size=11, bold=True)
    ws2['D4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['D36'] = '50米(s)'
    ws2['D36'].font = Font(size=11, bold=True)
    ws2['D36'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['E4'] = '跳绳(个)'
    ws2['E4'].font = Font(size=11, bold=True)
    ws2['E4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['E36'] = '跳绳(个)'
    ws2['E36'].font = Font(size=11, bold=True)
    ws2['E36'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['F4'] = '坐位体前屈(cm)'
    ws2['F4'].font = Font(size=11, bold=True)
    ws2['F4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['F36'] = '坐位体前屈(cm)'
    ws2['F36'].font = Font(size=11, bold=True)
    ws2['F36'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['H4'] = '仰卧起坐(个)'
    ws2['H4'].font = Font(size=11, bold=True)
    ws2['H4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['H36'] = '仰卧起坐(个)'
    ws2['H36'].font = Font(size=11, bold=True)
    ws2['H36'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['J4'] = '立定跳远(cm)'
    ws2['J4'].font = Font(size=11, bold=True)
    ws2['J4'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['J36'] = '立定跳远(cm)'
    ws2['J36'].font = Font(size=11, bold=True)
    ws2['J36'].alignment = Alignment(horizontal='center', vertical='center')

    # 4. 处理第一个表格区域：A4:I32
    ## 4.1 为第4行（表格标题行）设置双线边框
    for col in range(8, 12):  # A-I列
        cell = ws2.cell(row=4, column=col)
        # 第4行单元格：上下左右都为双线
        cell.border = Border(
            top=double_side,
            bottom=thin_side,
            left=double_side if col == 1 else thin_side,  # 第一列左边框为双线
            right=double_side if col == 11 else thin_side  # 最后一列右边框为双线
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")  # 内容居中

    ## 4.2 为A5:I32设置普通细边框
    for row in range(5, 33):  # 5-32行
        for col in range(8, 12):  # A-I列
            cell = ws2.cell(row=row, column=col)
            # 构建细边框（边缘列/行补充边框）
            left = double_side if col == 1 else thin_side
            right = double_side if col == 11 else thin_side
            bottom = double_side if row == 32 else thin_side
            cell.border = Border(
                top=thin_side,
                bottom=bottom,
                left=left,
                right=right
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 5. 处理第二个表格区域：A36:I64
    ## 5.1 为第36行（表格标题行）设置双线边框
    for col in range(8, 12):  # A-I列
        cell = ws2.cell(row=36, column=col)
        cell.border = Border(
            top=double_side,
            bottom=thin_side,
            left=double_side if col == 1 else thin_side,
            right=double_side if col == 11 else thin_side
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ## 5.2 为A37:I64设置普通细边框
    for row in range(37, 65):  # 37-64行
        for col in range(8, 12):  # A-I列
            cell = ws2.cell(row=row, column=col)
            left = double_side if col == 1 else thin_side
            right = double_side if col == 11 else thin_side
            bottom = double_side if row == 64 else thin_side
            cell.border = Border(
                top=thin_side,
                bottom=bottom,
                left=left,
                right=right
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")




    # ==============================================
    # 清空 D5 到 D32（包含合并单元格 DxEy）
    for row in range(5, 33):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'A{row}'] = None
        # 清空 D5 到 D32（包含合并单元格 DxEy）
    for row in range(5, 33):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'C{row}'] = None
    # 清空 D5 到 D32（包含合并单元格 DxEy）
    for row in range(5, 33):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'D{row}'] = None
        # 清空 D5 到 D32（包含合并单元格 DxEy）
    for row in range(5, 33):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'E{row}'] = None
    for row in range(5, 33):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'F{row}'] = None
    for row in range(5, 33):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'H{row}'] = None
    for row in range(5, 33):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'J{row}'] = None

        # 清空 D5 到 D32（包含合并单元格 DxEy）
    for row in range(37, 65):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'A{row}'] = None
        # 清空 D5 到 D32（包含合并单元格 DxEy）
    for row in range(37, 65):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'C{row}'] = None
    for row in range(37, 65):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'D{row}'] = None
        # 清空 D5 到 D32（包含合并单元格 DxEy）
    for row in range(37, 65):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'E{row}'] = None
    for row in range(37, 65):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'F{row}'] = None

    for row in range(37, 65):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'H{row}'] = None
    for row in range(37, 65):
        # 直接获取合并后的主单元格，安全赋值
        ws2[f'J{row}'] = None






    # 保存文件
    wb.save(save_path)
    wb.close()