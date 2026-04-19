# 导入Excel操作所需库
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment,Font,Border, Side,PatternFill,numbers



def yinianji_sanxiang_zhuzhuangtu(sh2):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """


    # 3. 表格数据
    data = [  # 获取             50米                跳绳              坐位体前屈      三项成绩 名称
        ['', sh2["A2"].value, sh2["A4"].value, sh2["A6"].value, sh2["A8"].value],
        [sh2["F1"].value, sh2["F2"].value, sh2["F4"].value, sh2["F6"].value, sh2["F8"].value],  # 优秀率
        [sh2["G1"].value, sh2["G2"].value, sh2["G4"].value, sh2["G6"].value, sh2["G8"].value],  # 良好率
        [sh2["H1"].value, sh2["H2"].value, sh2["H4"].value, sh2["H6"].value, sh2["H8"].value],  # 优良率
        [sh2["I1"].value, sh2["I2"].value, sh2["I4"].value, sh2["I6"].value, sh2["I8"].value],  # 及格率
        [sh2["J1"].value, sh2["J2"].value, sh2["J4"].value, sh2["J6"].value, sh2["J8"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh2.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh2['B20'] = '选择项目'
    sh2['B22'] = '50米'  # 默认显示
    sh2.merge_cells('B20:B21')
    sh2['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,总分三项\n（70分）"', allow_blank=True)
    sh2.add_data_validation(dv)
    dv.add('B22')
    sh2.merge_cells('B22:B25')
    sh2['B22'].font = Font(size=14, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh2['C20'] = '指标'
    # sh2['D20'] = '百分比'
    sh2['C21'] = '优秀率'
    sh2['C22'] = '良好率'
    sh2['C23'] = '优良率'
    sh2['C24'] = '及格率'
    sh2['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh2['D21'] = '=HLOOKUP($B$22,$F$20:$J$25,2,FALSE)'
    sh2['D22'] = '=HLOOKUP($B$22,$F$20:$J$25,3,FALSE)'
    sh2['D23'] = '=HLOOKUP($B$22,$F$20:$J$25,4,FALSE)'
    sh2['D24'] = '=HLOOKUP($B$22,$F$20:$J$25,5,FALSE)'
    sh2['D25'] = '=HLOOKUP($B$22,$F$20:$J$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J']:  # C、D、E、F 列
            sh2[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "三项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh2, min_col=3, min_row=21, max_row=25)
    values = Reference(sh2, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh2.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def yinianji_wuxiang_zhuzhuangtu(sh3):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈        肺活量               BMI            五项成绩 名称
        ['',              sh3["A2"].value, sh3["A4"].value, sh3["A6"].value, sh3["A8"].value, sh3["A10"].value,  sh3["A12"].value],
        [sh3["F1"].value, sh3["F2"].value, sh3["F4"].value, sh3["F6"].value, sh3["F8"].value, sh3["F10"].value,  sh3["F12"].value],  # 优秀率
        [sh3["G1"].value, sh3["G2"].value, sh3["G4"].value, sh3["G6"].value, sh3["G8"].value, sh3["G10"].value,  sh3["G12"].value],  # 良好率
        [sh3["H1"].value, sh3["H2"].value, sh3["H4"].value, sh3["H6"].value, sh3["H8"].value, sh3["H10"].value,  sh3["H12"].value],  # 优良率
        [sh3["I1"].value, sh3["I2"].value, sh3["I4"].value, sh3["I6"].value, sh3["I8"].value, sh3["I10"].value,  sh3["I12"].value],  # 及格率
        [sh3["J1"].value, sh3["J2"].value, sh3["J4"].value, sh3["J6"].value, sh3["J8"].value, sh3["J10"].value,  sh3["J12"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh3.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh3['B20'] = '选择项目'
    sh3['B22'] = '50米'  # 默认显示
    sh3.merge_cells('B20:B21')
    sh3['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,肺活量,BMI,总分五项\n（100分）"', allow_blank=True)
    sh3.add_data_validation(dv)
    dv.add('B22')
    sh3.merge_cells('B22:B25')
    sh3['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh3['C20'] = '指标'
    # sh3['D20'] = '百分比'
    sh3['C21'] = '优秀率'
    sh3['C22'] = '良好率'
    sh3['C23'] = '优良率'
    sh3['C24'] = '及格率'
    sh3['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh3['D21'] = '=HLOOKUP($B$22,$F$20:$L$25,2,FALSE)'
    sh3['D22'] = '=HLOOKUP($B$22,$F$20:$L$25,3,FALSE)'
    sh3['D23'] = '=HLOOKUP($B$22,$F$20:$L$25,4,FALSE)'
    sh3['D24'] = '=HLOOKUP($B$22,$F$20:$L$25,5,FALSE)'
    sh3['D25'] = '=HLOOKUP($B$22,$F$20:$L$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L']:  # C、D、E、F 列
            sh3[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "五项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh3, min_col=3, min_row=21, max_row=25)
    values = Reference(sh3, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh3.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")


def ernianji_sanxiang_zhuzhuangtu(sh2):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """


    # 3. 表格数据
    data = [  # 获取             50米                跳绳              坐位体前屈      三项成绩 名称
        ['', sh2["A2"].value, sh2["A4"].value, sh2["A6"].value, sh2["A8"].value],
        [sh2["F1"].value, sh2["F2"].value, sh2["F4"].value, sh2["F6"].value, sh2["F8"].value],  # 优秀率
        [sh2["G1"].value, sh2["G2"].value, sh2["G4"].value, sh2["G6"].value, sh2["G8"].value],  # 良好率
        [sh2["H1"].value, sh2["H2"].value, sh2["H4"].value, sh2["H6"].value, sh2["H8"].value],  # 优良率
        [sh2["I1"].value, sh2["I2"].value, sh2["I4"].value, sh2["I6"].value, sh2["I8"].value],  # 及格率
        [sh2["J1"].value, sh2["J2"].value, sh2["J4"].value, sh2["J6"].value, sh2["J8"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh2.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh2['B20'] = '选择项目'
    sh2['B22'] = '50米'  # 默认显示
    sh2.merge_cells('B20:B21')
    sh2['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,总分三项\n（65分）"', allow_blank=True)
    sh2.add_data_validation(dv)
    dv.add('B22')
    sh2.merge_cells('B22:B25')
    sh2['B22'].font = Font(size=14, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh2['C20'] = '指标'
    # sh2['D20'] = '百分比'
    sh2['C21'] = '优秀率'
    sh2['C22'] = '良好率'
    sh2['C23'] = '优良率'
    sh2['C24'] = '及格率'
    sh2['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh2['D21'] = '=HLOOKUP($B$22,$F$20:$J$25,2,FALSE)'
    sh2['D22'] = '=HLOOKUP($B$22,$F$20:$J$25,3,FALSE)'
    sh2['D23'] = '=HLOOKUP($B$22,$F$20:$J$25,4,FALSE)'
    sh2['D24'] = '=HLOOKUP($B$22,$F$20:$J$25,5,FALSE)'
    sh2['D25'] = '=HLOOKUP($B$22,$F$20:$J$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J']:  # C、D、E、F 列
            sh2[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "三项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh2, min_col=3, min_row=21, max_row=25)
    values = Reference(sh2, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh2.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def ernianji_sixiang_zhuzhuangtu(sh3):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈        仰卧起坐               四项成绩             名称
        ['',              sh3["A2"].value, sh3["A4"].value, sh3["A6"].value, sh3["A8"].value, sh3["A10"].value],
        [sh3["F1"].value, sh3["F2"].value, sh3["F4"].value, sh3["F6"].value, sh3["F8"].value, sh3["F10"].value],  # 优秀率
        [sh3["G1"].value, sh3["G2"].value, sh3["G4"].value, sh3["G6"].value, sh3["G8"].value, sh3["G10"].value],  # 良好率
        [sh3["H1"].value, sh3["H2"].value, sh3["H4"].value, sh3["H6"].value, sh3["H8"].value, sh3["H10"].value],  # 优良率
        [sh3["I1"].value, sh3["I2"].value, sh3["I4"].value, sh3["I6"].value, sh3["I8"].value, sh3["I10"].value],  # 及格率
        [sh3["J1"].value, sh3["J2"].value, sh3["J4"].value, sh3["J6"].value, sh3["J8"].value, sh3["J10"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh3.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh3['B20'] = '选择项目'
    sh3['B22'] = '50米'  # 默认显示
    sh3.merge_cells('B20:B21')
    sh3['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,总分四项\n（80分）"', allow_blank=True)
    sh3.add_data_validation(dv)
    dv.add('B22')
    sh3.merge_cells('B22:B25')
    sh3['B22'].font = Font(size=14, bold=True)


    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh3['C20'] = '指标'
    # sh3['D20'] = '百分比'
    sh3['C21'] = '优秀率'
    sh3['C22'] = '良好率'
    sh3['C23'] = '优良率'
    sh3['C24'] = '及格率'
    sh3['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh3['D21'] = '=HLOOKUP($B$22,$F$20:$L$25,2,FALSE)'
    sh3['D22'] = '=HLOOKUP($B$22,$F$20:$L$25,3,FALSE)'
    sh3['D23'] = '=HLOOKUP($B$22,$F$20:$L$25,4,FALSE)'
    sh3['D24'] = '=HLOOKUP($B$22,$F$20:$L$25,5,FALSE)'
    sh3['D25'] = '=HLOOKUP($B$22,$F$20:$L$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L']:  # C、D、E、F 列
            sh3[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "四项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh3, min_col=3, min_row=21, max_row=25)
    values = Reference(sh3, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh3.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def ernianji_liuxiang_zhuzhuangtu(sh4):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈       仰卧起坐              肺活量               BMI            六项成绩 名称
        ['',              sh4["A2"].value, sh4["A4"].value, sh4["A6"].value, sh4["A8"].value, sh4["A10"].value,  sh4["A12"].value,  sh4["A14"].value],
        [sh4["F1"].value, sh4["F2"].value, sh4["F4"].value, sh4["F6"].value, sh4["F8"].value, sh4["F10"].value,  sh4["F12"].value,  sh4["F14"].value],  # 优秀率
        [sh4["G1"].value, sh4["G2"].value, sh4["G4"].value, sh4["G6"].value, sh4["G8"].value, sh4["G10"].value,  sh4["G12"].value,  sh4["G14"].value],  # 良好率
        [sh4["H1"].value, sh4["H2"].value, sh4["H4"].value, sh4["H6"].value, sh4["H8"].value, sh4["H10"].value,  sh4["H12"].value,  sh4["H14"].value],  # 优良率
        [sh4["I1"].value, sh4["I2"].value, sh4["I4"].value, sh4["I6"].value, sh4["I8"].value, sh4["I10"].value,  sh4["I12"].value,  sh4["I14"].value],  # 及格率
        [sh4["J1"].value, sh4["J2"].value, sh4["J4"].value, sh4["J6"].value, sh4["J8"].value, sh4["J10"].value,  sh4["J12"].value,  sh4["J14"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh4.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh4['B20'] = '选择项目'
    sh4['B22'] = '50米'  # 默认显示
    sh4.merge_cells('B20:B21')
    sh4['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,肺活量,BMI,总分六项\n（100分）"', allow_blank=True)
    sh4.add_data_validation(dv)
    dv.add('B22')
    sh4.merge_cells('B22:B25')
    sh4['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh4['C20'] = '指标'
    # sh4['D20'] = '百分比'
    sh4['C21'] = '优秀率'
    sh4['C22'] = '良好率'
    sh4['C23'] = '优良率'
    sh4['C24'] = '及格率'
    sh4['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh4['D21'] = '=HLOOKUP($B$22,$F$20:$M$25,2,FALSE)'
    sh4['D22'] = '=HLOOKUP($B$22,$F$20:$M$25,3,FALSE)'
    sh4['D23'] = '=HLOOKUP($B$22,$F$20:$M$25,4,FALSE)'
    sh4['D24'] = '=HLOOKUP($B$22,$F$20:$M$25,5,FALSE)'
    sh4['D25'] = '=HLOOKUP($B$22,$F$20:$M$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:  # C、D、E、F 列
            sh4[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "六项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh4, min_col=3, min_row=21, max_row=25)
    values = Reference(sh4, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"
    #隐藏单元格内容
    sh4['M20'].number_format = ";;;"
    sh4['M21'].number_format = ";;;"
    sh4['M22'].number_format = ";;;"
    sh4['M23'].number_format = ";;;"
    sh4['M24'].number_format = ";;;"
    sh4['M25'].number_format = ";;;"


    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh4.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")


def sannianji_sixiang_zhuzhuangtu(sh2):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈        仰卧起坐               四项成绩             名称
        ['',              sh2["A2"].value, sh2["A4"].value, sh2["A6"].value, sh2["A8"].value, sh2["A10"].value],
        [sh2["F1"].value, sh2["F2"].value, sh2["F4"].value, sh2["F6"].value, sh2["F8"].value, sh2["F10"].value],  # 优秀率
        [sh2["G1"].value, sh2["G2"].value, sh2["G4"].value, sh2["G6"].value, sh2["G8"].value, sh2["G10"].value],  # 良好率
        [sh2["H1"].value, sh2["H2"].value, sh2["H4"].value, sh2["H6"].value, sh2["H8"].value, sh2["H10"].value],  # 优良率
        [sh2["I1"].value, sh2["I2"].value, sh2["I4"].value, sh2["I6"].value, sh2["I8"].value, sh2["I10"].value],  # 及格率
        [sh2["J1"].value, sh2["J2"].value, sh2["J4"].value, sh2["J6"].value, sh2["J8"].value, sh2["J10"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh2.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh2['B20'] = '选择项目'
    sh2['B22'] = '50米'  # 默认显示
    sh2.merge_cells('B20:B21')
    sh2['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,总分四项\n（65分）"', allow_blank=True)
    sh2.add_data_validation(dv)
    dv.add('B22')
    sh2.merge_cells('B22:B25')
    sh2['B22'].font = Font(size=14, bold=True)


    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh2['C20'] = '指标'
    # sh2['D20'] = '百分比'
    sh2['C21'] = '优秀率'
    sh2['C22'] = '良好率'
    sh2['C23'] = '优良率'
    sh2['C24'] = '及格率'
    sh2['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh2['D21'] = '=HLOOKUP($B$22,$F$20:$L$25,2,FALSE)'
    sh2['D22'] = '=HLOOKUP($B$22,$F$20:$L$25,3,FALSE)'
    sh2['D23'] = '=HLOOKUP($B$22,$F$20:$L$25,4,FALSE)'
    sh2['D24'] = '=HLOOKUP($B$22,$F$20:$L$25,5,FALSE)'
    sh2['D25'] = '=HLOOKUP($B$22,$F$20:$L$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L']:  # C、D、E、F 列
            sh2[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "四项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh2, min_col=3, min_row=21, max_row=25)
    values = Reference(sh2, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh2.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def sannianji_wuxiang_zhuzhuangtu(sh3):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈        仰卧起坐               立定跳远            五项成绩 名称
        ['',              sh3["A2"].value, sh3["A4"].value, sh3["A6"].value, sh3["A8"].value, sh3["A10"].value,  sh3["A12"].value],
        [sh3["F1"].value, sh3["F2"].value, sh3["F4"].value, sh3["F6"].value, sh3["F8"].value, sh3["F10"].value,  sh3["F12"].value],  # 优秀率
        [sh3["G1"].value, sh3["G2"].value, sh3["G4"].value, sh3["G6"].value, sh3["G8"].value, sh3["G10"].value,  sh3["G12"].value],  # 良好率
        [sh3["H1"].value, sh3["H2"].value, sh3["H4"].value, sh3["H6"].value, sh3["H8"].value, sh3["H10"].value,  sh3["H12"].value],  # 优良率
        [sh3["I1"].value, sh3["I2"].value, sh3["I4"].value, sh3["I6"].value, sh3["I8"].value, sh3["I10"].value,  sh3["I12"].value],  # 及格率
        [sh3["J1"].value, sh3["J2"].value, sh3["J4"].value, sh3["J6"].value, sh3["J8"].value, sh3["J10"].value,  sh3["J12"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh3.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh3['B20'] = '选择项目'
    sh3['B22'] = '50米'  # 默认显示
    sh3.merge_cells('B20:B21')
    sh3['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,立定跳远,总分五项\n（80分）"', allow_blank=True)
    sh3.add_data_validation(dv)
    dv.add('B22')
    sh3.merge_cells('B22:B25')
    sh3['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh3['C20'] = '指标'
    # sh3['D20'] = '百分比'
    sh3['C21'] = '优秀率'
    sh3['C22'] = '良好率'
    sh3['C23'] = '优良率'
    sh3['C24'] = '及格率'
    sh3['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh3['D21'] = '=HLOOKUP($B$22,$F$20:$L$25,2,FALSE)'
    sh3['D22'] = '=HLOOKUP($B$22,$F$20:$L$25,3,FALSE)'
    sh3['D23'] = '=HLOOKUP($B$22,$F$20:$L$25,4,FALSE)'
    sh3['D24'] = '=HLOOKUP($B$22,$F$20:$L$25,5,FALSE)'
    sh3['D25'] = '=HLOOKUP($B$22,$F$20:$L$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L']:  # C、D、E、F 列
            sh3[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "五项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh3, min_col=3, min_row=21, max_row=25)
    values = Reference(sh3, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh3.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def sannianji_qixiang_zhuzhuangtu(sh4):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈       仰卧起坐             立定跳远            肺活量               BMI            七项成绩 名称
        ['',              sh4["A2"].value, sh4["A4"].value, sh4["A6"].value, sh4["A8"].value, sh4["A10"].value,  sh4["A12"].value,  sh4["A14"].value,  sh4["A16"].value],
        [sh4["F1"].value, sh4["F2"].value, sh4["F4"].value, sh4["F6"].value, sh4["F8"].value, sh4["F10"].value,  sh4["F12"].value,  sh4["F14"].value,  sh4["F16"].value],  # 优秀率
        [sh4["G1"].value, sh4["G2"].value, sh4["G4"].value, sh4["G6"].value, sh4["G8"].value, sh4["G10"].value,  sh4["G12"].value,  sh4["G14"].value,  sh4["G16"].value],  # 良好率
        [sh4["H1"].value, sh4["H2"].value, sh4["H4"].value, sh4["H6"].value, sh4["H8"].value, sh4["H10"].value,  sh4["H12"].value,  sh4["H14"].value,  sh4["H16"].value],  # 优良率
        [sh4["I1"].value, sh4["I2"].value, sh4["I4"].value, sh4["I6"].value, sh4["I8"].value, sh4["I10"].value,  sh4["I12"].value,  sh4["I14"].value,  sh4["I16"].value],  # 及格率
        [sh4["J1"].value, sh4["J2"].value, sh4["J4"].value, sh4["J6"].value, sh4["J8"].value, sh4["J10"].value,  sh4["J12"].value,  sh4["J14"].value,  sh4["J16"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh4.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh4['B20'] = '选择项目'
    sh4['B22'] = '50米'  # 默认显示
    sh4.merge_cells('B20:B21')
    sh4['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,立定跳远,肺活量,BMI,总分七项\n（100分）"', allow_blank=True)
    sh4.add_data_validation(dv)
    dv.add('B22')
    sh4.merge_cells('B22:B25')
    sh4['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh4['C20'] = '指标'
    # sh4['D20'] = '百分比'
    sh4['C21'] = '优秀率'
    sh4['C22'] = '良好率'
    sh4['C23'] = '优良率'
    sh4['C24'] = '及格率'
    sh4['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh4['D21'] = '=HLOOKUP($B$22,$F$20:$N$25,2,FALSE)'
    sh4['D22'] = '=HLOOKUP($B$22,$F$20:$N$25,3,FALSE)'
    sh4['D23'] = '=HLOOKUP($B$22,$F$20:$N$25,4,FALSE)'
    sh4['D24'] = '=HLOOKUP($B$22,$F$20:$N$25,5,FALSE)'
    sh4['D25'] = '=HLOOKUP($B$22,$F$20:$N$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:  # C、D、E、F 列
            sh4[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "七项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh4, min_col=3, min_row=21, max_row=25)
    values = Reference(sh4, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"
    #隐藏单元格内容
    sh4['M20'].number_format = ";;;"
    sh4['M21'].number_format = ";;;"
    sh4['M22'].number_format = ";;;"
    sh4['M23'].number_format = ";;;"
    sh4['M24'].number_format = ";;;"
    sh4['M25'].number_format = ";;;"
    sh4['N20'].number_format = ";;;"
    sh4['N21'].number_format = ";;;"
    sh4['N22'].number_format = ";;;"
    sh4['N23'].number_format = ";;;"
    sh4['N24'].number_format = ";;;"
    sh4['N25'].number_format = ";;;"


    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh4.add_chart(chart, "E18")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")



def sinianji_sixiang_zhuzhuangtu(sh2):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈        仰卧起坐               四项成绩             名称
        ['',              sh2["A2"].value, sh2["A4"].value, sh2["A6"].value, sh2["A8"].value, sh2["A10"].value],
        [sh2["F1"].value, sh2["F2"].value, sh2["F4"].value, sh2["F6"].value, sh2["F8"].value, sh2["F10"].value],  # 优秀率
        [sh2["G1"].value, sh2["G2"].value, sh2["G4"].value, sh2["G6"].value, sh2["G8"].value, sh2["G10"].value],  # 良好率
        [sh2["H1"].value, sh2["H2"].value, sh2["H4"].value, sh2["H6"].value, sh2["H8"].value, sh2["H10"].value],  # 优良率
        [sh2["I1"].value, sh2["I2"].value, sh2["I4"].value, sh2["I6"].value, sh2["I8"].value, sh2["I10"].value],  # 及格率
        [sh2["J1"].value, sh2["J2"].value, sh2["J4"].value, sh2["J6"].value, sh2["J8"].value, sh2["J10"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh2.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh2['B20'] = '选择项目'
    sh2['B22'] = '50米'  # 默认显示
    sh2.merge_cells('B20:B21')
    sh2['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,总分四项\n（65分）"', allow_blank=True)
    sh2.add_data_validation(dv)
    dv.add('B22')
    sh2.merge_cells('B22:B25')
    sh2['B22'].font = Font(size=14, bold=True)


    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh2['C20'] = '指标'
    # sh2['D20'] = '百分比'
    sh2['C21'] = '优秀率'
    sh2['C22'] = '良好率'
    sh2['C23'] = '优良率'
    sh2['C24'] = '及格率'
    sh2['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh2['D21'] = '=HLOOKUP($B$22,$F$20:$L$25,2,FALSE)'
    sh2['D22'] = '=HLOOKUP($B$22,$F$20:$L$25,3,FALSE)'
    sh2['D23'] = '=HLOOKUP($B$22,$F$20:$L$25,4,FALSE)'
    sh2['D24'] = '=HLOOKUP($B$22,$F$20:$L$25,5,FALSE)'
    sh2['D25'] = '=HLOOKUP($B$22,$F$20:$L$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L']:  # C、D、E、F 列
            sh2[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "四项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh2, min_col=3, min_row=21, max_row=25)
    values = Reference(sh2, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh2.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def sinianji_wuxiang_zhuzhuangtu(sh3):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈        仰卧起坐               立定跳远            五项成绩 名称
        ['',              sh3["A2"].value, sh3["A4"].value, sh3["A6"].value, sh3["A8"].value, sh3["A10"].value,  sh3["A12"].value],
        [sh3["F1"].value, sh3["F2"].value, sh3["F4"].value, sh3["F6"].value, sh3["F8"].value, sh3["F10"].value,  sh3["F12"].value],  # 优秀率
        [sh3["G1"].value, sh3["G2"].value, sh3["G4"].value, sh3["G6"].value, sh3["G8"].value, sh3["G10"].value,  sh3["G12"].value],  # 良好率
        [sh3["H1"].value, sh3["H2"].value, sh3["H4"].value, sh3["H6"].value, sh3["H8"].value, sh3["H10"].value,  sh3["H12"].value],  # 优良率
        [sh3["I1"].value, sh3["I2"].value, sh3["I4"].value, sh3["I6"].value, sh3["I8"].value, sh3["I10"].value,  sh3["I12"].value],  # 及格率
        [sh3["J1"].value, sh3["J2"].value, sh3["J4"].value, sh3["J6"].value, sh3["J8"].value, sh3["J10"].value,  sh3["J12"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh3.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh3['B20'] = '选择项目'
    sh3['B22'] = '50米'  # 默认显示
    sh3.merge_cells('B20:B21')
    sh3['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,立定跳远,总分五项\n（80分）"', allow_blank=True)
    sh3.add_data_validation(dv)
    dv.add('B22')
    sh3.merge_cells('B22:B25')
    sh3['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh3['C20'] = '指标'
    # sh3['D20'] = '百分比'
    sh3['C21'] = '优秀率'
    sh3['C22'] = '良好率'
    sh3['C23'] = '优良率'
    sh3['C24'] = '及格率'
    sh3['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh3['D21'] = '=HLOOKUP($B$22,$F$20:$L$25,2,FALSE)'
    sh3['D22'] = '=HLOOKUP($B$22,$F$20:$L$25,3,FALSE)'
    sh3['D23'] = '=HLOOKUP($B$22,$F$20:$L$25,4,FALSE)'
    sh3['D24'] = '=HLOOKUP($B$22,$F$20:$L$25,5,FALSE)'
    sh3['D25'] = '=HLOOKUP($B$22,$F$20:$L$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L']:  # C、D、E、F 列
            sh3[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "五项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh3, min_col=3, min_row=21, max_row=25)
    values = Reference(sh3, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh3.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def sinianji_qixiang_zhuzhuangtu(sh4):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈       仰卧起坐             立定跳远            肺活量               BMI            七项成绩 名称
        ['',              sh4["A2"].value, sh4["A4"].value, sh4["A6"].value, sh4["A8"].value, sh4["A10"].value,  sh4["A12"].value,  sh4["A14"].value,  sh4["A16"].value],
        [sh4["F1"].value, sh4["F2"].value, sh4["F4"].value, sh4["F6"].value, sh4["F8"].value, sh4["F10"].value,  sh4["F12"].value,  sh4["F14"].value,  sh4["F16"].value],  # 优秀率
        [sh4["G1"].value, sh4["G2"].value, sh4["G4"].value, sh4["G6"].value, sh4["G8"].value, sh4["G10"].value,  sh4["G12"].value,  sh4["G14"].value,  sh4["G16"].value],  # 良好率
        [sh4["H1"].value, sh4["H2"].value, sh4["H4"].value, sh4["H6"].value, sh4["H8"].value, sh4["H10"].value,  sh4["H12"].value,  sh4["H14"].value,  sh4["H16"].value],  # 优良率
        [sh4["I1"].value, sh4["I2"].value, sh4["I4"].value, sh4["I6"].value, sh4["I8"].value, sh4["I10"].value,  sh4["I12"].value,  sh4["I14"].value,  sh4["I16"].value],  # 及格率
        [sh4["J1"].value, sh4["J2"].value, sh4["J4"].value, sh4["J6"].value, sh4["J8"].value, sh4["J10"].value,  sh4["J12"].value,  sh4["J14"].value,  sh4["J16"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh4.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh4['B20'] = '选择项目'
    sh4['B22'] = '50米'  # 默认显示
    sh4.merge_cells('B20:B21')
    sh4['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,立定跳远,肺活量,BMI,总分七项\n（100分）"', allow_blank=True)
    sh4.add_data_validation(dv)
    dv.add('B22')
    sh4.merge_cells('B22:B25')
    sh4['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh4['C20'] = '指标'
    # sh4['D20'] = '百分比'
    sh4['C21'] = '优秀率'
    sh4['C22'] = '良好率'
    sh4['C23'] = '优良率'
    sh4['C24'] = '及格率'
    sh4['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh4['D21'] = '=HLOOKUP($B$22,$F$20:$N$25,2,FALSE)'
    sh4['D22'] = '=HLOOKUP($B$22,$F$20:$N$25,3,FALSE)'
    sh4['D23'] = '=HLOOKUP($B$22,$F$20:$N$25,4,FALSE)'
    sh4['D24'] = '=HLOOKUP($B$22,$F$20:$N$25,5,FALSE)'
    sh4['D25'] = '=HLOOKUP($B$22,$F$20:$N$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:  # C、D、E、F 列
            sh4[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "七项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh4, min_col=3, min_row=21, max_row=25)
    values = Reference(sh4, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"
    #隐藏单元格内容
    sh4['M20'].number_format = ";;;"
    sh4['M21'].number_format = ";;;"
    sh4['M22'].number_format = ";;;"
    sh4['M23'].number_format = ";;;"
    sh4['M24'].number_format = ";;;"
    sh4['M25'].number_format = ";;;"
    sh4['N20'].number_format = ";;;"
    sh4['N21'].number_format = ";;;"
    sh4['N22'].number_format = ";;;"
    sh4['N23'].number_format = ";;;"
    sh4['N24'].number_format = ";;;"
    sh4['N25'].number_format = ";;;"


    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh4.add_chart(chart, "E18")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")


def wunianji_wuxiang_zhuzhuangtu(sh2):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈        仰卧起坐               50*8            五项成绩 名称
        ['',              sh2["A2"].value, sh2["A4"].value, sh2["A6"].value, sh2["A8"].value, sh2["A10"].value,  sh2["A12"].value],
        [sh2["F1"].value, sh2["F2"].value, sh2["F4"].value, sh2["F6"].value, sh2["F8"].value, sh2["F10"].value,  sh2["F12"].value],  # 优秀率
        [sh2["G1"].value, sh2["G2"].value, sh2["G4"].value, sh2["G6"].value, sh2["G8"].value, sh2["G10"].value,  sh2["G12"].value],  # 良好率
        [sh2["H1"].value, sh2["H2"].value, sh2["H4"].value, sh2["H6"].value, sh2["H8"].value, sh2["H10"].value,  sh2["H12"].value],  # 优良率
        [sh2["I1"].value, sh2["I2"].value, sh2["I4"].value, sh2["I6"].value, sh2["I8"].value, sh2["I10"].value,  sh2["I12"].value],  # 及格率
        [sh2["J1"].value, sh2["J2"].value, sh2["J4"].value, sh2["J6"].value, sh2["J8"].value, sh2["J10"].value,  sh2["J12"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh2.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh2['B20'] = '选择项目'
    sh2['B22'] = '50米'  # 默认显示
    sh2.merge_cells('B20:B21')
    sh2['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,50*8,总分五项\n（70分）"', allow_blank=True)
    sh2.add_data_validation(dv)
    dv.add('B22')
    sh2.merge_cells('B22:B25')
    sh2['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh2['C20'] = '指标'
    # sh2['D20'] = '百分比'
    sh2['C21'] = '优秀率'
    sh2['C22'] = '良好率'
    sh2['C23'] = '优良率'
    sh2['C24'] = '及格率'
    sh2['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh2['D21'] = '=HLOOKUP($B$22,$F$20:$L$25,2,FALSE)'
    sh2['D22'] = '=HLOOKUP($B$22,$F$20:$L$25,3,FALSE)'
    sh2['D23'] = '=HLOOKUP($B$22,$F$20:$L$25,4,FALSE)'
    sh2['D24'] = '=HLOOKUP($B$22,$F$20:$L$25,5,FALSE)'
    sh2['D25'] = '=HLOOKUP($B$22,$F$20:$L$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L']:  # C、D、E、F 列
            sh2[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "五项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh2, min_col=3, min_row=21, max_row=25)
    values = Reference(sh2, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh2.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def wunianji_liuxiang_zhuzhuangtu(sh3):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈       仰卧起坐             50*8            直臂悬垂                   六项成绩 名称
        ['',              sh3["A2"].value, sh3["A4"].value, sh3["A6"].value, sh3["A8"].value, sh3["A10"].value,  sh3["A12"].value,  sh3["A14"].value],
        [sh3["F1"].value, sh3["F2"].value, sh3["F4"].value, sh3["F6"].value, sh3["F8"].value, sh3["F10"].value,  sh3["F12"].value,  sh3["F14"].value],  # 优秀率
        [sh3["G1"].value, sh3["G2"].value, sh3["G4"].value, sh3["G6"].value, sh3["G8"].value, sh3["G10"].value,  sh3["G12"].value,  sh3["G14"].value],  # 良好率
        [sh3["H1"].value, sh3["H2"].value, sh3["H4"].value, sh3["H6"].value, sh3["H8"].value, sh3["H10"].value,  sh3["H12"].value,  sh3["H14"].value],  # 优良率
        [sh3["I1"].value, sh3["I2"].value, sh3["I4"].value, sh3["I6"].value, sh3["I8"].value, sh3["I10"].value,  sh3["I12"].value,  sh3["I14"].value],  # 及格率
        [sh3["J1"].value, sh3["J2"].value, sh3["J4"].value, sh3["J6"].value, sh3["J8"].value, sh3["J10"].value,  sh3["J12"].value,  sh3["J14"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh3.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh3['B20'] = '选择项目'
    sh3['B22'] = '50米'  # 默认显示
    sh3.merge_cells('B20:B21')
    sh3['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,50*8,直臂悬垂,总分六项\n（70分）"', allow_blank=True)
    sh3.add_data_validation(dv)
    dv.add('B22')
    sh3.merge_cells('B22:B25')
    sh3['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh3['C20'] = '指标'
    # sh3['D20'] = '百分比'
    sh3['C21'] = '优秀率'
    sh3['C22'] = '良好率'
    sh3['C23'] = '优良率'
    sh3['C24'] = '及格率'
    sh3['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh3['D21'] = '=HLOOKUP($B$22,$F$20:$N$25,2,FALSE)'
    sh3['D22'] = '=HLOOKUP($B$22,$F$20:$N$25,3,FALSE)'
    sh3['D23'] = '=HLOOKUP($B$22,$F$20:$N$25,4,FALSE)'
    sh3['D24'] = '=HLOOKUP($B$22,$F$20:$N$25,5,FALSE)'
    sh3['D25'] = '=HLOOKUP($B$22,$F$20:$N$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:  # C、D、E、F 列
            sh3[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "六项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh3, min_col=3, min_row=21, max_row=25)
    values = Reference(sh3, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"
    #隐藏单元格内容
    sh3['M20'].number_format = ";;;"
    sh3['M21'].number_format = ";;;"
    sh3['M22'].number_format = ";;;"
    sh3['M23'].number_format = ";;;"
    sh3['M24'].number_format = ";;;"
    sh3['M25'].number_format = ";;;"
    sh3['N20'].number_format = ";;;"
    sh3['N21'].number_format = ";;;"
    sh3['N22'].number_format = ";;;"
    sh3['N23'].number_format = ";;;"
    sh3['N24'].number_format = ";;;"
    sh3['N25'].number_format = ";;;"


    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh3.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def wunianji_baxiang_zhuzhuangtu(sh4):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈       仰卧起坐             50*8             直臂悬垂            肺活量               BMI            八项成绩 名称
        ['',              sh4["A2"].value, sh4["A4"].value, sh4["A6"].value, sh4["A8"].value, sh4["A10"].value,  sh4["A12"].value,  sh4["A14"].value,  sh4["A16"].value,  sh4["A18"].value],
        [sh4["F1"].value, sh4["F2"].value, sh4["F4"].value, sh4["F6"].value, sh4["F8"].value, sh4["F10"].value,  sh4["F12"].value,  sh4["F14"].value,  sh4["F16"].value,  sh4["F18"].value],  # 优秀率
        [sh4["G1"].value, sh4["G2"].value, sh4["G4"].value, sh4["G6"].value, sh4["G8"].value, sh4["G10"].value,  sh4["G12"].value,  sh4["G14"].value,  sh4["G16"].value,  sh4["G18"].value],  # 良好率
        [sh4["H1"].value, sh4["H2"].value, sh4["H4"].value, sh4["H6"].value, sh4["H8"].value, sh4["H10"].value,  sh4["H12"].value,  sh4["H14"].value,  sh4["H16"].value,  sh4["H18"].value],  # 优良率
        [sh4["I1"].value, sh4["I2"].value, sh4["I4"].value, sh4["I6"].value, sh4["I8"].value, sh4["I10"].value,  sh4["I12"].value,  sh4["I14"].value,  sh4["I16"].value,  sh4["I18"].value],  # 及格率
        [sh4["J1"].value, sh4["J2"].value, sh4["J4"].value, sh4["J6"].value, sh4["J8"].value, sh4["J10"].value,  sh4["J12"].value,  sh4["J14"].value,  sh4["J16"].value,  sh4["J18"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh4.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh4['B20'] = '选择项目'
    sh4['B22'] = '50米'  # 默认显示
    sh4.merge_cells('B20:B21')
    sh4['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,50*8,直臂悬垂,肺活量,BMI,总分八项\n（100分）"', allow_blank=True)
    sh4.add_data_validation(dv)
    dv.add('B22')
    sh4.merge_cells('B22:B25')
    sh4['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh4['C20'] = '指标'
    # sh4['D20'] = '百分比'
    sh4['C21'] = '优秀率'
    sh4['C22'] = '良好率'
    sh4['C23'] = '优良率'
    sh4['C24'] = '及格率'
    sh4['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh4['D21'] = '=HLOOKUP($B$22,$F$20:$O$25,2,FALSE)'
    sh4['D22'] = '=HLOOKUP($B$22,$F$20:$O$25,3,FALSE)'
    sh4['D23'] = '=HLOOKUP($B$22,$F$20:$O$25,4,FALSE)'
    sh4['D24'] = '=HLOOKUP($B$22,$F$20:$O$25,5,FALSE)'
    sh4['D25'] = '=HLOOKUP($B$22,$F$20:$O$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:  # C、D、E、F 列
            sh4[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "八项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh4, min_col=3, min_row=21, max_row=25)
    values = Reference(sh4, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"
    #隐藏单元格内容
    sh4['M20'].number_format = ";;;"
    sh4['M21'].number_format = ";;;"
    sh4['M22'].number_format = ";;;"
    sh4['M23'].number_format = ";;;"
    sh4['M24'].number_format = ";;;"
    sh4['M25'].number_format = ";;;"
    sh4['N20'].number_format = ";;;"
    sh4['N21'].number_format = ";;;"
    sh4['N22'].number_format = ";;;"
    sh4['N23'].number_format = ";;;"
    sh4['N24'].number_format = ";;;"
    sh4['N25'].number_format = ";;;"
    sh4['O20'].number_format = ";;;"
    sh4['O21'].number_format = ";;;"
    sh4['O22'].number_format = ";;;"
    sh4['O23'].number_format = ";;;"
    sh4['O24'].number_format = ";;;"
    sh4['O25'].number_format = ";;;"


    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh4.add_chart(chart, "E20")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")



def liunianji_wuxiang_zhuzhuangtu(sh2):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈        仰卧起坐               50*8            五项成绩 名称
        ['',              sh2["A2"].value, sh2["A4"].value, sh2["A6"].value, sh2["A8"].value, sh2["A10"].value,  sh2["A12"].value],
        [sh2["F1"].value, sh2["F2"].value, sh2["F4"].value, sh2["F6"].value, sh2["F8"].value, sh2["F10"].value,  sh2["F12"].value],  # 优秀率
        [sh2["G1"].value, sh2["G2"].value, sh2["G4"].value, sh2["G6"].value, sh2["G8"].value, sh2["G10"].value,  sh2["G12"].value],  # 良好率
        [sh2["H1"].value, sh2["H2"].value, sh2["H4"].value, sh2["H6"].value, sh2["H8"].value, sh2["H10"].value,  sh2["H12"].value],  # 优良率
        [sh2["I1"].value, sh2["I2"].value, sh2["I4"].value, sh2["I6"].value, sh2["I8"].value, sh2["I10"].value,  sh2["I12"].value],  # 及格率
        [sh2["J1"].value, sh2["J2"].value, sh2["J4"].value, sh2["J6"].value, sh2["J8"].value, sh2["J10"].value,  sh2["J12"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh2.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh2['B20'] = '选择项目'
    sh2['B22'] = '50米'  # 默认显示
    sh2.merge_cells('B20:B21')
    sh2['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,50*8,总分五项\n（70分）"', allow_blank=True)
    sh2.add_data_validation(dv)
    dv.add('B22')
    sh2.merge_cells('B22:B25')
    sh2['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh2['C20'] = '指标'
    # sh2['D20'] = '百分比'
    sh2['C21'] = '优秀率'
    sh2['C22'] = '良好率'
    sh2['C23'] = '优良率'
    sh2['C24'] = '及格率'
    sh2['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh2['D21'] = '=HLOOKUP($B$22,$F$20:$L$25,2,FALSE)'
    sh2['D22'] = '=HLOOKUP($B$22,$F$20:$L$25,3,FALSE)'
    sh2['D23'] = '=HLOOKUP($B$22,$F$20:$L$25,4,FALSE)'
    sh2['D24'] = '=HLOOKUP($B$22,$F$20:$L$25,5,FALSE)'
    sh2['D25'] = '=HLOOKUP($B$22,$F$20:$L$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L']:  # C、D、E、F 列
            sh2[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "五项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh2, min_col=3, min_row=21, max_row=25)
    values = Reference(sh2, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"

    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh2.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def liunianji_liuxiang_zhuzhuangtu(sh3):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈       仰卧起坐             50*8            直臂悬垂                   六项成绩 名称
        ['',              sh3["A2"].value, sh3["A4"].value, sh3["A6"].value, sh3["A8"].value, sh3["A10"].value,  sh3["A12"].value,  sh3["A14"].value],
        [sh3["F1"].value, sh3["F2"].value, sh3["F4"].value, sh3["F6"].value, sh3["F8"].value, sh3["F10"].value,  sh3["F12"].value,  sh3["F14"].value],  # 优秀率
        [sh3["G1"].value, sh3["G2"].value, sh3["G4"].value, sh3["G6"].value, sh3["G8"].value, sh3["G10"].value,  sh3["G12"].value,  sh3["G14"].value],  # 良好率
        [sh3["H1"].value, sh3["H2"].value, sh3["H4"].value, sh3["H6"].value, sh3["H8"].value, sh3["H10"].value,  sh3["H12"].value,  sh3["H14"].value],  # 优良率
        [sh3["I1"].value, sh3["I2"].value, sh3["I4"].value, sh3["I6"].value, sh3["I8"].value, sh3["I10"].value,  sh3["I12"].value,  sh3["I14"].value],  # 及格率
        [sh3["J1"].value, sh3["J2"].value, sh3["J4"].value, sh3["J6"].value, sh3["J8"].value, sh3["J10"].value,  sh3["J12"].value,  sh3["J14"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh3.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh3['B20'] = '选择项目'
    sh3['B22'] = '50米'  # 默认显示
    sh3.merge_cells('B20:B21')
    sh3['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,50*8,直臂悬垂,总分六项\n（70分）"', allow_blank=True)
    sh3.add_data_validation(dv)
    dv.add('B22')
    sh3.merge_cells('B22:B25')
    sh3['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh3['C20'] = '指标'
    # sh3['D20'] = '百分比'
    sh3['C21'] = '优秀率'
    sh3['C22'] = '良好率'
    sh3['C23'] = '优良率'
    sh3['C24'] = '及格率'
    sh3['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh3['D21'] = '=HLOOKUP($B$22,$F$20:$N$25,2,FALSE)'
    sh3['D22'] = '=HLOOKUP($B$22,$F$20:$N$25,3,FALSE)'
    sh3['D23'] = '=HLOOKUP($B$22,$F$20:$N$25,4,FALSE)'
    sh3['D24'] = '=HLOOKUP($B$22,$F$20:$N$25,5,FALSE)'
    sh3['D25'] = '=HLOOKUP($B$22,$F$20:$N$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:  # C、D、E、F 列
            sh3[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "六项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh3, min_col=3, min_row=21, max_row=25)
    values = Reference(sh3, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"
    #隐藏单元格内容
    sh3['M20'].number_format = ";;;"
    sh3['M21'].number_format = ";;;"
    sh3['M22'].number_format = ";;;"
    sh3['M23'].number_format = ";;;"
    sh3['M24'].number_format = ";;;"
    sh3['M25'].number_format = ";;;"
    sh3['N20'].number_format = ";;;"
    sh3['N21'].number_format = ";;;"
    sh3['N22'].number_format = ";;;"
    sh3['N23'].number_format = ";;;"
    sh3['N24'].number_format = ";;;"
    sh3['N25'].number_format = ";;;"


    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh3.add_chart(chart, "E17")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")
def liunianji_baxiang_zhuzhuangtu(sh4):
    """
    功能：下拉菜单切换项目 → 显示该项目 优秀/良好/优良/及格/不及格 全部指标
    无FormControl、无宏、无报错、所有openpyxl版本通用
    """

    # 3. 表格数据
    data = [  # 获取             50米            跳绳            坐位体前屈       仰卧起坐             50*8             直臂悬垂            肺活量               BMI            八项成绩 名称
        ['',              sh4["A2"].value, sh4["A4"].value, sh4["A6"].value, sh4["A8"].value, sh4["A10"].value,  sh4["A12"].value,  sh4["A14"].value,  sh4["A16"].value,  sh4["A18"].value],
        [sh4["F1"].value, sh4["F2"].value, sh4["F4"].value, sh4["F6"].value, sh4["F8"].value, sh4["F10"].value,  sh4["F12"].value,  sh4["F14"].value,  sh4["F16"].value,  sh4["F18"].value],  # 优秀率
        [sh4["G1"].value, sh4["G2"].value, sh4["G4"].value, sh4["G6"].value, sh4["G8"].value, sh4["G10"].value,  sh4["G12"].value,  sh4["G14"].value,  sh4["G16"].value,  sh4["G18"].value],  # 良好率
        [sh4["H1"].value, sh4["H2"].value, sh4["H4"].value, sh4["H6"].value, sh4["H8"].value, sh4["H10"].value,  sh4["H12"].value,  sh4["H14"].value,  sh4["H16"].value,  sh4["H18"].value],  # 优良率
        [sh4["I1"].value, sh4["I2"].value, sh4["I4"].value, sh4["I6"].value, sh4["I8"].value, sh4["I10"].value,  sh4["I12"].value,  sh4["I14"].value,  sh4["I16"].value,  sh4["I18"].value],  # 及格率
        [sh4["J1"].value, sh4["J2"].value, sh4["J4"].value, sh4["J6"].value, sh4["J8"].value, sh4["J10"].value,  sh4["J12"].value,  sh4["J14"].value,  sh4["J16"].value,  sh4["J18"].value]  # 不及格率
    ]


    # 3. 从 B20 开始逐行写入
    start_row = 20  # 从第20行开始
    start_col = 6  # 从B列开始（A=1, B=2, C=3...）

    for row_idx, row in enumerate(data):
        for col_idx, value in enumerate(row):
            # 定位单元格：行=20+偏移，列=B+偏移
            sh4.cell(
                row=start_row + row_idx,
                column=start_col + col_idx,
                value=value
            )


    # ======================
    # 下拉选择框（切换项目）
    # ======================
    sh4['B20'] = '选择项目'
    sh4['B22'] = '50米'  # 默认显示
    sh4.merge_cells('B20:B21')
    sh4['B20'].font = Font(size=14, bold=True)

    # 创建下拉菜单
    dv = DataValidation(type="list", formula1='"50米,跳绳,坐位体前屈,仰卧起坐,50*8,直臂悬垂,肺活量,BMI,总分八项\n（100分）"', allow_blank=True)
    sh4.add_data_validation(dv)
    dv.add('B22')
    sh4.merge_cells('B22:B25')
    sh4['B22'].font = Font(size=13, bold=True)

    # ======================
    # 动态图表数据区（自动更新）
    # ======================
    sh4['C20'] = '指标'
    # sh4['D20'] = '百分比'
    sh4['C21'] = '优秀率'
    sh4['C22'] = '良好率'
    sh4['C23'] = '优良率'
    sh4['C24'] = '及格率'
    sh4['C25'] = '不及格率'

    # 公式：根据下拉选择自动匹配数据
    sh4['D21'] = '=HLOOKUP($B$22,$F$20:$O$25,2,FALSE)'
    sh4['D22'] = '=HLOOKUP($B$22,$F$20:$O$25,3,FALSE)'
    sh4['D23'] = '=HLOOKUP($B$22,$F$20:$O$25,4,FALSE)'
    sh4['D24'] = '=HLOOKUP($B$22,$F$20:$O$25,5,FALSE)'
    sh4['D25'] = '=HLOOKUP($B$22,$F$20:$O$25,6,FALSE)'



    # 5. C21到F25区域设置为百分比格式
    for row in range(21, 26):  # 21 到 25 行
        for col in ['D','F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:  # C、D、E、F 列
            sh4[f'{col}{row}'].number_format = "0.00%"

    # ======================
    # 柱状图（只显示选中的单个项目）
    # ======================
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "八项测试成绩（项目可切换）"
    chart.y_axis.title = "百分率"
    chart.x_axis.title = "各项目成绩等级"

    # 纵轴设置
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    chart.y_axis.majorUnit = 0.2
    chart.y_axis.number_format = "0%"

    # 数据来源：动态区域
    categories = Reference(sh4, min_col=3, min_row=21, max_row=25)
    values = Reference(sh4, min_col=4, min_row=20, max_row=25)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)

    # 数据标签
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.number_format = "0%"
    #隐藏单元格内容
    sh4['M20'].number_format = ";;;"
    sh4['M21'].number_format = ";;;"
    sh4['M22'].number_format = ";;;"
    sh4['M23'].number_format = ";;;"
    sh4['M24'].number_format = ";;;"
    sh4['M25'].number_format = ";;;"
    sh4['N20'].number_format = ";;;"
    sh4['N21'].number_format = ";;;"
    sh4['N22'].number_format = ";;;"
    sh4['N23'].number_format = ";;;"
    sh4['N24'].number_format = ";;;"
    sh4['N25'].number_format = ";;;"
    sh4['O20'].number_format = ";;;"
    sh4['O21'].number_format = ";;;"
    sh4['O22'].number_format = ";;;"
    sh4['O23'].number_format = ";;;"
    sh4['O24'].number_format = ";;;"
    sh4['O25'].number_format = ";;;"


    # 图表大小位置
    chart.width = 18
    chart.height = 9
    sh4.add_chart(chart, "E20")

    # 保存
    # wb.save("✅单项目切换图表.xlsx")
    # print("生成完成！打开Excel，点击 G2 单元格下拉切换项目！")

# =======================
# 调用生成
# =======================
# if __name__ == '__main__':
    # yinianji_sanxiang_zhuzhuangtu()