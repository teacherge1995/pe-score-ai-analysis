import openpyxl
from openpyxl.styles import Alignment,Font,Border, Side,PatternFill,numbers

#一二年级 三项  50米 跳绳 坐位体前屈
def san_xunlianfangan_plan(sh2):
    """
    读取一年级体育成绩单Excel，自动生成详细训练方案并写入B36单元格

    参数：
        file_path: Excel文件路径（例："1年级体育成绩单_1.1.xlsx"）

    返回：
        成功返回训练方案文本，失败返回错误提示
    """
    try:


        # 2. 读取项目名称
        item_50m = sh2["G20"].value
        item_tiaosheng = sh2["H20"].value
        item_zuowei = sh2["I20"].value

        # 3. 读取各项比率（空值默认为0）
        # 50米
        excel_50m = sh2["G21"].value or 0
        good_50m = sh2["G22"].value or 0
        ex_good_50m = sh2["G23"].value or 0
        pass_50m = sh2["G24"].value or 0
        fail_50m = sh2["G25"].value or 0

        # 跳绳
        excel_tiaosheng = sh2["H21"].value or 0
        good_tiaosheng = sh2["H22"].value or 0
        ex_good_tiaosheng = sh2["H23"].value or 0
        pass_tiaosheng = sh2["H24"].value or 0
        fail_tiaosheng = sh2["H25"].value or 0

        # 坐位体前屈
        excel_zuowei = sh2["I21"].value or 0
        good_zuowei = sh2["I22"].value or 0
        ex_good_zuowei = sh2["I23"].value or 0
        pass_zuowei = sh2["I24"].value or 0
        fail_zuowei = sh2["I25"].value or 0

        # 肺活量
        excel_feihuol = sh2["J21"].value or 0
        good_feihuol = sh2["J22"].value or 0
        ex_good_feihuol = sh2["J23"].value or 0
        pass_feihuol = sh2["J24"].value or 0
        fail_feihuol = sh2["J25"].value or 0
        # Bmi
        excel_BMI = sh2["K21"].value or 0
        good_BMI = sh2["K22"].value or 0
        ex_good_BMI = sh2["K23"].value or 0
        pass_BMI = sh2["K24"].value or 0
        fail_BMI = sh2["K25"].value or 0



        # 4. 内部方法：根据项目生成具体训练方案
        def get_training_plan(item_name, fail_rate, ex_good_rate):
            fail = fail_rate * 100
            ex_good = ex_good_rate * 100

            if "50米" in item_name or "50m" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础起跑练习（原地摆臂、起跑反应练习），30米加速跑，趣味追逐跑，每周3次，降低强度提升兴趣，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点辅导起跑与反应速度，增加30米、40米间歇跑，每组2次，组间休息30秒，提升速度耐力。"
                elif ex_good >= 60:
                    return f"{item_name}：保持现有训练，增加行进间高抬腿、后蹬跑，提升爆发力与步频，培养尖子生冲刺能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化途中跑节奏，加入反应起跑练习，每组3次50米跑，巩固良好水平，提升优良率。"
                else:
                    return f"{item_name}：加强摆臂与蹬地练习，开展小组竞赛，激发积极性，提升整体优秀与良好人数。"

            elif "跳绳" in item_name:
                if fail > 20:
                    return f"{item_name}：从无绳徒手跳、单脚跳、双脚跳基础练起，每次500个，循序渐进让学生掌握基本节奏，降低不及格率。"
                elif fail > 10:
                    return f"{item_name}：分组练习，慢速连续跳+计时跳，重点纠正绊绳问题，每天5分钟专项训练，提升稳定性。"
                elif ex_good >= 60:
                    return f"{item_name}：保持快速跳训练，增加双摇尝试，并脚跳进阶动作，提升脚踝力量与熟练度。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力跳训练，30秒计时跳，提升连续跳绳能力，巩固优良水平。"
                else:
                    return f"{item_name}：增加竞赛跳绳、花样跳，提升兴趣与熟练度，扩大优秀、良好学生群体。"

            elif "坐位体前屈" in item_name:
                if fail > 20:
                    return f"{item_name}：每日课前静态拉伸（大腿后侧、腰部），坐姿体前屈保持15秒/组，做3组，提升身体柔韧度。"
                elif fail > 10:
                    return f"{item_name}：加强动态拉伸+伙伴辅助拉伸，循序渐进增加伸展幅度，帮助学生轻松达标。"
                elif ex_good >= 60:
                    return f"{item_name}：保持日常拉伸习惯，增加进阶柔韧练习，保持优秀柔韧水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日柔韧训练，巩固成绩，稳步提升优良率。"
                else:
                    return f"{item_name}：增加拉伸频率，课课练、天天练，提升整体柔韧素质。"
            else:
                return f"{item_name}：坚持常规训练，分层指导，全面提升成绩。"

        # 5. 生成三个项目方案
        plan1 = get_training_plan(item_50m, fail_50m, ex_good_50m)
        plan2 = get_training_plan(item_tiaosheng, fail_tiaosheng, ex_good_tiaosheng)
        plan3 = get_training_plan(item_zuowei, fail_zuowei, ex_good_zuowei)

        # 6. 组合最终方案
        final_plan = f"""建议训练方案：
        1. {plan1}
        2. {plan2}
        3. {plan3}
        整体坚持课课练、天天练，采用趣味化、分层化教学，全面提升学生身体素质。"""

        # # 7. 写入B36并保存
        sh2["B36"] = final_plan
        sh2.merge_cells('B36:M43')
        sh2['B36'].font = Font(size=12, bold=True)
        sh2['B36'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')  # 水平+垂直居中



    except Exception as e:
        return f"执行失败：{str(e)}"


#一年级  五项  50米 跳绳 坐位体前屈 肺活量 BMI
def di_wu_xunlianfangan_plan(sh2):
    """
    读取一年级体育成绩单Excel，自动生成详细训练方案并写入B36单元格

    参数：
        sh2: Excel工作表对象

    返回：
        成功返回训练方案文本，失败返回错误提示
    """
    try:
        # 2. 读取项目名称
        item_50m = sh2["G20"].value
        item_tiaosheng = sh2["H20"].value
        item_zuowei = sh2["I20"].value
        item_feihuoli = sh2["J20"].value  # 肺活量项目名
        item_bmi = sh2["K20"].value  # BMI项目名

        # 3. 读取各项比率（空值默认为0）
        # 50米
        excel_50m = sh2["G21"].value or 0
        good_50m = sh2["G22"].value or 0
        ex_good_50m = sh2["G23"].value or 0
        pass_50m = sh2["G24"].value or 0
        fail_50m = sh2["G25"].value or 0

        # 跳绳
        excel_tiaosheng = sh2["H21"].value or 0
        good_tiaosheng = sh2["H22"].value or 0
        ex_good_tiaosheng = sh2["H23"].value or 0
        pass_tiaosheng = sh2["H24"].value or 0
        fail_tiaosheng = sh2["H25"].value or 0

        # 坐位体前屈
        excel_zuowei = sh2["I21"].value or 0
        good_zuowei = sh2["I22"].value or 0
        ex_good_zuowei = sh2["I23"].value or 0
        pass_zuowei = sh2["I24"].value or 0
        fail_zuowei = sh2["I25"].value or 0

        # 肺活量
        excel_feihuol = sh2["J21"].value or 0
        good_feihuol = sh2["J22"].value or 0
        ex_good_feihuol = sh2["J23"].value or 0
        pass_feihuol = sh2["J24"].value or 0
        fail_feihuol = sh2["J25"].value or 0

        # BMI
        excel_BMI = sh2["K21"].value or 0
        good_BMI = sh2["K22"].value or 0
        ex_good_BMI = sh2["K23"].value or 0
        pass_BMI = sh2["K24"].value or 0
        fail_BMI = sh2["K25"].value or 0

        # 4. 内部方法：根据项目生成具体训练方案
        def get_training_plan(item_name, fail_rate, ex_good_rate):
            # 空值防护
            if not item_name:
                return "未获取到项目名称"

            fail = fail_rate * 100
            ex_good = ex_good_rate * 100

            if "50米" in item_name or "50m" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础起跑练习（原地摆臂、起跑反应练习），30米加速跑，趣味追逐跑，每周3次，降低强度提升兴趣，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点辅导起跑与反应速度，增加30米、40米间歇跑，每组2次，组间休息30秒，提升速度耐力。"
                elif ex_good >= 60:
                    return f"{item_name}：保持现有训练，增加行进间高抬腿、后蹬跑，提升爆发力与步频，培养尖子生冲刺能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化途中跑节奏，加入反应起跑练习，每组3次50米跑，巩固良好水平，提升优良率。"
                else:
                    return f"{item_name}：加强摆臂与蹬地练习，开展小组竞赛，激发积极性，提升整体优秀与良好人数。"

            elif "跳绳" in item_name:
                if fail > 20:
                    return f"{item_name}：从无绳徒手跳、单脚跳、双脚跳基础练起，每次500个，循序渐进让学生掌握基本节奏，降低不及格率。"
                elif fail > 10:
                    return f"{item_name}：分组练习，慢速连续跳+计时跳，重点纠正绊绳问题，每天5分钟专项训练，提升稳定性。"
                elif ex_good >= 60:
                    return f"{item_name}：保持快速跳训练，增加双摇尝试，并脚跳进阶动作，提升脚踝力量与熟练度。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力跳训练，30秒计时跳，提升连续跳绳能力，巩固优良水平。"
                else:
                    return f"{item_name}：增加竞赛跳绳、花样跳，提升兴趣与熟练度，扩大优秀、良好学生群体。"

            elif "坐位体前屈" in item_name:
                if fail > 20:
                    return f"{item_name}：每日课前静态拉伸（大腿后侧、腰部），坐姿体前屈保持15秒/组，做3组，提升身体柔韧度。"
                elif fail > 10:
                    return f"{item_name}：加强动态拉伸+伙伴辅助拉伸，循序渐进增加伸展幅度，帮助学生轻松达标。"
                elif ex_good >= 60:
                    return f"{item_name}：保持日常拉伸习惯，增加进阶柔韧练习，保持优秀柔韧水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日柔韧训练，巩固成绩，稳步提升优良率。"
                else:
                    return f"{item_name}：增加拉伸频率，课课练、天天练，提升整体柔韧素质。"

            # 新增：肺活量训练方案
            elif "肺活量" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础呼吸训练（腹式呼吸、吹纸条练习），配合扩胸运动，每日10分钟，提升呼吸肌力量。"
                elif fail > 10:
                    return f"{item_name}：加强深呼吸训练，结合吹气球、吹奏类小游戏，提升肺部通气能力，降低不及格率。"
                elif ex_good >= 60:
                    return f"{item_name}：保持有氧训练+深呼吸练习，维持优秀心肺功能水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日呼吸训练，配合慢跑提升耐力，巩固优良成绩。"
                else:
                    return f"{item_name}：增加有氧活动，强化呼吸训练，全面提升肺活量水平。"

            # 新增：BMI训练方案
            elif "BMI" in item_name or "身体质量" in item_name:
                if fail > 20:
                    return f"{item_name}：分层指导，偏瘦学生加强营养+力量训练，偏胖学生增加有氧慢跑，每日15分钟，改善体态。"
                elif fail > 10:
                    return f"{item_name}：结合趣味运动（跳绳、球类），控制饮食作息，帮助学生维持健康体重范围。"
                elif ex_good >= 60:
                    return f"{item_name}：保持均衡饮食+规律运动，维持健康体态，树立良好生活习惯。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持日常运动，保持健康生活方式，巩固优良体态。"
                else:
                    return f"{item_name}：加强运动与饮食引导，全面提升学生健康体重达标率。"

            else:
                return f"{item_name}：坚持常规训练，分层指导，全面提升成绩。"

        # 5. 生成五个项目方案
        plan1 = get_training_plan(item_50m, fail_50m, ex_good_50m)
        plan2 = get_training_plan(item_tiaosheng, fail_tiaosheng, ex_good_tiaosheng)
        plan3 = get_training_plan(item_zuowei, fail_zuowei, ex_good_zuowei)
        plan4 = get_training_plan(item_feihuoli, fail_feihuol, ex_good_feihuol)
        plan5 = get_training_plan(item_bmi, fail_BMI, ex_good_BMI)

        # 6. 组合最终方案
        final_plan = f"""建议训练方案：
1. {plan1}
2. {plan2}
3. {plan3}
4. {plan4}
5. {plan5}
整体坚持课课练、天天练，采用趣味化、分层化教学，全面提升学生身体素质。"""

        # 7. 写入B36并设置格式
        sh2["B36"] = final_plan
        sh2.merge_cells('B36:M43')
        sh2['B36'].font = Font(size=12, bold=True)
        sh2['B36'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

        return final_plan

    except Exception as e:
        return f"执行失败：{str(e)}"



#二三四年级 四项  50米 跳绳 坐位体前屈 仰卧起坐
def si_xunlianfangan_plan(sh2):
    """
    读取一年级体育成绩单Excel，自动生成详细训练方案并写入B36单元格

    参数：
        sh2: Excel工作表对象

    返回：
        成功返回训练方案文本，失败返回错误提示
    """
    try:

        # 2. 读取项目名称
        item_50m = sh2["G20"].value
        item_tiaosheng = sh2["H20"].value
        item_zuowei = sh2["I20"].value
        item_yangwoqizuo = sh2["J20"].value  # 新增：仰卧起坐项目名称

        # 3. 读取各项比率（空值默认为0）
        # 50米
        excel_50m = sh2["G21"].value or 0
        good_50m = sh2["G22"].value or 0
        ex_good_50m = sh2["G23"].value or 0
        pass_50m = sh2["G24"].value or 0
        fail_50m = sh2["G25"].value or 0

        # 跳绳
        excel_tiaosheng = sh2["H21"].value or 0
        good_tiaosheng = sh2["H22"].value or 0
        ex_good_tiaosheng = sh2["H23"].value or 0
        pass_tiaosheng = sh2["H24"].value or 0
        fail_tiaosheng = sh2["H25"].value or 0

        # 坐位体前屈
        excel_zuowei = sh2["I21"].value or 0
        good_zuowei = sh2["I22"].value or 0
        ex_good_zuowei = sh2["I23"].value or 0
        pass_zuowei = sh2["I24"].value or 0
        fail_zuowei = sh2["I25"].value or 0

        # 仰卧起坐（新增）
        excel_yangwoqizuo = sh2["J21"].value or 0
        good_yangwoqizuo = sh2["J22"].value or 0
        ex_good_yangwoqizuo = sh2["J23"].value or 0  # 修正拼写错误
        pass_yangwoqizuo = sh2["J24"].value or 0
        fail_yangwoqizuo = sh2["J25"].value or 0

        # 4. 内部方法：根据项目生成具体训练方案
        def get_training_plan(item_name, fail_rate, ex_good_rate):
            fail = fail_rate * 100
            ex_good = ex_good_rate * 100

            if "50米" in item_name or "50m" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础起跑练习（原地摆臂、起跑反应练习），30米加速跑，趣味追逐跑，每周3次，降低强度提升兴趣，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点辅导起跑与反应速度，增加30米、40米间歇跑，每组2次，组间休息30秒，提升速度耐力。"
                elif ex_good >= 60:
                    return f"{item_name}：保持现有训练，增加行进间高抬腿、后蹬跑，提升爆发力与步频，培养尖子生冲刺能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化途中跑节奏，加入反应起跑练习，每组3次50米跑，巩固良好水平，提升优良率。"
                else:
                    return f"{item_name}：加强摆臂与蹬地练习，开展小组竞赛，激发积极性，提升整体优秀与良好人数。"

            elif "跳绳" in item_name:
                if fail > 20:
                    return f"{item_name}：从无绳徒手跳、单脚跳、双脚跳基础练起，每次500个，循序渐进让学生掌握基本节奏，降低不及格率。"
                elif fail > 10:
                    return f"{item_name}：分组练习，慢速连续跳+计时跳，重点纠正绊绳问题，每天5分钟专项训练，提升稳定性。"
                elif ex_good >= 60:
                    return f"{item_name}：保持快速跳训练，增加双摇尝试，并脚跳进阶动作，提升脚踝力量与熟练度。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力跳训练，30秒计时跳，提升连续跳绳能力，巩固优良水平。"
                else:
                    return f"{item_name}：增加竞赛跳绳、花样跳，提升兴趣与熟练度，扩大优秀、良好学生群体。"

            elif "坐位体前屈" in item_name:
                if fail > 20:
                    return f"{item_name}：每日课前静态拉伸（大腿后侧、腰部），坐姿体前屈保持15秒/组，做3组，提升身体柔韧度。"
                elif fail > 10:
                    return f"{item_name}：加强动态拉伸+伙伴辅助拉伸，循序渐进增加伸展幅度，帮助学生轻松达标。"
                elif ex_good >= 60:
                    return f"{item_name}：保持日常拉伸习惯，增加进阶柔韧练习，保持优秀柔韧水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日柔韧训练，巩固成绩，稳步提升优良率。"
                else:
                    return f"{item_name}：增加拉伸频率，课课练、天天练，提升整体柔韧素质。"

            # 新增：仰卧起坐训练方案
            elif "仰卧起坐" in item_name:
                if fail > 20:
                    return f"{item_name}：从卷腹基础动作练起，每组10次，做3组，掌握标准动作，帮助薄弱学生快速达标。"
                elif fail > 10:
                    return f"{item_name}：分组规范训练，1分钟计时练习，重点纠正起身发力姿势，提升核心力量与完成次数。"
                elif ex_good >= 60:
                    return f"{item_name}：保持高强度训练，增加负重卷腹、快速仰卧起坐练习，突破速度与耐力，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力训练，40秒计时间歇练习，巩固优秀水平，稳步提升优良率。"
                else:
                    return f"{item_name}：每日坚持计时训练，开展小组比拼，提升核心力量，扩大优秀良好群体。"

            else:
                return f"{item_name}：坚持常规训练，分层指导，全面提升成绩。"

        # 5. 生成四个项目方案（新增仰卧起坐）
        plan1 = get_training_plan(item_50m, fail_50m, ex_good_50m)
        plan2 = get_training_plan(item_tiaosheng, fail_tiaosheng, ex_good_tiaosheng)
        plan3 = get_training_plan(item_zuowei, fail_zuowei, ex_good_zuowei)
        plan4 = get_training_plan(item_yangwoqizuo, fail_yangwoqizuo, ex_good_yangwoqizuo)

        # 6. 组合最终方案（4项）
        final_plan = f"""建议训练方案：
1. {plan1}
2. {plan2}
3. {plan3}
4. {plan4}
整体坚持课课练、天天练，采用趣味化、分层化教学，全面提升学生身体素质。"""

        # 7. 写入B36并保存
        sh2["B36"] = final_plan
        sh2.merge_cells('B36:M43')
        sh2['B36'].font = Font(size=12, bold=True)
        sh2['B36'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

        return final_plan

    except Exception as e:
        return f"执行失败：{str(e)}"



# 二年级 六项 50米 跳绳 坐位体前屈 仰卧起坐 肺活量 BMI
def liu_xunlianfangan_plan(sh2):

    try:

        # 2. 读取项目名称
        item_50m = sh2["G20"].value
        item_tiaosheng = sh2["H20"].value
        item_zuowei = sh2["I20"].value
        item_yangwoqizuo = sh2["J20"].value
        item_feihuol = sh2["K20"].value      # 新增：肺活量
        item_BMI = sh2["L20"].value          # 新增：BMI

        # 3. 读取各项比率（空值默认为0）
        # 50米
        excel_50m = sh2["G21"].value or 0
        good_50m = sh2["G22"].value or 0
        ex_good_50m = sh2["G23"].value or 0
        pass_50m = sh2["G24"].value or 0
        fail_50m = sh2["G25"].value or 0

        # 跳绳
        excel_tiaosheng = sh2["H21"].value or 0
        good_tiaosheng = sh2["H22"].value or 0
        ex_good_tiaosheng = sh2["H23"].value or 0
        pass_tiaosheng = sh2["H24"].value or 0
        fail_tiaosheng = sh2["H25"].value or 0

        # 坐位体前屈
        excel_zuowei = sh2["I21"].value or 0
        good_zuowei = sh2["I22"].value or 0
        ex_good_zuowei = sh2["I23"].value or 0
        pass_zuowei = sh2["I24"].value or 0
        fail_zuowei = sh2["I25"].value or 0

        # 仰卧起坐
        excel_yangwoqizuo = sh2["J21"].value or 0
        good_yangwoqizuo = sh2["J22"].value or 0
        ex_good_yangwoqizuo = sh2["J23"].value or 0
        pass_yangwoqizuo = sh2["J24"].value or 0
        fail_yangwoqizuo = sh2["J25"].value or 0

        # 肺活量（新增）
        excel_feihuol = sh2["K21"].value or 0
        good_feihuol = sh2["K22"].value or 0
        ex_good_feihuol = sh2["K23"].value or 0
        pass_feihuol = sh2["K24"].value or 0
        fail_feihuol = sh2["K25"].value or 0

        # BMI（新增）
        excel_BMI = sh2["L21"].value or 0
        good_BMI = sh2["L22"].value or 0
        ex_good_BMI = sh2["L23"].value or 0
        pass_BMI = sh2["L24"].value or 0
        fail_BMI = sh2["L25"].value or 0

        # 4. 内部方法：根据项目生成具体训练方案
        def get_training_plan(item_name, fail_rate, ex_good_rate):
            fail = fail_rate * 100
            ex_good = ex_good_rate * 100

            if "50米" in item_name or "50m" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础起跑练习（原地摆臂、起跑反应练习），30米加速跑，趣味追逐跑，每周3次，降低强度提升兴趣，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点辅导起跑与反应速度，增加30米、40米间歇跑，每组2次，组间休息30秒，提升速度耐力。"
                elif ex_good >= 60:
                    return f"{item_name}：保持现有训练，增加行进间高抬腿、后蹬跑，提升爆发力与步频，培养尖子生冲刺能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化途中跑节奏，加入反应起跑练习，每组3次50米跑，巩固良好水平，提升优良率。"
                else:
                    return f"{item_name}：加强摆臂与蹬地练习，开展小组竞赛，激发积极性，提升整体优秀与良好人数。"

            elif "跳绳" in item_name:
                if fail > 20:
                    return f"{item_name}：从无绳徒手跳、单脚跳、双脚跳基础练起，每次500个，循序渐进让学生掌握基本节奏，降低不及格率。"
                elif fail > 10:
                    return f"{item_name}：分组练习，慢速连续跳+计时跳，重点纠正绊绳问题，每天5分钟专项训练，提升稳定性。"
                elif ex_good >= 60:
                    return f"{item_name}：保持快速跳训练，增加双摇尝试，并脚跳进阶动作，提升脚踝力量与熟练度。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力跳训练，30秒计时跳，提升连续跳绳能力，巩固优良水平。"
                else:
                    return f"{item_name}：增加竞赛跳绳、花样跳，提升兴趣与熟练度，扩大优秀、良好学生群体。"

            elif "坐位体前屈" in item_name:
                if fail > 20:
                    return f"{item_name}：每日课前静态拉伸（大腿后侧、腰部），坐姿体前屈保持15秒/组，做3组，提升身体柔韧度。"
                elif fail > 10:
                    return f"{item_name}：加强动态拉伸+伙伴辅助拉伸，循序渐进增加伸展幅度，帮助学生轻松达标。"
                elif ex_good >= 60:
                    return f"{item_name}：保持日常拉伸习惯，增加进阶柔韧练习，保持优秀柔韧水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日柔韧训练，巩固成绩，稳步提升优良率。"
                else:
                    return f"{item_name}：增加拉伸频率，课课练、天天练，提升整体柔韧素质。"

            elif "仰卧起坐" in item_name:
                if fail > 20:
                    return f"{item_name}：从卷腹基础动作练起，每组10次，做3组，掌握标准动作，帮助薄弱学生快速达标。"
                elif fail > 10:
                    return f"{item_name}：分组规范训练，1分钟计时练习，重点纠正起身发力姿势，提升核心力量与完成次数。"
                elif ex_good >= 60:
                    return f"{item_name}：保持高强度训练，增加负重卷腹、快速仰卧起坐练习，突破速度与耐力，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力训练，40秒计时间歇练习，巩固优秀水平，稳步提升优良率。"
                else:
                    return f"{item_name}：每日坚持计时训练，开展小组比拼，提升核心力量，扩大优秀良好群体。"

            # 新增：肺活量训练方案
            elif "肺活量" in item_name:
                if fail > 20:
                    return f"{item_name}：开展腹式呼吸、吹纸条、吹气球基础练习，掌握正确呼吸方法，提升肺部通气能力。"
                elif fail > 10:
                    return f"{item_name}：加强深呼吸训练，配合慢跑提升心肺功能，规范测试吹气技巧，稳步提升成绩。"
                elif ex_good >= 60:
                    return f"{item_name}：保持有氧训练，强化深呼吸与吹气技巧，维持优秀心肺水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持日常呼吸训练与有氧运动，巩固优良成绩，提升整体水平。"
                else:
                    return f"{item_name}：增加呼吸专项练习，课课练，提升学生肺部功能与测试成绩。"

            # 新增：BMI训练方案
            elif "BMI" in item_name:
                if fail > 20:
                    return f"{item_name}：针对偏瘦/超重学生分层指导，合理搭配饮食，增加慢跑、球类运动，改善身体形态。"
                elif fail > 10:
                    return f"{item_name}：加强有氧锻炼，控制不良饮食习惯，个性化指导，帮助学生维持健康体重范围。"
                elif ex_good >= 60:
                    return f"{item_name}：保持健康生活习惯，坚持规律运动与均衡饮食，维持理想身体形态。"
                elif ex_good >= 40:
                    return f"{item_name}：持续健康管理，加强运动与饮食引导，巩固优良BMI水平。"
                else:
                    return f"{item_name}：普及健康知识，加强日常运动与饮食管理，提升整体健康体质达标率。"

            else:
                return f"{item_name}：坚持常规训练，分层指导，全面提升成绩。"

        # 5. 生成六个项目方案
        plan1 = get_training_plan(item_50m, fail_50m, ex_good_50m)
        plan2 = get_training_plan(item_tiaosheng, fail_tiaosheng, ex_good_tiaosheng)
        plan3 = get_training_plan(item_zuowei, fail_zuowei, ex_good_zuowei)
        plan4 = get_training_plan(item_yangwoqizuo, fail_yangwoqizuo, ex_good_yangwoqizuo)
        plan5 = get_training_plan(item_feihuol, fail_feihuol, ex_good_feihuol)
        plan6 = get_training_plan(item_BMI, fail_BMI, ex_good_BMI)

        # 6. 组合最终方案（6项）
        final_plan = f"""建议训练方案：
1. {plan1}
2. {plan2}
3. {plan3}
4. {plan4}
5. {plan5}
6. {plan6}
整体坚持课课练、天天练，采用趣味化、分层化教学，全面提升学生身体素质。"""

        # 7. 写入B36并保存
        sh2["B36"] = final_plan
        sh2.merge_cells('B36:M45')  # 扩大合并区域适配6项内容
        sh2['B36'].font = Font(size=12, bold=True)
        sh2['B36'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

        return final_plan

    except Exception as e:
        return f"执行失败：{str(e)}"


# 三四年级 五项  50米 跳绳 坐位体前屈 仰卧起坐 立定跳远
def zhong_wu_xunlianfangan_plan(sh2):

    try:

        # 2. 读取项目名称
        item_50m = sh2["G20"].value
        item_tiaosheng = sh2["H20"].value
        item_zuowei = sh2["I20"].value
        item_yangwoqizuo = sh2["J20"].value
        item_litiaoyuan = sh2["K20"].value  # 新增：立定跳远

        # 3. 读取各项比率（空值默认为0）
        # 50米
        excel_50m = sh2["G21"].value or 0
        good_50m = sh2["G22"].value or 0
        ex_good_50m = sh2["G23"].value or 0
        pass_50m = sh2["G24"].value or 0
        fail_50m = sh2["G25"].value or 0

        # 跳绳
        excel_tiaosheng = sh2["H21"].value or 0
        good_tiaosheng = sh2["H22"].value or 0
        ex_good_tiaosheng = sh2["H23"].value or 0
        pass_tiaosheng = sh2["H24"].value or 0
        fail_tiaosheng = sh2["H25"].value or 0

        # 坐位体前屈
        excel_zuowei = sh2["I21"].value or 0
        good_zuowei = sh2["I22"].value or 0
        ex_good_zuowei = sh2["I23"].value or 0
        pass_zuowei = sh2["I24"].value or 0
        fail_zuowei = sh2["I25"].value or 0

        # 仰卧起坐
        excel_yangwoqizuo = sh2["J21"].value or 0
        good_yangwoqizuo = sh2["J22"].value or 0
        ex_good_yangwoqizuo = sh2["J23"].value or 0
        pass_yangwoqizuo = sh2["J24"].value or 0
        fail_yangwoqizuo = sh2["J25"].value or 0

        # 立定跳远（新增）
        excel_litiaoyuan = sh2["K21"].value or 0
        good_litiaoyuan = sh2["K22"].value or 0
        ex_good_litiaoyuan = sh2["K23"].value or 0
        pass_litiaoyuan = sh2["K24"].value or 0
        fail_litiaoyuan = sh2["K25"].value or 0

        # 4. 内部方法：根据项目生成具体训练方案
        def get_training_plan(item_name, fail_rate, ex_good_rate):
            fail = fail_rate * 100
            ex_good = ex_good_rate * 100

            if "50米" in item_name or "50m" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础起跑练习（原地摆臂、起跑反应练习），30米加速跑，趣味追逐跑，每周3次，降低强度提升兴趣，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点辅导起跑与反应速度，增加30米、40米间歇跑，每组2次，组间休息30秒，提升速度耐力。"
                elif ex_good >= 60:
                    return f"{item_name}：保持现有训练，增加行进间高抬腿、后蹬跑，提升爆发力与步频，培养尖子生冲刺能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化途中跑节奏，加入反应起跑练习，每组3次50米跑，巩固良好水平，提升优良率。"
                else:
                    return f"{item_name}：加强摆臂与蹬地练习，开展小组竞赛，激发积极性，提升整体优秀与良好人数。"

            elif "跳绳" in item_name:
                if fail > 20:
                    return f"{item_name}：从无绳徒手跳、单脚跳、双脚跳基础练起，每次500个，循序渐进让学生掌握基本节奏，降低不及格率。"
                elif fail > 10:
                    return f"{item_name}：分组练习，慢速连续跳+计时跳，重点纠正绊绳问题，每天5分钟专项训练，提升稳定性。"
                elif ex_good >= 60:
                    return f"{item_name}：保持快速跳训练，增加双摇尝试，并脚跳进阶动作，提升脚踝力量与熟练度。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力跳训练，30秒计时跳，提升连续跳绳能力，巩固优良水平。"
                else:
                    return f"{item_name}：增加竞赛跳绳、花样跳，提升兴趣与熟练度，扩大优秀、良好学生群体。"

            elif "坐位体前屈" in item_name:
                if fail > 20:
                    return f"{item_name}：每日课前静态拉伸（大腿后侧、腰部），坐姿体前屈保持15秒/组，做3组，提升身体柔韧度。"
                elif fail > 10:
                    return f"{item_name}：加强动态拉伸+伙伴辅助拉伸，循序渐进增加伸展幅度，帮助学生轻松达标。"
                elif ex_good >= 60:
                    return f"{item_name}：保持日常拉伸习惯，增加进阶柔韧练习，保持优秀柔韧水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日柔韧训练，巩固成绩，稳步提升优良率。"
                else:
                    return f"{item_name}：增加拉伸频率，课课练、天天练，提升整体柔韧素质。"

            elif "仰卧起坐" in item_name:
                if fail > 20:
                    return f"{item_name}：从卷腹基础动作练起，每组10次，做3组，掌握标准动作，帮助薄弱学生快速达标。"
                elif fail > 10:
                    return f"{item_name}：分组规范训练，1分钟计时练习，重点纠正起身发力姿势，提升核心力量与完成次数。"
                elif ex_good >= 60:
                    return f"{item_name}：保持高强度训练，增加负重卷腹、快速仰卧起坐练习，突破速度与耐力，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力训练，40秒计时间歇练习，巩固优秀水平，稳步提升优良率。"
                else:
                    return f"{item_name}：每日坚持计时训练，开展小组比拼，提升核心力量，扩大优秀良好群体。"

            # 新增：立定跳远训练方案
            elif "立定跳远" in item_name:
                if fail > 20:
                    return f"{item_name}：从摆臂、屈膝、蹬地基础动作练起，原地跳跃练习，掌握标准发力姿势，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点强化下肢爆发力，开展蛙跳、台阶跳练习，规范落地缓冲技巧。"
                elif ex_good >= 60:
                    return f"{item_name}：保持爆发力训练，增加收腹跳、立定三级跳练习，提升远度，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：强化完整动作练习，提升发力连贯性，巩固优良水平，稳步提升成绩。"
                else:
                    return f"{item_name}：开展跳跃竞赛游戏，提升下肢力量与动作熟练度，扩大优秀良好群体。"

            else:
                return f"{item_name}：坚持常规训练，分层指导，全面提升成绩。"

        # 5. 生成五个项目方案
        plan1 = get_training_plan(item_50m, fail_50m, ex_good_50m)
        plan2 = get_training_plan(item_tiaosheng, fail_tiaosheng, ex_good_tiaosheng)
        plan3 = get_training_plan(item_zuowei, fail_zuowei, ex_good_zuowei)
        plan4 = get_training_plan(item_yangwoqizuo, fail_yangwoqizuo, ex_good_yangwoqizuo)
        plan5 = get_training_plan(item_litiaoyuan, fail_litiaoyuan, ex_good_litiaoyuan)

        # 6. 组合最终方案（5项）
        final_plan = f"""建议训练方案：
1. {plan1}
2. {plan2}
3. {plan3}
4. {plan4}
5. {plan5}
整体坚持课课练、天天练，采用趣味化、分层化教学，全面提升学生身体素质。"""

        # 7. 写入B36并保存（扩大单元格适配内容）
        sh2["B36"] = final_plan
        sh2.merge_cells('B36:M45')
        sh2['B36'].font = Font(size=12, bold=True)
        sh2['B36'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

        return final_plan

    except Exception as e:
        return f"执行失败：{str(e)}"


# 三四年级 七项  50米 跳绳 坐位体前屈 仰卧起坐 立定跳远 肺活量 BMI
def qi_xunlianfangan_plan(sh2):
    try:

        # 2. 读取项目名称
        item_50m = sh2["G20"].value
        item_tiaosheng = sh2["H20"].value
        item_zuowei = sh2["I20"].value
        item_yangwoqizuo = sh2["J20"].value
        item_litiaoyuan = sh2["K20"].value
        item_feihuoli = sh2["L20"].value  # 新增：肺活量
        item_BMI = sh2["M20"].value  # 新增：BMI

        # 3. 读取各项比率（空值默认为0）
        # 50米
        excel_50m = sh2["G21"].value or 0
        good_50m = sh2["G22"].value or 0
        ex_good_50m = sh2["G23"].value or 0
        pass_50m = sh2["G24"].value or 0
        fail_50m = sh2["G25"].value or 0

        # 跳绳
        excel_tiaosheng = sh2["H21"].value or 0
        good_tiaosheng = sh2["H22"].value or 0
        ex_good_tiaosheng = sh2["H23"].value or 0
        pass_tiaosheng = sh2["H24"].value or 0
        fail_tiaosheng = sh2["H25"].value or 0

        # 坐位体前屈
        excel_zuowei = sh2["I21"].value or 0
        good_zuowei = sh2["I22"].value or 0
        ex_good_zuowei = sh2["I23"].value or 0
        pass_zuowei = sh2["I24"].value or 0
        fail_zuowei = sh2["I25"].value or 0

        # 仰卧起坐
        excel_yangwoqizuo = sh2["J21"].value or 0
        good_yangwoqizuo = sh2["J22"].value or 0
        ex_good_yangwoqizuo = sh2["J23"].value or 0
        pass_yangwoqizuo = sh2["J24"].value or 0
        fail_yangwoqizuo = sh2["J25"].value or 0

        # 立定跳远
        excel_litiaoyuan = sh2["K21"].value or 0
        good_litiaoyuan = sh2["K22"].value or 0
        ex_good_litiaoyuan = sh2["K23"].value or 0
        pass_litiaoyuan = sh2["K24"].value or 0
        fail_litiaoyuan = sh2["K25"].value or 0

        # 肺活量（新增）
        excel_feihuoli = sh2["L21"].value or 0
        good_feihuoli = sh2["L22"].value or 0
        ex_good_feihuoli = sh2["L23"].value or 0
        pass_feihuoli = sh2["L24"].value or 0
        fail_feihuoli = sh2["L25"].value or 0

        # BMI（新增）
        excel_BMI = sh2["M21"].value or 0
        good_BMI = sh2["M22"].value or 0
        ex_good_BMI = sh2["M23"].value or 0
        pass_BMI = sh2["M24"].value or 0
        fail_BMI = sh2["M25"].value or 0

        # 4. 内部方法：根据项目生成具体训练方案
        def get_training_plan(item_name, fail_rate, ex_good_rate):
            fail = fail_rate * 100
            ex_good = ex_good_rate * 100

            if "50米" in item_name or "50m" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础起跑练习（原地摆臂、起跑反应练习），30米加速跑，趣味追逐跑，每周3次，降低强度提升兴趣，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点辅导起跑与反应速度，增加30米、40米间歇跑，每组2次，组间休息30秒，提升速度耐力。"
                elif ex_good >= 60:
                    return f"{item_name}：保持现有训练，增加行进间高抬腿、后蹬跑，提升爆发力与步频，培养尖子生冲刺能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化途中跑节奏，加入反应起跑练习，每组3次50米跑，巩固良好水平，提升优良率。"
                else:
                    return f"{item_name}：加强摆臂与蹬地练习，开展小组竞赛，激发积极性，提升整体优秀与良好人数。"

            elif "跳绳" in item_name:
                if fail > 20:
                    return f"{item_name}：从无绳徒手跳、单脚跳、双脚跳基础练起，每次500个，循序渐进让学生掌握基本节奏，降低不及格率。"
                elif fail > 10:
                    return f"{item_name}：分组练习，慢速连续跳+计时跳，重点纠正绊绳问题，每天5分钟专项训练，提升稳定性。"
                elif ex_good >= 60:
                    return f"{item_name}：保持快速跳训练，增加双摇尝试，并脚跳进阶动作，提升脚踝力量与熟练度。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力跳训练，30秒计时跳，提升连续跳绳能力，巩固优良水平。"
                else:
                    return f"{item_name}：增加竞赛跳绳、花样跳，提升兴趣与熟练度，扩大优秀、良好学生群体。"

            elif "坐位体前屈" in item_name:
                if fail > 20:
                    return f"{item_name}：每日课前静态拉伸（大腿后侧、腰部），坐姿体前屈保持15秒/组，做3组，提升身体柔韧度。"
                elif fail > 10:
                    return f"{item_name}：加强动态拉伸+伙伴辅助拉伸，循序渐进增加伸展幅度，帮助学生轻松达标。"
                elif ex_good >= 60:
                    return f"{item_name}：保持日常拉伸习惯，增加进阶柔韧练习，保持优秀柔韧水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日柔韧训练，巩固成绩，稳步提升优良率。"
                else:
                    return f"{item_name}：增加拉伸频率，课课练、天天练，提升整体柔韧素质。"

            elif "仰卧起坐" in item_name:
                if fail > 20:
                    return f"{item_name}：从卷腹基础动作练起，每组10次，做3组，掌握标准动作，帮助薄弱学生快速达标。"
                elif fail > 10:
                    return f"{item_name}：分组规范训练，1分钟计时练习，重点纠正起身发力姿势，提升核心力量与完成次数。"
                elif ex_good >= 60:
                    return f"{item_name}：保持高强度训练，增加负重卷腹、快速仰卧起坐练习，突破速度与耐力，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力训练，40秒计时间歇练习，巩固优秀水平，稳步提升优良率。"
                else:
                    return f"{item_name}：每日坚持计时训练，开展小组比拼，提升核心力量，扩大优秀良好群体。"

            elif "立定跳远" in item_name:
                if fail > 20:
                    return f"{item_name}：从摆臂、屈膝、蹬地基础动作练起，原地跳跃练习，掌握标准发力姿势，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点强化下肢爆发力，开展蛙跳、台阶跳练习，规范落地缓冲技巧。"
                elif ex_good >= 60:
                    return f"{item_name}：保持爆发力训练，增加收腹跳、立定三级跳练习，提升远度，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：强化完整动作练习，提升发力连贯性，巩固优良水平，稳步提升成绩。"
                else:
                    return f"{item_name}：开展跳跃竞赛游戏，提升下肢力量与动作熟练度，扩大优秀良好群体。"

            # 新增：肺活量训练方案
            elif "肺活量" in item_name:
                if fail > 20:
                    return f"{item_name}：开展腹式呼吸、吹纸条、吹气球基础练习，掌握正确呼吸方法，提升肺部通气能力。"
                elif fail > 10:
                    return f"{item_name}：加强深呼吸训练+憋气练习，配合慢跑提升心肺功能，规范测试吹气技巧，稳步提升成绩。"
                elif ex_good >= 60:
                    return f"{item_name}：保持有氧训练，强化深呼吸与吹气技巧，维持优秀心肺水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持日常呼吸训练与有氧运动，巩固优良成绩，提升整体水平。"
                else:
                    return f"{item_name}：增加呼吸专项练习，课课练，提升学生肺部功能与测试成绩。"

            # 新增：BMI训练方案
            elif "BMI" in item_name:
                if fail > 20:
                    return f"{item_name}：针对偏瘦/超重学生分层指导，合理搭配饮食，增加慢跑、球类运动，改善身体形态。"
                elif fail > 10:
                    return f"{item_name}：加强有氧锻炼，控制不良饮食习惯，个性化指导，帮助学生维持健康体重范围。"
                elif ex_good >= 60:
                    return f"{item_name}：保持健康生活习惯，坚持规律运动与均衡饮食，维持理想身体形态。"
                elif ex_good >= 40:
                    return f"{item_name}：持续健康管理，加强运动与饮食引导，巩固优良BMI水平。"
                else:
                    return f"{item_name}：普及健康知识，加强日常运动与饮食管理，提升整体健康体质达标率。"

            else:
                return f"{item_name}：坚持常规训练，分层指导，全面提升成绩。"

        # 5. 生成七个项目方案
        plan1 = get_training_plan(item_50m, fail_50m, ex_good_50m)
        plan2 = get_training_plan(item_tiaosheng, fail_tiaosheng, ex_good_tiaosheng)
        plan3 = get_training_plan(item_zuowei, fail_zuowei, ex_good_zuowei)
        plan4 = get_training_plan(item_yangwoqizuo, fail_yangwoqizuo, ex_good_yangwoqizuo)
        plan5 = get_training_plan(item_litiaoyuan, fail_litiaoyuan, ex_good_litiaoyuan)
        plan6 = get_training_plan(item_feihuoli, fail_feihuoli, ex_good_feihuoli)
        plan7 = get_training_plan(item_BMI, fail_BMI, ex_good_BMI)

        # 6. 组合最终方案（7项）
        final_plan = f"""建议训练方案：
1. {plan1}
2. {plan2}
3. {plan3}
4. {plan4}
5. {plan5}
6. {plan6}
7. {plan7}
整体坚持课课练、天天练，采用趣味化、分层化教学，全面提升学生身体素质。"""

        # 7. 写入B36并保存（扩大单元格适配7项内容）
        sh2["B36"] = final_plan
        sh2.merge_cells('B36:M48')
        sh2['B36'].font = Font(size=12, bold=True)
        sh2['B36'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

        return final_plan

    except Exception as e:
        return f"执行失败：{str(e)}"



# 五六年级 五项  50米 跳绳 坐位体前屈 仰卧起坐 50×8往返跑
def gao_wu_xunlianfangan_plan(sh2):
    """
    读取一年级体育成绩单Excel，自动生成详细训练方案并写入B36单元格

    参数：
        sh2: Excel工作表对象

    返回：
        成功返回训练方案文本，失败返回错误提示
    """
    try:

        # 2. 读取项目名称
        item_50m = sh2["G20"].value
        item_tiaosheng = sh2["H20"].value
        item_zuowei = sh2["I20"].value
        item_yangwoqizuo = sh2["J20"].value
        item_50x8 = sh2["K20"].value  # 新增：50×8往返跑

        # 3. 读取各项比率（空值默认为0）
        # 50米
        excel_50m = sh2["G21"].value or 0
        good_50m = sh2["G22"].value or 0
        ex_good_50m = sh2["G23"].value or 0
        pass_50m = sh2["G24"].value or 0
        fail_50m = sh2["G25"].value or 0

        # 跳绳
        excel_tiaosheng = sh2["H21"].value or 0
        good_tiaosheng = sh2["H22"].value or 0
        ex_good_tiaosheng = sh2["H23"].value or 0
        pass_tiaosheng = sh2["H24"].value or 0
        fail_tiaosheng = sh2["H25"].value or 0

        # 坐位体前屈
        excel_zuowei = sh2["I21"].value or 0
        good_zuowei = sh2["I22"].value or 0
        ex_good_zuowei = sh2["I23"].value or 0
        pass_zuowei = sh2["I24"].value or 0
        fail_zuowei = sh2["I25"].value or 0

        # 仰卧起坐
        excel_yangwoqizuo = sh2["J21"].value or 0
        good_yangwoqizuo = sh2["J22"].value or 0
        ex_good_yangwoqizuo = sh2["J23"].value or 0
        pass_yangwoqizuo = sh2["J24"].value or 0
        fail_yangwoqizuo = sh2["J25"].value or 0

        # 50×8往返跑（新增）
        excel_50x8 = sh2["K21"].value or 0
        good_50x8 = sh2["K22"].value or 0
        ex_good_50x8 = sh2["K23"].value or 0
        pass_50x8 = sh2["K24"].value or 0
        fail_50x8 = sh2["K25"].value or 0

        # 4. 内部方法：根据项目生成具体训练方案
        def get_training_plan(item_name, fail_rate, ex_good_rate):
            fail = fail_rate * 100
            ex_good = ex_good_rate * 100

            if "50米" in item_name or "50m" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础起跑练习（原地摆臂、起跑反应练习），30米加速跑，趣味追逐跑，每周3次，降低强度提升兴趣，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点辅导起跑与反应速度，增加30米、40米间歇跑，每组2次，组间休息30秒，提升速度耐力。"
                elif ex_good >= 60:
                    return f"{item_name}：保持现有训练，增加行进间高抬腿、后蹬跑，提升爆发力与步频，培养尖子生冲刺能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化途中跑节奏，加入反应起跑练习，每组3次50米跑，巩固良好水平，提升优良率。"
                else:
                    return f"{item_name}：加强摆臂与蹬地练习，开展小组竞赛，激发积极性，提升整体优秀与良好人数。"

            elif "跳绳" in item_name:
                if fail > 20:
                    return f"{item_name}：从无绳徒手跳、单脚跳、双脚跳基础练起，每次500个，循序渐进让学生掌握基本节奏，降低不及格率。"
                elif fail > 10:
                    return f"{item_name}：分组练习，慢速连续跳+计时跳，重点纠正绊绳问题，每天5分钟专项训练，提升稳定性。"
                elif ex_good >= 60:
                    return f"{item_name}：保持快速跳训练，增加双摇尝试，并脚跳进阶动作，提升脚踝力量与熟练度。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力跳训练，30秒计时跳，提升连续跳绳能力，巩固优良水平。"
                else:
                    return f"{item_name}：增加竞赛跳绳、花样跳，提升兴趣与熟练度，扩大优秀、良好学生群体。"

            elif "坐位体前屈" in item_name:
                if fail > 20:
                    return f"{item_name}：每日课前静态拉伸（大腿后侧、腰部），坐姿体前屈保持15秒/组，做3组，提升身体柔韧度。"
                elif fail > 10:
                    return f"{item_name}：加强动态拉伸+伙伴辅助拉伸，循序渐进增加伸展幅度，帮助学生轻松达标。"
                elif ex_good >= 60:
                    return f"{item_name}：保持日常拉伸习惯，增加进阶柔韧练习，保持优秀柔韧水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日柔韧训练，巩固成绩，稳步提升优良率。"
                else:
                    return f"{item_name}：增加拉伸频率，课课练、天天练，提升整体柔韧素质。"

            elif "仰卧起坐" in item_name:
                if fail > 20:
                    return f"{item_name}：从卷腹基础动作练起，每组10次，做3组，掌握标准动作，帮助薄弱学生快速达标。"
                elif fail > 10:
                    return f"{item_name}：分组规范训练，1分钟计时练习，重点纠正起身发力姿势，提升核心力量与完成次数。"
                elif ex_good >= 60:
                    return f"{item_name}：保持高强度训练，增加负重卷腹、快速仰卧起坐练习，突破速度与耐力，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力训练，40秒计时间歇练习，巩固优秀水平，稳步提升优良率。"
                else:
                    return f"{item_name}：每日坚持计时训练，开展小组比拼，提升核心力量，扩大优秀良好群体。"

            # 新增：50×8往返跑训练方案
            elif "50×8" in item_name or "往返跑" in item_name:
                if fail > 20:
                    return f"{item_name}：从折返跑基础动作、绕杆转弯练习练起，配合慢跑提升耐力，帮助薄弱学生掌握技巧并达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练20米折返跑，重点规范转弯减速技巧，提升耐力与节奏感，降低失误率。"
                elif ex_good >= 60:
                    return f"{item_name}：保持耐力跑训练，增加间歇折返练习，提升速度耐力，培养尖子生稳定发挥能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化完整流程训练，提升体能与转弯技巧，巩固优良水平，稳步提升成绩。"
                else:
                    return f"{item_name}：开展折返跑竞赛，提升耐力与动作规范性，扩大优秀良好群体。"

            else:
                return f"{item_name}：坚持常规训练，分层指导，全面提升成绩。"

        # 5. 生成五个项目方案
        plan1 = get_training_plan(item_50m, fail_50m, ex_good_50m)
        plan2 = get_training_plan(item_tiaosheng, fail_tiaosheng, ex_good_tiaosheng)
        plan3 = get_training_plan(item_zuowei, fail_zuowei, ex_good_zuowei)
        plan4 = get_training_plan(item_yangwoqizuo, fail_yangwoqizuo, ex_good_yangwoqizuo)
        plan5 = get_training_plan(item_50x8, fail_50x8, ex_good_50x8)

        # 6. 组合最终方案（5项）
        final_plan = f"""建议训练方案：
1. {plan1}
2. {plan2}
3. {plan3}
4. {plan4}
5. {plan5}
整体坚持课课练、天天练，采用趣味化、分层化教学，全面提升学生身体素质。"""

        # 7. 写入B36并保存（扩大单元格适配内容）
        sh2["B36"] = final_plan
        sh2.merge_cells('B36:M45')
        sh2['B36'].font = Font(size=12, bold=True)
        sh2['B36'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

        return final_plan

    except Exception as e:
        return f"执行失败：{str(e)}"


# 五六年级 六项  50米 跳绳 坐位体前屈 仰卧起坐 50×8往返跑 直臂悬垂
def gao_liu_xunlianfangan_plan(sh2):
    """
    读取一年级体育成绩单Excel，自动生成详细训练方案并写入B36单元格

    参数：
        sh2: Excel工作表对象

    返回：
        成功返回训练方案文本，失败返回错误提示
    """
    try:

        # 2. 读取项目名称
        item_50m = sh2["G20"].value
        item_tiaosheng = sh2["H20"].value
        item_zuowei = sh2["I20"].value
        item_yangwoqizuo = sh2["J20"].value
        item_50x8 = sh2["K20"].value
        item_xianchui = sh2["L20"].value  # 新增：直臂悬垂

        # 3. 读取各项比率（空值默认为0）
        # 50米
        excel_50m = sh2["G21"].value or 0
        good_50m = sh2["G22"].value or 0
        ex_good_50m = sh2["G23"].value or 0
        pass_50m = sh2["G24"].value or 0
        fail_50m = sh2["G25"].value or 0

        # 跳绳
        excel_tiaosheng = sh2["H21"].value or 0
        good_tiaosheng = sh2["H22"].value or 0
        ex_good_tiaosheng = sh2["H23"].value or 0
        pass_tiaosheng = sh2["H24"].value or 0
        fail_tiaosheng = sh2["H25"].value or 0

        # 坐位体前屈
        excel_zuowei = sh2["I21"].value or 0
        good_zuowei = sh2["I22"].value or 0
        ex_good_zuowei = sh2["I23"].value or 0
        pass_zuowei = sh2["I24"].value or 0
        fail_zuowei = sh2["I25"].value or 0

        # 仰卧起坐
        excel_yangwoqizuo = sh2["J21"].value or 0
        good_yangwoqizuo = sh2["J22"].value or 0
        ex_good_yangwoqizuo = sh2["J23"].value or 0
        pass_yangwoqizuo = sh2["J24"].value or 0
        fail_yangwoqizuo = sh2["J25"].value or 0

        # 50×8往返跑
        excel_50x8 = sh2["K21"].value or 0
        good_50x8 = sh2["K22"].value or 0
        ex_good_50x8 = sh2["K23"].value or 0
        pass_50x8 = sh2["K24"].value or 0
        fail_50x8 = sh2["K25"].value or 0

        # 直臂悬垂（新增）
        excel_xianchui = sh2["L21"].value or 0
        good_xianchui = sh2["L22"].value or 0
        ex_good_xianchui = sh2["L23"].value or 0
        pass_xianchui = sh2["L24"].value or 0
        fail_xianchui = sh2["L25"].value or 0

        # 4. 内部方法：根据项目生成具体训练方案
        def get_training_plan(item_name, fail_rate, ex_good_rate):
            fail = fail_rate * 100
            ex_good = ex_good_rate * 100

            if "50米" in item_name or "50m" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础起跑练习（原地摆臂、起跑反应练习），30米加速跑，趣味追逐跑，每周3次，降低强度提升兴趣，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点辅导起跑与反应速度，增加30米、40米间歇跑，每组2次，组间休息30秒，提升速度耐力。"
                elif ex_good >= 60:
                    return f"{item_name}：保持现有训练，增加行进间高抬腿、后蹬跑，提升爆发力与步频，培养尖子生冲刺能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化途中跑节奏，加入反应起跑练习，每组3次50米跑，巩固良好水平，提升优良率。"
                else:
                    return f"{item_name}：加强摆臂与蹬地练习，开展小组竞赛，激发积极性，提升整体优秀与良好人数。"

            elif "跳绳" in item_name:
                if fail > 20:
                    return f"{item_name}：从无绳徒手跳、单脚跳、双脚跳基础练起，每次500个，循序渐进让学生掌握基本节奏，降低不及格率。"
                elif fail > 10:
                    return f"{item_name}：分组练习，慢速连续跳+计时跳，重点纠正绊绳问题，每天5分钟专项训练，提升稳定性。"
                elif ex_good >= 60:
                    return f"{item_name}：保持快速跳训练，增加双摇尝试，并脚跳进阶动作，提升脚踝力量与熟练度。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力跳训练，30秒计时跳，提升连续跳绳能力，巩固优良水平。"
                else:
                    return f"{item_name}：增加竞赛跳绳、花样跳，提升兴趣与熟练度，扩大优秀、良好学生群体。"

            elif "坐位体前屈" in item_name:
                if fail > 20:
                    return f"{item_name}：每日课前静态拉伸（大腿后侧、腰部），坐姿体前屈保持15秒/组，做3组，提升身体柔韧度。"
                elif fail > 10:
                    return f"{item_name}：加强动态拉伸+伙伴辅助拉伸，循序渐进增加伸展幅度，帮助学生轻松达标。"
                elif ex_good >= 60:
                    return f"{item_name}：保持日常拉伸习惯，增加进阶柔韧练习，保持优秀柔韧水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日柔韧训练，巩固成绩，稳步提升优良率。"
                else:
                    return f"{item_name}：增加拉伸频率，课课练、天天练，提升整体柔韧素质。"

            elif "仰卧起坐" in item_name:
                if fail > 20:
                    return f"{item_name}：从卷腹基础动作练起，每组10次，做3组，掌握标准动作，帮助薄弱学生快速达标。"
                elif fail > 10:
                    return f"{item_name}：分组规范训练，1分钟计时练习，重点纠正起身发力姿势，提升核心力量与完成次数。"
                elif ex_good >= 60:
                    return f"{item_name}：保持高强度训练，增加负重卷腹、快速仰卧起坐练习，突破速度与耐力，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力训练，40秒计时间歇练习，巩固优秀水平，稳步提升优良率。"
                else:
                    return f"{item_name}：每日坚持计时训练，开展小组比拼，提升核心力量，扩大优秀良好群体。"

            elif "50×8" in item_name or "往返跑" in item_name:
                if fail > 20:
                    return f"{item_name}：从折返跑基础动作、绕杆转弯练习练起，配合慢跑提升耐力，帮助薄弱学生掌握技巧并达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练20米折返跑，重点规范转弯减速技巧，提升耐力与节奏感，降低失误率。"
                elif ex_good >= 60:
                    return f"{item_name}：保持耐力跑训练，增加间歇折返练习，提升速度耐力，培养尖子生稳定发挥能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化完整流程训练，提升体能与转弯技巧，巩固优良水平，稳步提升成绩。"
                else:
                    return f"{item_name}：开展折返跑竞赛，提升耐力与动作规范性，扩大优秀良好群体。"

            # 新增：直臂悬垂训练方案
            elif "直臂悬垂" in item_name or "悬垂" in item_name:
                if fail > 20:
                    return f"{item_name}：从握杠、直臂支撑基础练习开始，培养正确握姿与手臂力量，帮助薄弱学生稳定悬垂达标。"
                elif fail > 10:
                    return f"{item_name}：分组进行悬垂计时训练，强化手臂、肩背力量，规范身体姿态，提升悬垂时长。"
                elif ex_good >= 60:
                    return f"{item_name}：保持力量训练，增加负重悬垂、静态控制练习，提升上肢耐力，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：强化标准悬垂练习，提升上肢与核心力量，巩固优良水平，稳步提升成绩。"
                else:
                    return f"{item_name}：开展悬垂竞赛游戏，增强上肢力量与握力，扩大优秀良好群体。"

            else:
                return f"{item_name}：坚持常规训练，分层指导，全面提升成绩。"

        # 5. 生成六个项目方案
        plan1 = get_training_plan(item_50m, fail_50m, ex_good_50m)
        plan2 = get_training_plan(item_tiaosheng, fail_tiaosheng, ex_good_tiaosheng)
        plan3 = get_training_plan(item_zuowei, fail_zuowei, ex_good_zuowei)
        plan4 = get_training_plan(item_yangwoqizuo, fail_yangwoqizuo, ex_good_yangwoqizuo)
        plan5 = get_training_plan(item_50x8, fail_50x8, ex_good_50x8)
        plan6 = get_training_plan(item_xianchui, fail_xianchui, ex_good_xianchui)

        # 6. 组合最终方案（6项）
        final_plan = f"""建议训练方案：
1. {plan1}
2. {plan2}
3. {plan3}
4. {plan4}
5. {plan5}
6. {plan6}
整体坚持课课练、天天练，采用趣味化、分层化教学，全面提升学生身体素质。"""

        # 7. 写入B36并保存（扩大单元格适配6项内容）
        sh2["B36"] = final_plan
        sh2.merge_cells('B36:M47')
        sh2['B36'].font = Font(size=12, bold=True)
        sh2['B36'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

        return final_plan

    except Exception as e:
        return f"执行失败：{str(e)}"


# 五六年级 八项  50米 跳绳 坐位体前屈 仰卧起坐 50×8往返跑 直臂悬垂 肺活量 BMI
def gao_ba_xunlianfangan_plan(sh2):
    """
    读取一年级体育成绩单Excel，自动生成详细训练方案并写入B36单元格

    参数：
        sh2: Excel工作表对象

    返回：
        成功返回训练方案文本，失败返回错误提示
    """
    try:

        # 2. 读取项目名称
        item_50m = sh2["G20"].value
        item_tiaosheng = sh2["H20"].value
        item_zuowei = sh2["I20"].value
        item_yangwoqizuo = sh2["J20"].value
        item_50x8 = sh2["K20"].value
        item_xianchui = sh2["L20"].value
        item_feihuoli = sh2["M20"].value  # 新增：肺活量
        item_BMI = sh2["N20"].value        # 新增：BMI

        # 3. 读取各项比率（空值默认为0）
        # 50米
        excel_50m = sh2["G21"].value or 0
        good_50m = sh2["G22"].value or 0
        ex_good_50m = sh2["G23"].value or 0
        pass_50m = sh2["G24"].value or 0
        fail_50m = sh2["G25"].value or 0

        # 跳绳
        excel_tiaosheng = sh2["H21"].value or 0
        good_tiaosheng = sh2["H22"].value or 0
        ex_good_tiaosheng = sh2["H23"].value or 0
        pass_tiaosheng = sh2["H24"].value or 0
        fail_tiaosheng = sh2["H25"].value or 0

        # 坐位体前屈
        excel_zuowei = sh2["I21"].value or 0
        good_zuowei = sh2["I22"].value or 0
        ex_good_zuowei = sh2["I23"].value or 0
        pass_zuowei = sh2["I24"].value or 0
        fail_zuowei = sh2["I25"].value or 0

        # 仰卧起坐
        excel_yangwoqizuo = sh2["J21"].value or 0
        good_yangwoqizuo = sh2["J22"].value or 0
        ex_good_yangwoqizuo = sh2["J23"].value or 0
        pass_yangwoqizuo = sh2["J24"].value or 0
        fail_yangwoqizuo = sh2["J25"].value or 0

        # 50×8往返跑
        excel_50x8 = sh2["K21"].value or 0
        good_50x8 = sh2["K22"].value or 0
        ex_good_50x8 = sh2["K23"].value or 0
        pass_50x8 = sh2["K24"].value or 0
        fail_50x8 = sh2["K25"].value or 0

        # 直臂悬垂
        excel_xianchui = sh2["L21"].value or 0
        good_xianchui = sh2["L22"].value or 0
        ex_good_xianchui = sh2["L23"].value or 0
        pass_xianchui = sh2["L24"].value or 0
        fail_xianchui = sh2["L25"].value or 0

        # 肺活量（新增）
        excel_feihuoli = sh2["M21"].value or 0
        good_feihuoli = sh2["M22"].value or 0
        ex_good_feihuoli = sh2["M23"].value or 0
        pass_feihuoli = sh2["M24"].value or 0
        fail_feihuoli = sh2["M25"].value or 0

        # BMI（新增）
        excel_BMI = sh2["N21"].value or 0
        good_BMI = sh2["N22"].value or 0
        ex_good_BMI = sh2["N23"].value or 0
        pass_BMI = sh2["N24"].value or 0
        fail_BMI = sh2["N25"].value or 0

        # 4. 内部方法：根据项目生成具体训练方案
        def get_training_plan(item_name, fail_rate, ex_good_rate):
            fail = fail_rate * 100
            ex_good = ex_good_rate * 100

            if "50米" in item_name or "50m" in item_name:
                if fail > 20:
                    return f"{item_name}：开展基础起跑练习（原地摆臂、起跑反应练习），30米加速跑，趣味追逐跑，每周3次，降低强度提升兴趣，帮助薄弱学生达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练，重点辅导起跑与反应速度，增加30米、40米间歇跑，每组2次，组间休息30秒，提升速度耐力。"
                elif ex_good >= 60:
                    return f"{item_name}：保持现有训练，增加行进间高抬腿、后蹬跑，提升爆发力与步频，培养尖子生冲刺能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化途中跑节奏，加入反应起跑练习，每组3次50米跑，巩固良好水平，提升优良率。"
                else:
                    return f"{item_name}：加强摆臂与蹬地练习，开展小组竞赛，激发积极性，提升整体优秀与良好人数。"

            elif "跳绳" in item_name:
                if fail > 20:
                    return f"{item_name}：从无绳徒手跳、单脚跳、双脚跳基础练起，每次500个，循序渐进让学生掌握基本节奏，降低不及格率。"
                elif fail > 10:
                    return f"{item_name}：分组练习，慢速连续跳+计时跳，重点纠正绊绳问题，每天5分钟专项训练，提升稳定性。"
                elif ex_good >= 60:
                    return f"{item_name}：保持快速跳训练，增加双摇尝试，并脚跳进阶动作，提升脚踝力量与熟练度。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力跳训练，30秒计时跳，提升连续跳绳能力，巩固优良水平。"
                else:
                    return f"{item_name}：增加竞赛跳绳、花样跳，提升兴趣与熟练度，扩大优秀、良好学生群体。"

            elif "坐位体前屈" in item_name:
                if fail > 20:
                    return f"{item_name}：每日课前静态拉伸（大腿后侧、腰部），坐姿体前屈保持15秒/组，做3组，提升身体柔韧度。"
                elif fail > 10:
                    return f"{item_name}：加强动态拉伸+伙伴辅助拉伸，循序渐进增加伸展幅度，帮助学生轻松达标。"
                elif ex_good >= 60:
                    return f"{item_name}：保持日常拉伸习惯，增加进阶柔韧练习，保持优秀柔韧水平。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持每日柔韧训练，巩固成绩，稳步提升优良率。"
                else:
                    return f"{item_name}：增加拉伸频率，课课练、天天练，提升整体柔韧素质。"

            elif "仰卧起坐" in item_name:
                if fail > 20:
                    return f"{item_name}：从卷腹基础动作练起，每组10次，做3组，掌握标准动作，帮助薄弱学生快速达标。"
                elif fail > 10:
                    return f"{item_name}：分组规范训练，1分钟计时练习，重点纠正起身发力姿势，提升核心力量与完成次数。"
                elif ex_good >= 60:
                    return f"{item_name}：保持高强度训练，增加负重卷腹、快速仰卧起坐练习，突破速度与耐力，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：加强耐力训练，40秒计时间歇练习，巩固优秀水平，稳步提升优良率。"
                else:
                    return f"{item_name}：每日坚持计时训练，开展小组比拼，提升核心力量，扩大优秀良好群体。"

            elif "50×8" in item_name or "往返跑" in item_name:
                if fail > 20:
                    return f"{item_name}：从折返跑基础动作、绕杆转弯练习练起，配合慢跑提升耐力，帮助薄弱学生掌握技巧并达标。"
                elif fail > 10:
                    return f"{item_name}：分组训练20米折返跑，重点规范转弯减速技巧，提升耐力与节奏感，降低失误率。"
                elif ex_good >= 60:
                    return f"{item_name}：保持耐力跑训练，增加间歇折返练习，提升速度耐力，培养尖子生稳定发挥能力。"
                elif ex_good >= 40:
                    return f"{item_name}：强化完整流程训练，提升体能与转弯技巧，巩固优良水平，稳步提升成绩。"
                else:
                    return f"{item_name}：开展折返跑竞赛，提升耐力与动作规范性，扩大优秀良好群体。"

            elif "直臂悬垂" in item_name or "悬垂" in item_name:
                if fail > 20:
                    return f"{item_name}：从握杠、直臂支撑基础练习开始，培养正确握姿与手臂力量，帮助薄弱学生稳定悬垂达标。"
                elif fail > 10:
                    return f"{item_name}：分组进行悬垂计时训练，强化手臂、肩背力量，规范身体姿态，提升悬垂时长。"
                elif ex_good >= 60:
                    return f"{item_name}：保持力量训练，增加负重悬垂、静态控制练习，提升上肢耐力，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：强化标准悬垂练习，提升上肢与核心力量，巩固优良水平，稳步提升成绩。"
                else:
                    return f"{item_name}：开展悬垂竞赛游戏，增强上肢力量与握力，扩大优秀良好群体。"

            # 新增：肺活量训练方案
            elif "肺活量" in item_name:
                if fail > 20:
                    return f"{item_name}：开展腹式呼吸、吹气球、吹纸条基础练习，掌握正确呼吸方法，提升肺部通气能力。"
                elif fail > 10:
                    return f"{item_name}：加强深呼吸与憋气训练，配合慢跑、游泳提升心肺功能，规范测试吹气技巧。"
                elif ex_good >= 60:
                    return f"{item_name}：保持有氧训练，强化深呼吸练习，提升肺功能，培养尖子生。"
                elif ex_good >= 40:
                    return f"{item_name}：坚持呼吸训练与有氧运动，巩固优良水平，稳步提升肺活量。"
                else:
                    return f"{item_name}：开展呼吸训练小游戏，增强心肺功能，扩大优秀良好群体。"

            # 新增：BMI训练方案
            elif "BMI" in item_name:
                if fail > 20:
                    return f"{item_name}：针对偏瘦/超重学生分层指导，合理搭配饮食，增加慢跑、球类运动，改善身体形态。"
                elif fail > 10:
                    return f"{item_name}：加强有氧锻炼，控制不良饮食习惯，个性化指导，帮助学生维持健康体重。"
                elif ex_good >= 60:
                    return f"{item_name}：保持健康生活习惯，坚持规律运动与均衡饮食，维持理想身体形态。"
                elif ex_good >= 40:
                    return f"{item_name}：持续健康管理，加强运动与饮食引导，巩固优良BMI水平。"
                else:
                    return f"{item_name}：普及健康知识，加强日常运动与饮食管理，提升整体健康体质达标率。"

            else:
                return f"{item_name}：坚持常规训练，分层指导，全面提升成绩。"

        # 5. 生成八个项目方案
        plan1 = get_training_plan(item_50m, fail_50m, ex_good_50m)
        plan2 = get_training_plan(item_tiaosheng, fail_tiaosheng, ex_good_tiaosheng)
        plan3 = get_training_plan(item_zuowei, fail_zuowei, ex_good_zuowei)
        plan4 = get_training_plan(item_yangwoqizuo, fail_yangwoqizuo, ex_good_yangwoqizuo)
        plan5 = get_training_plan(item_50x8, fail_50x8, ex_good_50x8)
        plan6 = get_training_plan(item_xianchui, fail_xianchui, ex_good_xianchui)
        plan7 = get_training_plan(item_feihuoli, fail_feihuoli, ex_good_feihuoli)
        plan8 = get_training_plan(item_BMI, fail_BMI, ex_good_BMI)

        # 6. 组合最终方案（8项）
        final_plan = f"""建议训练方案：
1. {plan1}
2. {plan2}
3. {plan3}
4. {plan4}
5. {plan5}
6. {plan6}
7. {plan7}
8. {plan8}
整体坚持课课练、天天练，采用趣味化、分层化教学，全面提升学生身体素质。"""

        # 7. 写入B36并保存（扩大单元格适配8项内容）
        sh2["B37"] = final_plan
        sh2.merge_cells('B37:M51')
        sh2['B37'].font = Font(size=12, bold=True)
        sh2['B37'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

        return final_plan

    except Exception as e:
        return f"执行失败：{str(e)}"