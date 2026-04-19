# 导入系统路径处理模块
import os
# 导入pandas库
import pandas as pd
# 导入openpyxl（用于操作Excel单元格格式、合并单元格）
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font,Border, Side,PatternFill,numbers
from esinianji_jisuanhanshu import JiSuanPingJunFen,JiSuanYouLiangLv,JiSuan_nan_list,JiSuan_nv_list,JiSuan_tianjiapaiming,JiSuan_zongchengji,JiSuan_qudiaobujipaiming,tiaosheng_manfen
from zzhuzhuangtu import sinianji_sixiang_zhuzhuangtu,sinianji_wuxiang_zhuzhuangtu,sinianji_qixiang_zhuzhuangtu
from zxunliancelue import si_xunlianfangan_plan,zhong_wu_xunlianfangan_plan,qi_xunlianfangan_plan





def sinianji_chuliwenjian(df, origin_name):
    """
    :param df: 原始Excel的DataFrame
    :param origin_name: 原始文件名
    :return: 处理后文件的完整保存路径
    :raises Exception: 处理异常时抛出详细错误信息
    """
    try:
        # ===================== 1. 提取男生/女生数据 =====================
        duqu_nan = df.iloc[3:31].copy()  # 男生数据（3-30行）
        duqu_nv = df.iloc[35:68].copy()  # 女生数据（35-67行）

        # ===================== 2. 合并数据并清理空行 =====================
        duqu_quanbu = pd.concat([duqu_nan, duqu_nv], ignore_index=True)
        non_empty_df = duqu_quanbu.dropna(how='all').reset_index(drop=True)

        # 第二步：DataFrame数据转二维列表
        data_list = non_empty_df.values.tolist()
        nansheng = []  # 有排名的男生
        nvsheng = []  # 有排名的女生
        wenjianming , ext = os.path.splitext(origin_name)

        # # 全体学生
        quanbanpaiming = JiSuan_nan_list(data_list,wenjianming)[0] + JiSuan_nv_list(data_list,wenjianming)[0]
        # 全体学生 50排名
        quanbanpaiming = JiSuan_tianjiapaiming(4, quanbanpaiming)
        # 全体学生 50排名 跳绳排名
        quanbanpaiming = JiSuan_tianjiapaiming(8, quanbanpaiming)
        # # # 全体学生 50排名 跳绳排名  坐位体前屈排名
        quanbanpaiming = JiSuan_tianjiapaiming(13, quanbanpaiming)
        # # # 全体学生 50排名 跳绳排名  坐位体前屈排名  仰卧起坐
        quanbanpaiming = JiSuan_tianjiapaiming(17, quanbanpaiming)
        # # # 全体学生 50排名 跳绳排名  坐位体前屈排名  仰卧起坐  四项成绩
        quanbanpaiming = JiSuan_zongchengji(21, quanbanpaiming)
        # # # 全体学生 50排名 跳绳排名  坐位体前屈排名  仰卧起坐  四项成绩 立定跳远
        quanbanpaiming = JiSuan_tianjiapaiming(25, quanbanpaiming)
        # # # 全体学生 50排名 跳绳排名  坐位体前屈排名  仰卧起坐  四项成绩 立定跳远 五项成绩
        quanbanpaiming = JiSuan_zongchengji(29, quanbanpaiming)
        #全体学生 50排名 跳绳排名  坐位体前屈排名  仰卧起坐  四项成绩 立定跳远 五项成绩 肺活量
        quanbanpaiming = JiSuan_tianjiapaiming(33, quanbanpaiming)
        # 全体学生 50排名 跳绳排名  坐位体前屈排名  仰卧起坐  四项成绩 立定跳远 五项成绩 肺活量   BMI
        quanbanpaiming = JiSuan_tianjiapaiming(39, quanbanpaiming)
        # 全体学生 50排名 跳绳排名  坐位体前屈排名  仰卧起坐  四项成绩 立定跳远 五项成绩 肺活量   BMI 七项成绩
        quanbanpaiming = JiSuan_zongchengji(43, quanbanpaiming)
        #
        # # 完整的有排名的男生、女生
        for student in quanbanpaiming:
            if student[2] == '男':
                nansheng.append(student)
            elif student[2] == '女':
                nvsheng.append(student)

        # ========== 新增：处理  四项  「不计排名」的 ==========
        # 男生去掉不计排名   四项成绩
        sinan = JiSuan_qudiaobujipaiming(22,nansheng)
        # 女生去掉不计排名   四项成绩
        sinv = JiSuan_qudiaobujipaiming(22,nvsheng)
        # ========== 新增：处理  四项  「不计排名」的 ==========
        # 男生去掉不计排名   五项成绩
        wunan = JiSuan_qudiaobujipaiming(30, nansheng)
        # 女生去掉不计排名   五项成绩
        wunv = JiSuan_qudiaobujipaiming(30, nvsheng)
        # ========== 新增：处理   七项  「不计排名」的 ==========
        # 男生去掉不计排名   七项成绩
        qinan = JiSuan_qudiaobujipaiming(44,nansheng)
        # 女生去掉不计排名   七项成绩
        qinv = JiSuan_qudiaobujipaiming(44,nvsheng)
        #

        # ===================== 3. 准备保存路径 =====================
        base_dir = os.path.abspath(os.path.dirname(__file__))
        download_dir = os.path.join(base_dir, 'downloads')
        os.makedirs(download_dir, exist_ok=True)

        safe_origin_name = os.path.basename(origin_name)
        new_file_name = f"四年级体育成绩单_{safe_origin_name}"
        new_file_path = os.path.join(download_dir, new_file_name)

        # ===================== 4. 第一步写入：只写纯数据（无表头） =====================
        # 如果列表没有表头，直接用 pd.DataFrame() 转换（和原有 non_empty_df 格式一致）
        biaotou_list = ['序号', '姓名', '性别',  '50米', '', '','', '一分钟跳绳', '','', '', '', '坐位体前屈', '', '','',
                 '仰卧起坐', '', '','',  '基本四项（65分）', '', '','', '立定跳远', '','', '', '五项总分（80分）', '','', '','肺活量','','','','体重指数','','','','','','七项总分（100分）', '', '','','班级']

        data1 = ['', '', '', '成绩(s)', '得分', '排名','等级', #50米
                 '成绩(个)', '得分', '加分', '排名','等级', #跳绳
                 '成绩(cm)', '得分', '排名','等级', #坐位体前屈
                 '成绩(个)','得分', '排名','等级', #仰卧起坐
                 '总分', '平均分', '排名','等级',
                 '成绩(cm)', '得分', '排名','等级',#立定跳远
                 '总分', '平均分', '排名','等级',
                 '成绩(ml)', '得分', '排名','等级',#肺活量
                 '身高(cm)', '体重(kg)','BMI', '得分', '排名','等级',#BMI
                 '总分', '平均分','排名', '等级']

        quban_from_list = pd.DataFrame(quanbanpaiming)
        biaotou = pd.DataFrame([biaotou_list])
        data = pd.DataFrame([data1])

        # 先写入数据（从第3行开始，预留前2行给复杂表头）
        with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
            biaotou.to_excel(writer, sheet_name='成绩明细', index=False,
                                     header=False,         startrow=2 )
            data.to_excel(writer,    sheet_name='成绩明细', index=False,
                                     header=False,         startrow=3)
            # 写入成绩明细Sheet：数据从第3行开始（startrow=2），无索引、无表头
            quban_from_list.to_excel( writer,
                                     sheet_name='成绩明细', index=False,
                                     header=False,         startrow=4 )

            # 写入处理说明Sheet（保持原有逻辑）
            sheet2_data = []
            pd.DataFrame(sheet2_data).to_excel(
                writer, sheet_name='四项成绩',  index=False,  header=False  )
            # # ========== 新增：写入Sheet3（男女生汇总） ==========
            # 自定义Sheet3的内容（示例：男女生人数、平均分等统计信息，可按需修改）
            sheet3_data = []
            # 写入Sheet3（命名为「男女生汇总」，也可叫「Sheet3」）
            pd.DataFrame(sheet3_data).to_excel(
                writer,  sheet_name='五项成绩', index=False,   header=False  )
            # # 自定义Sheet4的内容（示例：男女生人数、平均分等统计信息，可按需修改）
            sheet4_data = []
            # 写入Sheet4（命名为「男女生汇总」，也可叫「Sheet4」）
            pd.DataFrame(sheet4_data).to_excel(
                writer, sheet_name='七项成绩', index=False, header=False)

        # # ===================== 5. 第二步：用openpyxl添加复杂合并表头 =====================
        # # sheet1
        # # 加载已生成的Excel文件（用于修改单元格格式）
        wb = load_workbook(new_file_path)
        sh1 = wb['成绩明细']  # 选中成绩明细Sheet
        zuida_liehao = sh1.max_column + 1 # 获取工作表最大列号，自适应所有列，避免遗漏
        # # 获取工作表实际有数据/样式的最大行号
        zuida_hanghao = sh1.max_row + 1

        # -------- 5.1 第一行：A1:I1合并，写入「济南市小学」 --------
        # merge_cells：合并单元格，格式为「起始列起始行:结束列结束行」
        sh1.merge_cells('A1:AU1')
        # 填写合并后的单元格内容
        sh1['A1'] = '济南市景山小学'
        # 设置文字居中对齐（水平+垂直）
        sh1['A1'].font = Font(size=20, bold=True, name='微软雅黑')

        # -------- 5.2 第二行：分4组合并单元格 --------
        # 第一组：A2:C2合并，写入「班级」
        sh1.merge_cells('A2:G2')
        sh1['A2'] = '班级'
        sh1['A2'].font = Font(size=13, bold=True)  # 仅设置字号20，如需优化可扩展：Font(size=20, bold=True, name='微软雅黑')
        # 第二组：D2:E2合并，写入「教师」
        sh1.merge_cells('H2:P2')
        sh1['H2'] = wenjianming
        sh1['H2'].font = Font(size=13, bold=True)
        # 第三组：F2:G2合并，写入「教师」（可根据需求修改为其他内容，比如「科目」）
        sh1.merge_cells('Q2:X2')
        sh1['Q2'] = '教师'
        sh1['Q2'].font = Font(size=13, bold=True)
        # 第四组：H2:I2合并，写入「教师」（可根据需求修改）
        sh1.merge_cells('Y2:AU2')
        sh1['Y2'] = ''

        sh1.merge_cells('A3:A4')
        sh1.merge_cells('B3:B4')
        sh1.merge_cells('C3:C4')

        sh1.merge_cells('D3:G3')
        sh1.merge_cells('H3:L3')
        sh1.merge_cells('M3:P3')
        sh1.merge_cells('Q3:T3')
        sh1.merge_cells('U3:X3')
        sh1.merge_cells('Y3:AB3')
        sh1.merge_cells('AC3:AF3')
        sh1.merge_cells('AG3:AJ3')
        sh1.merge_cells('AK3:AP3')
        sh1.merge_cells('AQ3:AT3')
        sh1.merge_cells('AU3:AU4')



        # =======================================颜色填充===================================
        # 定义颜色填充
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # 遍历所有单元格“优秀”的单元格填充为绿色，“良好”填充为黄色，“及格”填充为蓝色，“不及格”填充为红色
        # 学生成绩
        for row in sh1.iter_rows():
            for cell in row:
                if cell.value == '优秀':
                    cell.fill = green_fill
                elif cell.value == '良好':
                    cell.fill = yellow_fill
                elif cell.value == '及格':
                    cell.fill = blue_fill
                elif cell.value == '不及格':
                    cell.fill = red_fill
                elif cell.value == '不计排名':
                    cell.fill = red_fill



        # 遍历 sh1 的 I 列（第9列），从第1行到最后一行
        for row in range(1, sh1.max_row + 1):
            cell = sh1.cell(row=row, column=9)  # I列 = 第9列
            if cell.value == 120:
                cell.fill = green_fill


        # ==================================添加边框=======================================
        all_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        # 1. 定义================================居中对齐样式================================（只定义1次，重复使用，提升效率）
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # 双层循环：遍历所有行和所有列的单元格
        for row in sh1.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
                cell.border = all_border

        # ========================男女分界线===专用粗边框===========================（核心：style = 'medium'        实现边框加粗）
        # bold_side = Side(style='medium', color='black')  # medium=中粗实线，也可用'thick'=粗实线
        bold_borde = Border(left=Side(style='thin'),    right=Side(style='thin'),
                            top=Side(style='medium'),   bottom=Side(style='thin')   )
        # 2. 单独处理第23行：遍历所有列，应用粗边框
        for col_num in range(1, zuida_liehao):
            # 定位第23行第col_num列的单元格
            current_cell = sh1.cell(row=len(nansheng)+5+JiSuan_nan_list(data_list,wenjianming)[1], column=col_num)
            # 应用粗边框样式
            current_cell.border = bold_borde

        # -------- 5.3 可选：调整行高（让表头更美观） --------
        sh1.row_dimensions[1].height = 35  # 第一行高度
        sh1.row_dimensions[2].height = 28  # 第二行高度
        sh1.row_dimensions[3].height = 26  # 第二行高度
        sh1.row_dimensions[4].height = 31  # 第二行高度
        for row_num in range(5, 60):
            sh1.row_dimensions[row_num].height = 20
        # 同时固定前4行（A1~A4）和前三列（A~C），冻结位置设为 D5
        sh1.freeze_panes = 'D5'

        #===========================第三行加粗====================
        # 如果你只需要加粗，不需要调整字号，可写为 Font(bold=True)
        bold_font = Font(size=14, bold=True)
        for col_num in range(1, zuida_liehao):
            # 定位第3行第col_num列的单元格
            current_cell = sh1.cell(row=3, column=col_num)
            # 应用加粗字体样式
            current_cell.font = bold_font

        # 1. ==================定义字体样式===================3-60行调整字号为13=================
        font_13 = Font(size=13)  # 核心：设置字号为13，如需加粗可加 bold=True
        # 2. 关键：range(3, 61) 左闭右开，
        for row_num in range(4, zuida_hanghao):
            # 内层循环：遍历当前行的所有有效列（这里取A列到ZZ列，足够覆盖你的数据）
            # 如需更多列，可增大第二个参数（如100，对应CV列）
            for col_num in range(1, zuida_liehao):
                # 定位当前单元格
                current_cell = sh1.cell(row=row_num, column=col_num)
                # 应用13号字体样式
                current_cell.font = font_13

        #===================名次列改数字类型=====================
        paiming = [6,11,15,19,23,27,31,35,41,45]
        for i in paiming:
            # 遍历第6列的所有单元格
            for cell in sh1.iter_cols(min_col=i, max_col=i):
                for single_cell in cell:
                    try:
                        single_cell.value = float(single_cell.value)
                        single_cell.number_format = 'General'
                    except (ValueError, TypeError):
                        pass



        # # sheet2
        sh2 = wb['四项成绩']
        # =======================四项成绩====================================
        # 四项  五十米    # #应测人数，实测人数，平均分
        sh2['B2'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh2['C2'] = int(len(list(sinan)) + len(list(sinv)))
        sh2['D2'] = float(JiSuanPingJunFen(4,sinan,sinv)[0])
        # 四项 跳绳    # 应测人数，实测人数，平均分
        sh2['B4'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh2['C4'] = int(len(list(sinan)) + len(list(sinv)))
        sh2['D4'] = float(JiSuanPingJunFen(8,sinan,sinv)[0])
        # 四项 坐位体前屈     # 应测人数，实测人数，平均分
        sh2['B6'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh2['C6'] = int(len(list(sinan)) + len(list(sinv)))
        sh2['D6'] = float(JiSuanPingJunFen(13,sinan,sinv)[0])
        # 四项 仰卧起坐     # 应测人数，实测人数，平均分
        sh2['B8'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh2['C8'] = int(len(list(sinan)) + len(list(sinv)))
        sh2['D8'] = float(JiSuanPingJunFen(17, sinan, sinv)[0])
        # # 四项     # 应测人数，实测人数，平均分
        sh2['B10'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh2['C10'] = int(len(list(sinan)) + len(list(sinv)))
        sh2['D10'] = float(JiSuanPingJunFen(21, sinan, sinv)[0])
        #
        # # 四项  五十米   # # 男生 女生 50米 平均分
        sh2['K2'] = int(len(list(nansheng)))
        sh2['L2'] = int(len(list(sinan)))
        sh2['M2'] = float(JiSuanPingJunFen(4,sinan,sinv)[1])
        sh2['S2'] = int(len(list(nvsheng)))
        sh2['T2'] = int(len(list(sinv)))
        sh2['U2'] = float(JiSuanPingJunFen(4,sinan,sinv)[2])
        # # 四项 跳绳    # 男生 女生 跳绳 平均分
        sh2['K4'] = int(len(list(nansheng)))
        sh2['L4'] = int(len(list(sinan)))
        sh2['M4'] = float(JiSuanPingJunFen(8,sinan,sinv)[1])
        sh2['S4'] = int(len(list(nvsheng)))
        sh2['T4'] = int(len(list(sinv)))
        sh2['U4'] = float(JiSuanPingJunFen(8,sinan,sinv)[2])
        # # 四项  坐位体前屈       # 男生 女生 坐位体前屈 平均分
        sh2['K6'] = int(len(list(nansheng)))
        sh2['L6'] = int(len(list(sinan)))
        sh2['M6'] = float(JiSuanPingJunFen(13, sinan, sinv)[1])
        sh2['S6'] = int(len(list(nvsheng)))
        sh2['T6'] = int(len(list(sinv)))
        sh2['U6'] = float(JiSuanPingJunFen(13, sinan, sinv)[2])
        # # 四项  仰卧起坐       # 男生 女生 坐位体前屈 平均分
        sh2['K8'] = int(len(list(nansheng)))
        sh2['L8'] = int(len(list(sinan)))
        sh2['M8'] = float(JiSuanPingJunFen(17, sinan, sinv)[1])
        sh2['S8'] = int(len(list(nvsheng)))
        sh2['T8'] = int(len(list(sinv)))
        sh2['U8'] = float(JiSuanPingJunFen(17, sinan, sinv)[2])
        # # 四项 # 男生 女生  平均分
        sh2['K10'] = int(len(list(nansheng)))
        sh2['L10'] = int(len(list(sinan)))
        sh2['M10'] = float(JiSuanPingJunFen(21, sinan, sinv)[1])
        sh2['S10'] = int(len(list(nvsheng)))
        sh2['T10'] = int(len(list(sinv)))
        sh2['U10'] = float(JiSuanPingJunFen(21, sinan, sinv)[2])
        #
        # # 四项  五十米    # 各等级总人数
        sh2['F3'] = JiSuanYouLiangLv(6,sinan,sinv)[0]
        sh2['G3'] = JiSuanYouLiangLv(6,sinan,sinv)[1]
        sh2['H3'] = JiSuanYouLiangLv(6,sinan,sinv)[2]
        sh2['I3'] = JiSuanYouLiangLv(6,sinan,sinv)[3]
        sh2['J3'] = JiSuanYouLiangLv(6,sinan,sinv)[4]
        # # 四项  跳绳      # 各等级总人数
        sh2['F5'] = JiSuanYouLiangLv(11,sinan,sinv)[0]
        sh2['G5'] = JiSuanYouLiangLv(11,sinan,sinv)[1]
        sh2['H5'] = JiSuanYouLiangLv(11,sinan,sinv)[2]
        sh2['I5'] = JiSuanYouLiangLv(11,sinan,sinv)[3]
        sh2['J5'] = JiSuanYouLiangLv(11,sinan,sinv)[4]
        # # 四项  坐位体前屈        # 各等级总人数
        sh2['F7'] = JiSuanYouLiangLv(15, sinan, sinv)[0]
        sh2['G7'] = JiSuanYouLiangLv(15, sinan, sinv)[1]
        sh2['H7'] = JiSuanYouLiangLv(15, sinan, sinv)[2]
        sh2['I7'] = JiSuanYouLiangLv(15, sinan, sinv)[3]
        sh2['J7'] = JiSuanYouLiangLv(15, sinan, sinv)[4]
        # # 四项  仰卧起坐        # 各等级总人数
        sh2['F9'] = JiSuanYouLiangLv(19, sinan, sinv)[0]
        sh2['G9'] = JiSuanYouLiangLv(19, sinan, sinv)[1]
        sh2['H9'] = JiSuanYouLiangLv(19, sinan, sinv)[2]
        sh2['I9'] = JiSuanYouLiangLv(19, sinan, sinv)[3]
        sh2['J9'] = JiSuanYouLiangLv(19, sinan, sinv)[4]
        # # 四项  总分        # 各等级总人数
        sh2['F11'] = JiSuanYouLiangLv(23, sinan, sinv)[0]
        sh2['G11'] = JiSuanYouLiangLv(23, sinan, sinv)[1]
        sh2['H11'] = JiSuanYouLiangLv(23, sinan, sinv)[2]
        sh2['I11'] = JiSuanYouLiangLv(23, sinan, sinv)[3]
        sh2['J11'] = JiSuanYouLiangLv(23, sinan, sinv)[4]
        #
        # # 四项  五十米        # # 总等级各种率
        sh2['F2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[5])
        sh2['G2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[6])
        sh2['H2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[7])
        sh2['I2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[8])
        sh2['J2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[9])
        # # 四项  跳绳        # # 总等级各种率
        sh2['F4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[5])
        sh2['G4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[6])
        sh2['H4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[7])
        sh2['I4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[8])
        sh2['J4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[9])
        # # 四项   坐位体前屈        # # 总等级各种率
        sh2['F6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[5])
        sh2['G6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[6])
        sh2['H6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[7])
        sh2['I6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[8])
        sh2['J6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[9])
        # # 四项   仰卧起坐        # # 总等级各种率
        sh2['F8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[5])
        sh2['G8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[6])
        sh2['H8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[7])
        sh2['I8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[8])
        sh2['J8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[9])
        # # 四项   总分        # # 总等级各种率
        sh2['F10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[5])
        sh2['G10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[6])
        sh2['H10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[7])
        sh2['I10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[8])
        sh2['J10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[9])
        #
        # # 四项  五十米        # # 男生50米  各等级人数
        sh2['N3'] = JiSuanYouLiangLv(6,sinan,sinv)[10]
        sh2['O3'] = JiSuanYouLiangLv(6,sinan,sinv)[11]
        sh2['P3'] = JiSuanYouLiangLv(6,sinan,sinv)[12]
        sh2['Q3'] = JiSuanYouLiangLv(6,sinan,sinv)[13]
        sh2['R3'] = JiSuanYouLiangLv(6,sinan,sinv)[14]
        # # 四项  跳绳        # # 男生跳绳  各等级人数
        sh2['N5'] = JiSuanYouLiangLv(11, sinan, sinv)[10]
        sh2['O5'] = JiSuanYouLiangLv(11, sinan, sinv)[11]
        sh2['P5'] = JiSuanYouLiangLv(11, sinan, sinv)[12]
        sh2['Q5'] = JiSuanYouLiangLv(11, sinan, sinv)[13]
        sh2['R5'] = JiSuanYouLiangLv(11, sinan, sinv)[14]
        # # 四项  坐位体前屈      # # 男生坐位体前屈  各等级人数
        sh2['N7'] = JiSuanYouLiangLv(15, sinan, sinv)[10]
        sh2['O7'] = JiSuanYouLiangLv(15, sinan, sinv)[11]
        sh2['P7'] = JiSuanYouLiangLv(15, sinan, sinv)[12]
        sh2['Q7'] = JiSuanYouLiangLv(15, sinan, sinv)[13]
        sh2['R7'] = JiSuanYouLiangLv(15, sinan, sinv)[14]
        # # 四项  仰卧起坐      # # 男生坐位体前屈  各等级人数
        sh2['N9'] = JiSuanYouLiangLv(19, sinan, sinv)[10]
        sh2['O9'] = JiSuanYouLiangLv(19, sinan, sinv)[11]
        sh2['P9'] = JiSuanYouLiangLv(19, sinan, sinv)[12]
        sh2['Q9'] = JiSuanYouLiangLv(19, sinan, sinv)[13]
        sh2['R9'] = JiSuanYouLiangLv(19, sinan, sinv)[14]
        # # 四项  总分      # # 男生  各等级人数
        sh2['N11'] = JiSuanYouLiangLv(23, sinan, sinv)[10]
        sh2['O11'] = JiSuanYouLiangLv(23, sinan, sinv)[11]
        sh2['P11'] = JiSuanYouLiangLv(23, sinan, sinv)[12]
        sh2['Q11'] = JiSuanYouLiangLv(23, sinan, sinv)[13]
        sh2['R11'] = JiSuanYouLiangLv(23, sinan, sinv)[14]
        #
        # # 四项   五十米     # # 男生  50米各等级的率
        sh2['N2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[15])
        sh2['O2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[16])
        sh2['P2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[17])
        sh2['Q2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[18])
        sh2['R2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[19])
        # # 四项  跳绳        # # 男生  跳绳各等级的率
        sh2['N4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[15])
        sh2['O4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[16])
        sh2['P4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[17])
        sh2['Q4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[18])
        sh2['R4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[19])
        # # 四项   坐位体前屈        # # 男生  坐位体前屈各等级的率
        sh2['N6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[15])
        sh2['O6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[16])
        sh2['P6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[17])
        sh2['Q6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[18])
        sh2['R6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[19])
        # # 四项   坐位体前屈        # # 男生  坐位体前屈各等级的率
        sh2['N8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[15])
        sh2['O8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[16])
        sh2['P8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[17])
        sh2['Q8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[18])
        sh2['R8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[19])

        # # 四项   总分        # # 男生  坐位体前屈各等级的率
        sh2['N10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[15])
        sh2['O10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[16])
        sh2['P10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[17])
        sh2['Q10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[18])
        sh2['R10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[19])
        #
        # # 四项   五十米      # # 女生50米  各等级人数
        sh2['V3'] = JiSuanYouLiangLv(6,sinan,sinv)[20]
        sh2['W3'] = JiSuanYouLiangLv(6,sinan,sinv)[21]
        sh2['X3'] = JiSuanYouLiangLv(6,sinan,sinv)[22]
        sh2['Y3'] = JiSuanYouLiangLv(6,sinan,sinv)[23]
        sh2['Z3'] = JiSuanYouLiangLv(6,sinan,sinv)[24]
        # # 四项   跳绳      # # 女生跳绳  各等级人数
        sh2['V5'] = JiSuanYouLiangLv(11, sinan, sinv)[20]
        sh2['W5'] = JiSuanYouLiangLv(11, sinan, sinv)[21]
        sh2['X5'] = JiSuanYouLiangLv(11, sinan, sinv)[22]
        sh2['Y5'] = JiSuanYouLiangLv(11, sinan, sinv)[23]
        sh2['Z5'] = JiSuanYouLiangLv(11, sinan, sinv)[24]
        # # 四项  坐位体前屈    # # 女生坐位体前屈  各等级人数
        sh2['V7'] = JiSuanYouLiangLv(15, sinan, sinv)[20]
        sh2['W7'] = JiSuanYouLiangLv(15, sinan, sinv)[21]
        sh2['X7'] = JiSuanYouLiangLv(15, sinan, sinv)[22]
        sh2['Y7'] = JiSuanYouLiangLv(15, sinan, sinv)[23]
        sh2['Z7'] = JiSuanYouLiangLv(15, sinan, sinv)[24]
        # # 四项  仰卧起坐    # # 女生坐位体前屈  各等级人数
        sh2['V9'] = JiSuanYouLiangLv(19, sinan, sinv)[20]
        sh2['W9'] = JiSuanYouLiangLv(19, sinan, sinv)[21]
        sh2['X9'] = JiSuanYouLiangLv(19, sinan, sinv)[22]
        sh2['Y9'] = JiSuanYouLiangLv(19, sinan, sinv)[23]
        sh2['Z9'] = JiSuanYouLiangLv(19, sinan, sinv)[24]
        # # 四项  坐位体前屈    # # 女生坐位体前屈  各等级人数
        sh2['V11'] = JiSuanYouLiangLv(23, sinan, sinv)[20]
        sh2['W11'] = JiSuanYouLiangLv(23, sinan, sinv)[21]
        sh2['X11'] = JiSuanYouLiangLv(23, sinan, sinv)[22]
        sh2['Y11'] = JiSuanYouLiangLv(23, sinan, sinv)[23]
        sh2['Z11'] = JiSuanYouLiangLv(23, sinan, sinv)[24]
        #
        # # 四项   五十米        # # 女生50米各等级的率
        sh2['V2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[25])
        sh2['W2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[26])
        sh2['X2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[27])
        sh2['Y2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[28])
        sh2['Z2'] = float(JiSuanYouLiangLv(6,sinan,sinv)[29])
        # # 四项   跳绳  # # 女生跳绳各等级的率
        sh2['V4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[25])
        sh2['W4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[26])
        sh2['X4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[27])
        sh2['Y4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[28])
        sh2['Z4'] = float(JiSuanYouLiangLv(11, sinan, sinv)[29])
        # # 四项  坐位体前屈    # # 女生坐位体前屈各等级的率
        sh2['V6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[25])
        sh2['W6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[26])
        sh2['X6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[27])
        sh2['Y6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[28])
        sh2['Z6'] = float(JiSuanYouLiangLv(15, sinan, sinv)[29])
        # # 四项  仰卧起坐    # # 女生坐位体前屈各等级的率
        sh2['V8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[25])
        sh2['W8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[26])
        sh2['X8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[27])
        sh2['Y8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[28])
        sh2['Z8'] = float(JiSuanYouLiangLv(19, sinan, sinv)[29])
        # # 四项  总分    # # 女生   各等级的率
        sh2['V10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[25])
        sh2['W10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[26])
        sh2['X10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[27])
        sh2['Y10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[28])
        sh2['Z10'] = float(JiSuanYouLiangLv(23, sinan, sinv)[29])
        #
        #
        #跳绳满分率
        sh2['B12'] = ("达120分\n比率")
        sh2['B13'] = tiaosheng_manfen(8,sinan,sinv)[0]
        sh2['C12'] = ("达120分\n人数")
        sh2['C13'] = tiaosheng_manfen(8,sinan,sinv)[1]
        sh2['D12'] = ("100-119\n比率")
        sh2['D13'] = tiaosheng_manfen(8,sinan,sinv)[2]
        sh2['E12'] = ("100-119\n人数")
        sh2['E13'] = tiaosheng_manfen(8,sinan,sinv)[3]

        sh2['L12'] = ("男生\n达120分\n比率")
        sh2['L13'] = tiaosheng_manfen(8,sinan,sinv)[4]
        sh2['M12'] = ("男生\n达120分\n人数")
        sh2['M13'] = tiaosheng_manfen(8,sinan,sinv)[5]
        sh2['N12'] = ("男生\n100-119\n比率")
        sh2['N13'] = tiaosheng_manfen(8,sinan,sinv)[6]
        sh2['O12'] = ("男生\n100-119\n人数")
        sh2['O13'] = tiaosheng_manfen(8,sinan,sinv)[7]

        sh2['S12'] = ("女生\n达120分\n比率")
        sh2['S13'] = tiaosheng_manfen(8,sinan,sinv)[8]
        sh2['T12'] = ("女生\n达120分\n人数")
        sh2['T13'] = tiaosheng_manfen(8,sinan,sinv)[9]
        sh2['U12'] = ("女生\n100-119\n比率")
        sh2['U13'] = tiaosheng_manfen(8,sinan,sinv)[10]
        sh2['V12'] = ("女生\n100-119\n人数")
        sh2['V13'] = tiaosheng_manfen(8,sinan,sinv)[11]

        sh2['A2'] = "50米"
        sh2['A3'] = "人数"
        sh2['A4'] = "跳绳"
        sh2['A5'] = "人数"
        sh2['A6'] = "坐位体前屈"
        sh2['A7'] = "人数"
        sh2['A8'] = "仰卧起坐"
        sh2['A9'] = "人数"
        sh2['A10'] = "总分四项\n（65分）"
        sh2['A11'] = "人数"

        sh2['A12'] = "跳绳"

        sh2['B1'] = "应测人数"
        sh2['C1'] = "实测人数"
        sh2['D1'] = "平均分"
        sh2['E1'] = "名次"
        sh2['F1'] = "优秀率"
        sh2['G1'] = "良好率"
        sh2['H1'] = "优良率"
        sh2['I1'] = "及格率"
        sh2['J1'] = "不及格率"
        sh2['K1'] = "男生\n应测人数"
        sh2['L1'] = "男生\n实测人数"
        sh2['M1'] = "男生\n平均分"
        sh2['N1'] = "男生\n优秀率"
        sh2['O1'] = "男生\n良好率"
        sh2['P1'] = "男生\n优良率"
        sh2['Q1'] = "男生\n及格率"
        sh2['R1'] = "男生\n不及格率"
        sh2['S1'] = "女生\n应测人数"
        sh2['T1'] = "女生\n实测人数"
        sh2['U1'] = "女生\n平均分"
        sh2['V1'] = "女生\n优秀率"
        sh2['W1'] = "女生\n良好率"
        sh2['X1'] = "女生\n优良率"
        sh2['Y1'] = "女生\n及格率"
        sh2['Z1'] = "女生\n不及格率"

        # sh2.merge_cells('A10:Z11')
        # 同时固定前4行（A1~A4）和前三列（A~C），冻结位置设为 D5
        sh2.freeze_panes = 'B2'

        # 全部内容居中  四项成绩
        for row in sh2.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                #添加边框
                cell.border = all_border

        #调整宽度
        kuanhao =['A','B','C','J','K','L','R','S','T','Z']
        for i in kuanhao:
            sh2.column_dimensions[i].width = 12
        # 四项成绩
        # 定义百分数格式（保留两位小数）
        baifenshu = numbers.FORMAT_PERCENTAGE_00  # 格式为“85.00%”
        danyuangehao = ['B13','D13','L13','N13','S13','U13']
        for i in danyuangehao:
            sh2[i].number_format = '0.00%'
        # 区域1：男生/女生各项目的率（如优秀率、良好率等）
        xuhao = [2,4,6,8,10]
        #区域1：优秀率、良好率
        for i in xuhao:
            # 四项成绩
            for row in sh2.iter_rows(min_row=i, max_row=i, min_col=5, max_col=10):
                for cell in row:
                    cell.number_format = baifenshu
        # 区域2：男生/女生各项目的率（如优秀率、良好率等）
        for i in xuhao:
            for row in sh2.iter_rows(min_row=i, max_row=i, min_col=14, max_col=18):
                for cell in row:
                    cell.number_format = baifenshu
        # 区域3：男生/女生各项目的率（如优秀率、良好率等）
        for i in xuhao:
            for row in sh2.iter_rows(min_row=i, max_row=i, min_col=22, max_col=26):
                for cell in row:
                    cell.number_format = baifenshu


        # 步骤1：创建一个字体对象，设置字体大小为 20（其他属性保持默认，如字体、加粗等不变）
        target_font = Font(size=13)  # size 参数指定字体大小，这里设为 20
        # 步骤2：遍历工作表中所有有内容的单元格（高效遍历）
        # 若想遍历整个工作表（包括空单元格），可直接用 sh1.iter_rows() 无参数
        for row in sh2.iter_rows(min_row=1, max_row=sh2.max_row, min_col=1, max_col=sh2.max_column):
            for cell in row:
                # 步骤3：给单元格应用预设的字体对象
                cell.font = target_font

        # ================添加柱状图=================
        sinianji_sixiang_zhuzhuangtu(sh2)
        # 循环设置 B20~D25 居中 + 边框
        for row in sh2['B20:D25']:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  # 水平+垂直居中
                cell.border = all_border  # 添加边框


        # ==================添加训练策略=======================
        si_xunlianfangan_plan(sh2)



        # #
        # # # sheet3
        sh3 = wb['五项成绩']
        # =======================五项成绩====================================
        # 全部五项  五十米    # #应测人数，实测人数，平均分
        sh3['B2'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh3['C2'] = int(len(list(wunan)) + len(list(wunv)))
        sh3['D2'] = float(JiSuanPingJunFen(4, wunan, wunv)[0])
        # # 全部五项 跳绳    # 应测人数，实测人数，平均分
        sh3['B4'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh3['C4'] = int(len(list(wunan)) + len(list(wunv)))
        sh3['D4'] = float(JiSuanPingJunFen(8, wunan, wunv)[0])
        # 全部五项 坐位体前屈     # 应测人数，实测人数，平均分
        sh3['B6'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh3['C6'] = int(len(list(wunan)) + len(list(wunv)))
        sh3['D6'] = float(JiSuanPingJunFen(13, wunan, wunv)[0])
        # # 全部五项 仰卧起坐     # 应测人数，实测人数，平均分
        sh3['B8'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh3['C8'] = int(len(list(wunan)) + len(list(wunv)))
        sh3['D8'] = float(JiSuanPingJunFen(17, wunan, wunv)[0])
        # # 全部五项 立定跳远     # 应测人数，实测人数，平均分
        sh3['B10'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh3['C10'] = int(len(list(wunan)) + len(list(wunv)))
        sh3['D10'] = float(JiSuanPingJunFen(25, wunan, wunv)[0])
        # # 全部五项     # 应测人数，实测人数，平均分
        sh3['B12'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh3['C12'] = int(len(list(wunan)) + len(list(wunv)))
        sh3['D12'] = float(JiSuanPingJunFen(29, wunan, wunv)[0])
        # #
        # # # 全部五项  五十米   # # 男生 女生 50米 平均分
        sh3['K2'] = int(len(list(nansheng)))
        sh3['L2'] = int(len(list(wunan)))
        sh3['M2'] = float(JiSuanPingJunFen(4, wunan, wunv)[1])
        sh3['S2'] = int(len(list(nvsheng)))
        sh3['T2'] = int(len(list(wunv)))
        sh3['U2'] = float(JiSuanPingJunFen(4, wunan, wunv)[2])
        # # # 全部五项 跳绳    # 男生 女生 跳绳 平均分
        sh3['K4'] = int(len(list(nansheng)))
        sh3['L4'] = int(len(list(wunan)))
        sh3['M4'] = float(JiSuanPingJunFen(8, wunan, wunv)[1])
        sh3['S4'] = int(len(list(nvsheng)))
        sh3['T4'] = int(len(list(wunv)))
        sh3['U4'] = float(JiSuanPingJunFen(8, wunan, wunv)[2])
        # # # 全部五项  坐位体前屈       # 男生 女生 坐位体前屈 平均分
        sh3['K6'] = int(len(list(nansheng)))
        sh3['L6'] = int(len(list(wunan)))
        sh3['M6'] = float(JiSuanPingJunFen(13, wunan, wunv)[1])
        sh3['S6'] = int(len(list(nvsheng)))
        sh3['T6'] = int(len(list(wunv)))
        sh3['U6'] = float(JiSuanPingJunFen(13, wunan, wunv)[2])
        # # # 全部五项  仰卧起坐       # 男生 女生  平均分
        sh3['K8'] = int(len(list(nansheng)))
        sh3['L8'] = int(len(list(wunan)))
        sh3['M8'] = float(JiSuanPingJunFen(17, wunan, wunv)[1])
        sh3['S8'] = int(len(list(nvsheng)))
        sh3['T8'] = int(len(list(wunv)))
        sh3['U8'] = float(JiSuanPingJunFen(17, wunan, wunv)[2])
        # # # 全部五项  立定跳远       # 男生 女生  平均分
        sh3['K10'] = int(len(list(nansheng)))
        sh3['L10'] = int(len(list(wunan)))
        sh3['M10'] = float(JiSuanPingJunFen(25, wunan, wunv)[1])
        sh3['S10'] = int(len(list(nvsheng)))
        sh3['T10'] = int(len(list(wunv)))
        sh3['U10'] = float(JiSuanPingJunFen(25, wunan, wunv)[2])
        # # # 全部五项 # 男生 女生  平均分
        sh3['K12'] = int(len(list(nansheng)))
        sh3['L12'] = int(len(list(wunan)))
        sh3['M12'] = float(JiSuanPingJunFen(29, wunan, wunv)[1])
        sh3['S12'] = int(len(list(nvsheng)))
        sh3['T12'] = int(len(list(wunv)))
        sh3['U12'] = float(JiSuanPingJunFen(29, wunan, wunv)[2])
        # #
        # # # 全部五项  五十米    # 各等级总人数
        sh3['F3'] = JiSuanYouLiangLv(6, wunan, wunv)[0]
        sh3['G3'] = JiSuanYouLiangLv(6, wunan, wunv)[1]
        sh3['H3'] = JiSuanYouLiangLv(6, wunan, wunv)[2]
        sh3['I3'] = JiSuanYouLiangLv(6, wunan, wunv)[3]
        sh3['J3'] = JiSuanYouLiangLv(6, wunan, wunv)[4]
        # # # 全部五项  跳绳      # 各等级总人数
        sh3['F5'] = JiSuanYouLiangLv(11, wunan, wunv)[0]
        sh3['G5'] = JiSuanYouLiangLv(11, wunan, wunv)[1]
        sh3['H5'] = JiSuanYouLiangLv(11, wunan, wunv)[2]
        sh3['I5'] = JiSuanYouLiangLv(11, wunan, wunv)[3]
        sh3['J5'] = JiSuanYouLiangLv(11, wunan, wunv)[4]
        # # # 全部五项  坐位体前屈        # 各等级总人数
        sh3['F7'] = JiSuanYouLiangLv(15, wunan, wunv)[0]
        sh3['G7'] = JiSuanYouLiangLv(15, wunan, wunv)[1]
        sh3['H7'] = JiSuanYouLiangLv(15, wunan, wunv)[2]
        sh3['I7'] = JiSuanYouLiangLv(15, wunan, wunv)[3]
        sh3['J7'] = JiSuanYouLiangLv(15, wunan, wunv)[4]
        # # # 全部五项  仰卧起坐        # 各等级总人数
        sh3['F9'] = JiSuanYouLiangLv(19, wunan, wunv)[0]
        sh3['G9'] = JiSuanYouLiangLv(19, wunan, wunv)[1]
        sh3['H9'] = JiSuanYouLiangLv(19, wunan, wunv)[2]
        sh3['I9'] = JiSuanYouLiangLv(19, wunan, wunv)[3]
        sh3['J9'] = JiSuanYouLiangLv(19, wunan, wunv)[4]
        # # # 全部五项  立定跳远        # 各等级总人数
        sh3['F11'] = JiSuanYouLiangLv(27, wunan, wunv)[0]
        sh3['G11'] = JiSuanYouLiangLv(27, wunan, wunv)[1]
        sh3['H11'] = JiSuanYouLiangLv(27, wunan, wunv)[2]
        sh3['I11'] = JiSuanYouLiangLv(27, wunan, wunv)[3]
        sh3['J11'] = JiSuanYouLiangLv(27, wunan, wunv)[4]
        # # # 全部五项  总分        # 各等级总人数
        sh3['F13'] = JiSuanYouLiangLv(31, wunan, wunv)[0]
        sh3['G13'] = JiSuanYouLiangLv(31, wunan, wunv)[1]
        sh3['H13'] = JiSuanYouLiangLv(31, wunan, wunv)[2]
        sh3['I13'] = JiSuanYouLiangLv(31, wunan, wunv)[3]
        sh3['J13'] = JiSuanYouLiangLv(31, wunan, wunv)[4]
        # #
        # # # 全部五项  五十米        # # 总等级各种率
        sh3['F2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[5])
        sh3['G2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[6])
        sh3['H2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[7])
        sh3['I2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[8])
        sh3['J2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[9])
        # # # 全部五项  跳绳        # # 总等级各种率
        sh3['F4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[5])
        sh3['G4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[6])
        sh3['H4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[7])
        sh3['I4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[8])
        sh3['J4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[9])
        # # # 全部五项   坐位体前屈        # # 总等级各种率
        sh3['F6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[5])
        sh3['G6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[6])
        sh3['H6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[7])
        sh3['I6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[8])
        sh3['J6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[9])
        # # # 全部五项   仰卧起坐        # # 总等级各种率
        sh3['F8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[5])
        sh3['G8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[6])
        sh3['H8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[7])
        sh3['I8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[8])
        sh3['J8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[9])
        # # # 全部五项   立定跳远        # # 总等级各种率
        sh3['F10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[5])
        sh3['G10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[6])
        sh3['H10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[7])
        sh3['I10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[8])
        sh3['J10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[9])
        #
        # # # 全部五项   总分        # # 总等级各种率
        sh3['F12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[5])
        sh3['G12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[6])
        sh3['H12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[7])
        sh3['I12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[8])
        sh3['J12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[9])
        # #
        # # # 全部五项  五十米        # # 男生50米  各等级人数
        sh3['N3'] = JiSuanYouLiangLv(6, wunan, wunv)[10]
        sh3['O3'] = JiSuanYouLiangLv(6, wunan, wunv)[11]
        sh3['P3'] = JiSuanYouLiangLv(6, wunan, wunv)[12]
        sh3['Q3'] = JiSuanYouLiangLv(6, wunan, wunv)[13]
        sh3['R3'] = JiSuanYouLiangLv(6, wunan, wunv)[14]
        # # # 全部五项  跳绳        # # 男生跳绳  各等级人数
        sh3['N5'] = JiSuanYouLiangLv(11, wunan, wunv)[10]
        sh3['O5'] = JiSuanYouLiangLv(11, wunan, wunv)[11]
        sh3['P5'] = JiSuanYouLiangLv(11, wunan, wunv)[12]
        sh3['Q5'] = JiSuanYouLiangLv(11, wunan, wunv)[13]
        sh3['R5'] = JiSuanYouLiangLv(11, wunan, wunv)[14]
        # # # 全部五项  坐位体前屈      # # 男生坐位体前屈  各等级人数
        sh3['N7'] = JiSuanYouLiangLv(15, wunan, wunv)[10]
        sh3['O7'] = JiSuanYouLiangLv(15, wunan, wunv)[11]
        sh3['P7'] = JiSuanYouLiangLv(15, wunan, wunv)[12]
        sh3['Q7'] = JiSuanYouLiangLv(15, wunan, wunv)[13]
        sh3['R7'] = JiSuanYouLiangLv(15, wunan, wunv)[14]
        # # # 全部五项  仰卧起坐      # # 男生   各等级人数
        sh3['N9'] = JiSuanYouLiangLv(19, wunan, wunv)[10]
        sh3['O9'] = JiSuanYouLiangLv(19, wunan, wunv)[11]
        sh3['P9'] = JiSuanYouLiangLv(19, wunan, wunv)[12]
        sh3['Q9'] = JiSuanYouLiangLv(19, wunan, wunv)[13]
        sh3['R9'] = JiSuanYouLiangLv(19, wunan, wunv)[14]
        # # # 全部五项  立定跳远      # # 男生   各等级人数
        sh3['N11'] = JiSuanYouLiangLv(27, wunan, wunv)[10]
        sh3['O11'] = JiSuanYouLiangLv(27, wunan, wunv)[11]
        sh3['P11'] = JiSuanYouLiangLv(27, wunan, wunv)[12]
        sh3['Q11'] = JiSuanYouLiangLv(27, wunan, wunv)[13]
        sh3['R11'] = JiSuanYouLiangLv(27, wunan, wunv)[14]
        #
        # # # 全部五项  总分      # # 男生  各等级人数
        sh3['N13'] = JiSuanYouLiangLv(31, wunan, wunv)[10]
        sh3['O13'] = JiSuanYouLiangLv(31, wunan, wunv)[11]
        sh3['P13'] = JiSuanYouLiangLv(31, wunan, wunv)[12]
        sh3['Q13'] = JiSuanYouLiangLv(31, wunan, wunv)[13]
        sh3['R13'] = JiSuanYouLiangLv(31, wunan, wunv)[14]
        # #
        # # # 全部五项   五十米     # # 男生  50米各等级的率
        sh3['N2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[15])
        sh3['O2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[16])
        sh3['P2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[17])
        sh3['Q2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[18])
        sh3['R2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[19])
        # # # 全部五项  跳绳        # # 男生  跳绳各等级的率
        sh3['N4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[15])
        sh3['O4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[16])
        sh3['P4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[17])
        sh3['Q4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[18])
        sh3['R4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[19])
        # # # 全部五项   坐位体前屈        # # 男生  坐位体前屈各等级的率
        sh3['N6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[15])
        sh3['O6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[16])
        sh3['P6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[17])
        sh3['Q6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[18])
        sh3['R6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[19])
        # # # 全部五项   仰卧起坐        # # 男生    各等级的率
        sh3['N8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[15])
        sh3['O8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[16])
        sh3['P8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[17])
        sh3['Q8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[18])
        sh3['R8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[19])
        # # # 全部五项   立定跳远        # # 男生    各等级的率
        sh3['N10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[15])
        sh3['O10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[16])
        sh3['P10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[17])
        sh3['Q10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[18])
        sh3['R10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[19])
        # # # 全部五项   总分        # # 男生    各等级的率
        sh3['N12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[15])
        sh3['O12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[16])
        sh3['P12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[17])
        sh3['Q12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[18])
        sh3['R12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[19])
        # #
        # # # 全部五项   五十米      # # 女生50米  各等级人数
        sh3['V3'] = JiSuanYouLiangLv(6, wunan, wunv)[20]
        sh3['W3'] = JiSuanYouLiangLv(6, wunan, wunv)[21]
        sh3['X3'] = JiSuanYouLiangLv(6, wunan, wunv)[22]
        sh3['Y3'] = JiSuanYouLiangLv(6, wunan, wunv)[23]
        sh3['Z3'] = JiSuanYouLiangLv(6, wunan, wunv)[24]
        # # # 全部五项   跳绳      # # 女生跳绳  各等级人数
        sh3['V5'] = JiSuanYouLiangLv(11, wunan, wunv)[20]
        sh3['W5'] = JiSuanYouLiangLv(11, wunan, wunv)[21]
        sh3['X5'] = JiSuanYouLiangLv(11, wunan, wunv)[22]
        sh3['Y5'] = JiSuanYouLiangLv(11, wunan, wunv)[23]
        sh3['Z5'] = JiSuanYouLiangLv(11, wunan, wunv)[24]
        # # # 全部五项  坐位体前屈    # # 女生坐位体前屈  各等级人数
        sh3['V7'] = JiSuanYouLiangLv(15, wunan, wunv)[20]
        sh3['W7'] = JiSuanYouLiangLv(15, wunan, wunv)[21]
        sh3['X7'] = JiSuanYouLiangLv(15, wunan, wunv)[22]
        sh3['Y7'] = JiSuanYouLiangLv(15, wunan, wunv)[23]
        sh3['Z7'] = JiSuanYouLiangLv(15, wunan, wunv)[24]
        # # # 全部五项  仰卧起坐    # # 女生   各等级人数
        sh3['V9'] = JiSuanYouLiangLv(19, wunan, wunv)[20]
        sh3['W9'] = JiSuanYouLiangLv(19, wunan, wunv)[21]
        sh3['X9'] = JiSuanYouLiangLv(19, wunan, wunv)[22]
        sh3['Y9'] = JiSuanYouLiangLv(19, wunan, wunv)[23]
        sh3['Z9'] = JiSuanYouLiangLv(19, wunan, wunv)[24]
        # # # 全部五项  立定跳远    # # 女生   各等级人数
        sh3['V11'] = JiSuanYouLiangLv(27, wunan, wunv)[20]
        sh3['W11'] = JiSuanYouLiangLv(27, wunan, wunv)[21]
        sh3['X11'] = JiSuanYouLiangLv(27, wunan, wunv)[22]
        sh3['Y11'] = JiSuanYouLiangLv(27, wunan, wunv)[23]
        sh3['Z11'] = JiSuanYouLiangLv(27, wunan, wunv)[24]
        # # # 全部五项       # # 女生    各等级人数
        sh3['V13'] = JiSuanYouLiangLv(31, wunan, wunv)[20]
        sh3['W13'] = JiSuanYouLiangLv(31, wunan, wunv)[21]
        sh3['X13'] = JiSuanYouLiangLv(31, wunan, wunv)[22]
        sh3['Y13'] = JiSuanYouLiangLv(31, wunan, wunv)[23]
        sh3['Z13'] = JiSuanYouLiangLv(31, wunan, wunv)[24]
        # #
        # # # 全部五项   五十米        # # 女生50米各等级的率
        sh3['V2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[25])
        sh3['W2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[26])
        sh3['X2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[27])
        sh3['Y2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[28])
        sh3['Z2'] = float(JiSuanYouLiangLv(6, wunan, wunv)[29])
        # # # 全部五项   跳绳  # # 女生跳绳各等级的率
        sh3['V4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[25])
        sh3['W4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[26])
        sh3['X4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[27])
        sh3['Y4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[28])
        sh3['Z4'] = float(JiSuanYouLiangLv(11, wunan, wunv)[29])
        # # # 全部五项  坐位体前屈    # # 女生坐位体前屈各等级的率
        sh3['V6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[25])
        sh3['W6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[26])
        sh3['X6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[27])
        sh3['Y6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[28])
        sh3['Z6'] = float(JiSuanYouLiangLv(15, wunan, wunv)[29])
        # #
        # # # 全部五项  仰卧起坐    # # 女生   各等级的率
        sh3['V8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[25])
        sh3['W8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[26])
        sh3['X8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[27])
        sh3['Y8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[28])
        sh3['Z8'] = float(JiSuanYouLiangLv(19, wunan, wunv)[29])
        # # # 全部五项  立定跳远    # # 女生   各等级的率
        sh3['V10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[25])
        sh3['W10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[26])
        sh3['X10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[27])
        sh3['Y10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[28])
        sh3['Z10'] = float(JiSuanYouLiangLv(27, wunan, wunv)[29])
        # # # 全部五项  总分    # # 女生   各等级的率
        sh3['V12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[25])
        sh3['W12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[26])
        sh3['X12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[27])
        sh3['Y12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[28])
        sh3['Z12'] = float(JiSuanYouLiangLv(31, wunan, wunv)[29])
        # #
        sh3['A2'] = "50米"
        sh3['A3'] = "人数"
        sh3['A4'] = "跳绳"
        sh3['A5'] = "人数"
        sh3['A6'] = "坐位体前屈"
        sh3['A7'] = "人数"
        sh3['A8'] = "仰卧起坐"
        sh3['A9'] = "人数"
        sh3['A10'] = "立定跳远"
        sh3['A11'] = "人数"
        sh3['A12'] = "总分五项\n（80分）"
        sh3['A13'] = "人数"
        # #
        sh3['B1'] = "应测人数"
        sh3['C1'] = "实测人数"
        sh3['D1'] = "平均分"
        sh3['E1'] = "名次"
        sh3['F1'] = "优秀率"
        sh3['G1'] = "良好率"
        sh3['H1'] = "优良率"
        sh3['I1'] = "及格率"
        sh3['J1'] = "不及格率"
        sh3['K1'] = "男生\n应测人数"
        sh3['L1'] = "男生\n实测人数"
        sh3['M1'] = "男生\n平均分"
        sh3['N1'] = "男生\n优秀率"
        sh3['O1'] = "男生\n良好率"
        sh3['P1'] = "男生\n优良率"
        sh3['Q1'] = "男生\n及格率"
        sh3['R1'] = "男生\n不及格率"
        sh3['S1'] = "女生\n应测人数"
        sh3['T1'] = "女生\n实测人数"
        sh3['U1'] = "女生\n平均分"
        sh3['V1'] = "女生\n优秀率"
        sh3['W1'] = "女生\n良好率"
        sh3['X1'] = "女生\n优良率"
        sh3['Y1'] = "女生\n及格率"
        sh3['Z1'] = "女生\n不及格率"
        #
        # 同时固定前4行（A1~A4）和前三列（A~C），冻结位置设为 D5
        sh3.freeze_panes = 'B2'

        # 全部内容居中  四项成绩
        for row in sh3.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # 添加边框
                cell.border = all_border

        # 调整宽度
        kuanhao = ['A', 'B', 'C', 'J', 'K', 'L', 'R', 'S', 'T', 'Z']
        for i in kuanhao:
            sh3.column_dimensions[i].width = 12
        # 四项成绩
        # 定义百分数格式（保留两位小数）
        baifenshu = numbers.FORMAT_PERCENTAGE_00  # 格式为“85.00%”
        # 区域1：男生/女生各项目的率（如优秀率、良好率等）
        xuhao = [2, 4, 6, 8,10,12]
        # 区域1：优秀率、良好率
        for i in xuhao:
            # 四项成绩
            for row in sh3.iter_rows(min_row=i, max_row=i, min_col=5, max_col=10):
                for cell in row:
                    cell.number_format = baifenshu
        # 区域2：男生/女生各项目的率（如优秀率、良好率等）
        for i in xuhao:
            for row in sh3.iter_rows(min_row=i, max_row=i, min_col=14, max_col=18):
                for cell in row:
                    cell.number_format = baifenshu
        # 区域3：男生/女生各项目的率（如优秀率、良好率等）
        for i in xuhao:
            for row in sh3.iter_rows(min_row=i, max_row=i, min_col=22, max_col=26):
                for cell in row:
                    cell.number_format = baifenshu

        # 步骤1：创建一个字体对象，设置字体大小为 20（其他属性保持默认，如字体、加粗等不变）
        target_font = Font(size=13)  # size 参数指定字体大小，这里设为 20
        # 步骤2：遍历工作表中所有有内容的单元格（高效遍历）
        # 若想遍历整个工作表（包括空单元格），可直接用 sh1.iter_rows() 无参数
        for row in sh3.iter_rows(min_row=1, max_row=sh3.max_row, min_col=1, max_col=sh3.max_column):
            for cell in row:
                # 步骤3：给单元格应用预设的字体对象
                cell.font = target_font
        # ================添加柱状图=================
        sinianji_wuxiang_zhuzhuangtu(sh3)
        # 循环设置 B20~D25 居中 + 边框
        for row in sh3['B20:D25']:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  # 水平+垂直居中
                cell.border = all_border  # 添加边框

        # ==================添加训练策略=======================
        zhong_wu_xunlianfangan_plan(sh3)


        #
        # # sheet4
        sh4 = wb['七项成绩']
        # =======================四项成绩====================================
        # 全部七项  五十米    # #应测人数，实测人数，平均分
        sh4['B2'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh4['C2'] = int(len(list(qinan)) + len(list(qinv)))
        sh4['D2'] = float(JiSuanPingJunFen(4, qinan, qinv)[0])
        # # 全部七项 跳绳    # 应测人数，实测人数，平均分
        sh4['B4'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh4['C4'] = int(len(list(qinan)) + len(list(qinv)))
        sh4['D4'] = float(JiSuanPingJunFen(8, qinan, qinv)[0])
        # # 全部七项 坐位体前屈     # 应测人数，实测人数，平均分
        sh4['B6'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh4['C6'] = int(len(list(qinan)) + len(list(qinv)))
        sh4['D6'] = float(JiSuanPingJunFen(13, qinan, qinv)[0])
        # # 全部七项 仰卧起坐     # 应测人数，实测人数，平均分
        sh4['B8'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh4['C8'] = int(len(list(qinan)) + len(list(qinv)))
        sh4['D8'] = float(JiSuanPingJunFen(17, qinan, qinv)[0])
        # # 全部七项 立定跳远     # 应测人数，实测人数，平均分
        sh4['B10'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh4['C10'] = int(len(list(qinan)) + len(list(qinv)))
        sh4['D10'] = float(JiSuanPingJunFen(25, qinan, qinv)[0])
        # # 全部七项 肺活量     # 应测人数，实测人数，平均分
        sh4['B12'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh4['C12'] = int(len(list(qinan)) + len(list(qinv)))
        sh4['D12'] = float(JiSuanPingJunFen(33, qinan, qinv)[0])
        # # 全部七项 BMI     # 应测人数，实测人数，平均分
        sh4['B14'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh4['C14'] = int(len(list(qinan)) + len(list(qinv)))
        sh4['D14'] = float(JiSuanPingJunFen(39, qinan, qinv)[0])
        # # 全部七项     # 应测人数，实测人数，平均分
        sh4['B16'] = int(len(list(nansheng)) + len(list(nvsheng)))
        sh4['C16'] = int(len(list(qinan)) + len(list(qinv)))
        sh4['D16'] = float(JiSuanPingJunFen(43, qinan, qinv)[0])
        # #
        # # # 全部七项  五十米   # # 男生 女生 50米 平均分
        sh4['K2'] = int(len(list(nansheng)))
        sh4['L2'] = int(len(list(qinan)))
        sh4['M2'] = float(JiSuanPingJunFen(4, qinan, qinv)[1])
        sh4['S2'] = int(len(list(nvsheng)))
        sh4['T2'] = int(len(list(qinv)))
        sh4['U2'] = float(JiSuanPingJunFen(4, qinan, qinv)[2])
        # # 全部七项 跳绳    # 男生 女生 跳绳 平均分
        sh4['K4'] = int(len(list(nansheng)))
        sh4['L4'] = int(len(list(qinan)))
        sh4['M4'] = float(JiSuanPingJunFen(8, qinan, qinv)[1])
        sh4['S4'] = int(len(list(nvsheng)))
        sh4['T4'] = int(len(list(qinv)))
        sh4['U4'] = float(JiSuanPingJunFen(8, qinan, qinv)[2])
        # # 全部七项  坐位体前屈       # 男生 女生 坐位体前屈 平均分
        sh4['K6'] = int(len(list(nansheng)))
        sh4['L6'] = int(len(list(qinan)))
        sh4['M6'] = float(JiSuanPingJunFen(13, qinan, qinv)[1])
        sh4['S6'] = int(len(list(nvsheng)))
        sh4['T6'] = int(len(list(qinv)))
        sh4['U6'] = float(JiSuanPingJunFen(13, qinan, qinv)[2])
        # # 全部七项  仰卧起坐       # 男生 女生  平均分
        sh4['K8'] = int(len(list(nansheng)))
        sh4['L8'] = int(len(list(qinan)))
        sh4['M8'] = float(JiSuanPingJunFen(17, qinan, qinv)[1])
        sh4['S8'] = int(len(list(nvsheng)))
        sh4['T8'] = int(len(list(qinv)))
        sh4['U8'] = float(JiSuanPingJunFen(17, qinan, qinv)[2])
        # # 全部七项  立定跳远       # 男生 女生  平均分
        sh4['K10'] = int(len(list(nansheng)))
        sh4['L10'] = int(len(list(qinan)))
        sh4['M10'] = float(JiSuanPingJunFen(25, qinan, qinv)[1])
        sh4['S10'] = int(len(list(nvsheng)))
        sh4['T10'] = int(len(list(qinv)))
        sh4['U10'] = float(JiSuanPingJunFen(25, qinan, qinv)[2])
        # # 全部七项  肺活量       # 男生 女生  平均分
        sh4['K12'] = int(len(list(nansheng)))
        sh4['L12'] = int(len(list(qinan)))
        sh4['M12'] = float(JiSuanPingJunFen(33, qinan, qinv)[1])
        sh4['S12'] = int(len(list(nvsheng)))
        sh4['T12'] = int(len(list(qinv)))
        sh4['U12'] = float(JiSuanPingJunFen(33, qinan, qinv)[2])
        # # 全部七项  BMI        # 男生 女生  平均分
        sh4['K14'] = int(len(list(nansheng)))
        sh4['L14'] = int(len(list(qinan)))
        sh4['M14'] = float(JiSuanPingJunFen(39, qinan, qinv)[1])
        sh4['S14'] = int(len(list(nvsheng)))
        sh4['T14'] = int(len(list(qinv)))
        sh4['U14'] = float(JiSuanPingJunFen(39, qinan, qinv)[2])
        # # 全部七项 # 男生 女生  平均分
        sh4['K16'] = int(len(list(nansheng)))
        sh4['L16'] = int(len(list(qinan)))
        sh4['M16'] = float(JiSuanPingJunFen(43, qinan, qinv)[1])
        sh4['S16'] = int(len(list(nvsheng)))
        sh4['T16'] = int(len(list(qinv)))
        sh4['U16'] = float(JiSuanPingJunFen(43, qinan, qinv)[2])
        # #
        # # 全部七项  五十米    # 各等级总人数
        sh4['F3'] = JiSuanYouLiangLv(6, qinan, qinv)[0]
        sh4['G3'] = JiSuanYouLiangLv(6, qinan, qinv)[1]
        sh4['H3'] = JiSuanYouLiangLv(6, qinan, qinv)[2]
        sh4['I3'] = JiSuanYouLiangLv(6, qinan, qinv)[3]
        sh4['J3'] = JiSuanYouLiangLv(6, qinan, qinv)[4]
        # # 全部七项  跳绳      # 各等级总人数
        sh4['F5'] = JiSuanYouLiangLv(11, qinan, qinv)[0]
        sh4['G5'] = JiSuanYouLiangLv(11, qinan, qinv)[1]
        sh4['H5'] = JiSuanYouLiangLv(11, qinan, qinv)[2]
        sh4['I5'] = JiSuanYouLiangLv(11, qinan, qinv)[3]
        sh4['J5'] = JiSuanYouLiangLv(11, qinan, qinv)[4]
        # # 全部七项  坐位体前屈        # 各等级总人数
        sh4['F7'] = JiSuanYouLiangLv(15, qinan, qinv)[0]
        sh4['G7'] = JiSuanYouLiangLv(15, qinan, qinv)[1]
        sh4['H7'] = JiSuanYouLiangLv(15, qinan, qinv)[2]
        sh4['I7'] = JiSuanYouLiangLv(15, qinan, qinv)[3]
        sh4['J7'] = JiSuanYouLiangLv(15, qinan, qinv)[4]
        # # 全部七项  仰卧起坐        # 各等级总人数
        sh4['F9'] = JiSuanYouLiangLv(19, qinan, qinv)[0]
        sh4['G9'] = JiSuanYouLiangLv(19, qinan, qinv)[1]
        sh4['H9'] = JiSuanYouLiangLv(19, qinan, qinv)[2]
        sh4['I9'] = JiSuanYouLiangLv(19, qinan, qinv)[3]
        sh4['J9'] = JiSuanYouLiangLv(19, qinan, qinv)[4]
        # # 全部七项  立定跳远        # 各等级总人数
        sh4['F11'] = JiSuanYouLiangLv(27, qinan, qinv)[0]
        sh4['G11'] = JiSuanYouLiangLv(27, qinan, qinv)[1]
        sh4['H11'] = JiSuanYouLiangLv(27, qinan, qinv)[2]
        sh4['I11'] = JiSuanYouLiangLv(27, qinan, qinv)[3]
        sh4['J11'] = JiSuanYouLiangLv(27, qinan, qinv)[4]
        # # 全部七项  肺活量        # 各等级总人数
        sh4['F13'] = JiSuanYouLiangLv(35, qinan, qinv)[0]
        sh4['G13'] = JiSuanYouLiangLv(35, qinan, qinv)[1]
        sh4['H13'] = JiSuanYouLiangLv(35, qinan, qinv)[2]
        sh4['I13'] = JiSuanYouLiangLv(35, qinan, qinv)[3]
        sh4['J13'] = JiSuanYouLiangLv(35, qinan, qinv)[4]
        # # 全部七项  BMI        # 各等级总人数
        sh4['F15'] = JiSuanYouLiangLv(41, qinan, qinv)[0]
        sh4['G15'] = JiSuanYouLiangLv(41, qinan, qinv)[1]
        sh4['H15'] = JiSuanYouLiangLv(41, qinan, qinv)[2]
        sh4['I15'] = JiSuanYouLiangLv(41, qinan, qinv)[3]
        sh4['J15'] = JiSuanYouLiangLv(41, qinan, qinv)[4]
        # # 全部七项  总分        # 各等级总人数
        sh4['F17'] = JiSuanYouLiangLv(45, qinan, qinv)[0]
        sh4['G17'] = JiSuanYouLiangLv(45, qinan, qinv)[1]
        sh4['H17'] = JiSuanYouLiangLv(45, qinan, qinv)[2]
        sh4['I17'] = JiSuanYouLiangLv(45, qinan, qinv)[3]
        sh4['J17'] = JiSuanYouLiangLv(45, qinan, qinv)[4]
        #
        # # 全部七项  五十米        # # 总等级各种率
        sh4['F2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[5])
        sh4['G2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[6])
        sh4['H2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[7])
        sh4['I2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[8])
        sh4['J2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[9])
        # # 全部七项  跳绳        # # 总等级各种率
        sh4['F4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[5])
        sh4['G4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[6])
        sh4['H4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[7])
        sh4['I4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[8])
        sh4['J4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[9])
        # # 全部七项   坐位体前屈        # # 总等级各种率
        sh4['F6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[5])
        sh4['G6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[6])
        sh4['H6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[7])
        sh4['I6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[8])
        sh4['J6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[9])
        # # 全部七项   仰卧起坐        # # 总等级各种率
        sh4['F8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[5])
        sh4['G8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[6])
        sh4['H8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[7])
        sh4['I8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[8])
        sh4['J8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[9])
        # # 全部七项   立定跳远        # # 总等级各种率
        sh4['F10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[5])
        sh4['G10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[6])
        sh4['H10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[7])
        sh4['I10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[8])
        sh4['J10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[9])
        # # 全部七项   肺活量        # # 总等级各种率
        sh4['F12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[5])
        sh4['G12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[6])
        sh4['H12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[7])
        sh4['I12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[8])
        sh4['J12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[9])
        # # 全部七项   BMI        # # 总等级各种率
        sh4['F14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[5])
        sh4['G14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[6])
        sh4['H14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[7])
        sh4['I14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[8])
        sh4['J14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[9])
        # # 全部七项   总分        # # 总等级各种率
        sh4['F16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[5])
        sh4['G16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[6])
        sh4['H16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[7])
        sh4['I16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[8])
        sh4['J16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[9])
        #
        # # 全部七项  五十米        # # 男生50米  各等级人数
        sh4['N3'] = JiSuanYouLiangLv(6, qinan, qinv)[10]
        sh4['O3'] = JiSuanYouLiangLv(6, qinan, qinv)[11]
        sh4['P3'] = JiSuanYouLiangLv(6, qinan, qinv)[12]
        sh4['Q3'] = JiSuanYouLiangLv(6, qinan, qinv)[13]
        sh4['R3'] = JiSuanYouLiangLv(6, qinan, qinv)[14]
        # # 全部七项  跳绳        # # 男生跳绳  各等级人数
        sh4['N5'] = JiSuanYouLiangLv(11, qinan, qinv)[10]
        sh4['O5'] = JiSuanYouLiangLv(11, qinan, qinv)[11]
        sh4['P5'] = JiSuanYouLiangLv(11, qinan, qinv)[12]
        sh4['Q5'] = JiSuanYouLiangLv(11, qinan, qinv)[13]
        sh4['R5'] = JiSuanYouLiangLv(11, qinan, qinv)[14]
        # # 全部七项  坐位体前屈      # # 男生坐位体前屈  各等级人数
        sh4['N7'] = JiSuanYouLiangLv(15, qinan, qinv)[10]
        sh4['O7'] = JiSuanYouLiangLv(15, qinan, qinv)[11]
        sh4['P7'] = JiSuanYouLiangLv(15, qinan, qinv)[12]
        sh4['Q7'] = JiSuanYouLiangLv(15, qinan, qinv)[13]
        sh4['R7'] = JiSuanYouLiangLv(15, qinan, qinv)[14]
        # # 全部七项  仰卧起坐      # # 男生   各等级人数
        sh4['N9'] = JiSuanYouLiangLv(19, qinan, qinv)[10]
        sh4['O9'] = JiSuanYouLiangLv(19, qinan, qinv)[11]
        sh4['P9'] = JiSuanYouLiangLv(19, qinan, qinv)[12]
        sh4['Q9'] = JiSuanYouLiangLv(19, qinan, qinv)[13]
        sh4['R9'] = JiSuanYouLiangLv(19, qinan, qinv)[14]
        # # 全部七项  立定跳远      # # 男生   各等级人数
        sh4['N11'] = JiSuanYouLiangLv(27, qinan, qinv)[10]
        sh4['O11'] = JiSuanYouLiangLv(27, qinan, qinv)[11]
        sh4['P11'] = JiSuanYouLiangLv(27, qinan, qinv)[12]
        sh4['Q11'] = JiSuanYouLiangLv(27, qinan, qinv)[13]
        sh4['R11'] = JiSuanYouLiangLv(27, qinan, qinv)[14]
        # # 全部七项  肺活量      # # 男生   各等级人数
        sh4['N13'] = JiSuanYouLiangLv(35, qinan, qinv)[10]
        sh4['O13'] = JiSuanYouLiangLv(35, qinan, qinv)[11]
        sh4['P13'] = JiSuanYouLiangLv(35, qinan, qinv)[12]
        sh4['Q13'] = JiSuanYouLiangLv(35, qinan, qinv)[13]
        sh4['R13'] = JiSuanYouLiangLv(35, qinan, qinv)[14]
        # # 全部七项  BMI      # # 男生  各等级人数
        sh4['N15'] = JiSuanYouLiangLv(41, qinan, qinv)[10]
        sh4['O15'] = JiSuanYouLiangLv(41, qinan, qinv)[11]
        sh4['P15'] = JiSuanYouLiangLv(41, qinan, qinv)[12]
        sh4['Q15'] = JiSuanYouLiangLv(41, qinan, qinv)[13]
        sh4['R15'] = JiSuanYouLiangLv(41, qinan, qinv)[14]
        # # 全部七项  总分      # # 男生  各等级人数
        sh4['N17'] = JiSuanYouLiangLv(45, qinan, qinv)[10]
        sh4['O17'] = JiSuanYouLiangLv(45, qinan, qinv)[11]
        sh4['P17'] = JiSuanYouLiangLv(45, qinan, qinv)[12]
        sh4['Q17'] = JiSuanYouLiangLv(45, qinan, qinv)[13]
        sh4['R17'] = JiSuanYouLiangLv(45, qinan, qinv)[14]
        # #
        # # 全部七项   五十米     # # 男生  50米各等级的率
        sh4['N2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[15])
        sh4['O2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[16])
        sh4['P2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[17])
        sh4['Q2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[18])
        sh4['R2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[19])
        # # 全部七项  跳绳        # # 男生  跳绳各等级的率
        sh4['N4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[15])
        sh4['O4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[16])
        sh4['P4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[17])
        sh4['Q4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[18])
        sh4['R4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[19])
        # # 全部七项   坐位体前屈        # # 男生  坐位体前屈各等级的率
        sh4['N6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[15])
        sh4['O6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[16])
        sh4['P6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[17])
        sh4['Q6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[18])
        sh4['R6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[19])
        # # 全部七项   仰卧起坐        # # 男生    各等级的率
        sh4['N8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[15])
        sh4['O8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[16])
        sh4['P8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[17])
        sh4['Q8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[18])
        sh4['R8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[19])
        # # 全部七项   立定跳远        # # 男生    各等级的率
        sh4['N10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[15])
        sh4['O10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[16])
        sh4['P10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[17])
        sh4['Q10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[18])
        sh4['R10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[19])
        # # 全部七项   肺活量        # # 男生    各等级的率
        sh4['N12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[15])
        sh4['O12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[16])
        sh4['P12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[17])
        sh4['Q12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[18])
        sh4['R12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[19])
        # # 全部七项   BMI        # # 男生    各等级的率
        sh4['N14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[15])
        sh4['O14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[16])
        sh4['P14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[17])
        sh4['Q14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[18])
        sh4['R14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[19])
        # # 全部七项   总分        # # 男生    各等级的率
        sh4['N16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[15])
        sh4['O16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[16])
        sh4['P16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[17])
        sh4['Q16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[18])
        sh4['R16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[19])
        #
        # # 全部七项   五十米      # # 女生50米  各等级人数
        sh4['V3'] = JiSuanYouLiangLv(6, qinan, qinv)[20]
        sh4['W3'] = JiSuanYouLiangLv(6, qinan, qinv)[21]
        sh4['X3'] = JiSuanYouLiangLv(6, qinan, qinv)[22]
        sh4['Y3'] = JiSuanYouLiangLv(6, qinan, qinv)[23]
        sh4['Z3'] = JiSuanYouLiangLv(6, qinan, qinv)[24]
        # # 全部七项   跳绳      # # 女生跳绳  各等级人数
        sh4['V5'] = JiSuanYouLiangLv(11, qinan, qinv)[20]
        sh4['W5'] = JiSuanYouLiangLv(11, qinan, qinv)[21]
        sh4['X5'] = JiSuanYouLiangLv(11, qinan, qinv)[22]
        sh4['Y5'] = JiSuanYouLiangLv(11, qinan, qinv)[23]
        sh4['Z5'] = JiSuanYouLiangLv(11, qinan, qinv)[24]
        # # 全部七项  坐位体前屈    # # 女生坐位体前屈  各等级人数
        sh4['V7'] = JiSuanYouLiangLv(15, qinan, qinv)[20]
        sh4['W7'] = JiSuanYouLiangLv(15, qinan, qinv)[21]
        sh4['X7'] = JiSuanYouLiangLv(15, qinan, qinv)[22]
        sh4['Y7'] = JiSuanYouLiangLv(15, qinan, qinv)[23]
        sh4['Z7'] = JiSuanYouLiangLv(15, qinan, qinv)[24]
        # # 全部七项  仰卧起坐    # # 女生   各等级人数
        sh4['V9'] = JiSuanYouLiangLv(19, qinan, qinv)[20]
        sh4['W9'] = JiSuanYouLiangLv(19, qinan, qinv)[21]
        sh4['X9'] = JiSuanYouLiangLv(19, qinan, qinv)[22]
        sh4['Y9'] = JiSuanYouLiangLv(19, qinan, qinv)[23]
        sh4['Z9'] = JiSuanYouLiangLv(19, qinan, qinv)[24]
        # # 全部七项  立定跳远    # # 女生   各等级人数
        sh4['V11'] = JiSuanYouLiangLv(27, qinan, qinv)[20]
        sh4['W11'] = JiSuanYouLiangLv(27, qinan, qinv)[21]
        sh4['X11'] = JiSuanYouLiangLv(27, qinan, qinv)[22]
        sh4['Y11'] = JiSuanYouLiangLv(27, qinan, qinv)[23]
        sh4['Z11'] = JiSuanYouLiangLv(27, qinan, qinv)[24]
        # # 全部七项  肺活量    # # 女生   各等级人数
        sh4['V13'] = JiSuanYouLiangLv(35, qinan, qinv)[20]
        sh4['W13'] = JiSuanYouLiangLv(35, qinan, qinv)[21]
        sh4['X13'] = JiSuanYouLiangLv(35, qinan, qinv)[22]
        sh4['Y13'] = JiSuanYouLiangLv(35, qinan, qinv)[23]
        sh4['Z13'] = JiSuanYouLiangLv(35, qinan, qinv)[24]
        # # 全部七项  BMI     # # 女生    各等级人数
        sh4['V15'] = JiSuanYouLiangLv(41, qinan, qinv)[20]
        sh4['W15'] = JiSuanYouLiangLv(41, qinan, qinv)[21]
        sh4['X15'] = JiSuanYouLiangLv(41, qinan, qinv)[22]
        sh4['Y15'] = JiSuanYouLiangLv(41, qinan, qinv)[23]
        sh4['Z15'] = JiSuanYouLiangLv(41, qinan, qinv)[24]
        # # 全部七项       # # 女生    各等级人数
        sh4['V17'] = JiSuanYouLiangLv(45, qinan, qinv)[20]
        sh4['W17'] = JiSuanYouLiangLv(45, qinan, qinv)[21]
        sh4['X17'] = JiSuanYouLiangLv(45, qinan, qinv)[22]
        sh4['Y17'] = JiSuanYouLiangLv(45, qinan, qinv)[23]
        sh4['Z17'] = JiSuanYouLiangLv(45, qinan, qinv)[24]
        #
        # # 全部七项   五十米        # # 女生50米各等级的率
        sh4['V2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[25])
        sh4['W2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[26])
        sh4['X2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[27])
        sh4['Y2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[28])
        sh4['Z2'] = float(JiSuanYouLiangLv(6, qinan, qinv)[29])
        # # 全部七项   跳绳  # # 女生跳绳各等级的率
        sh4['V4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[25])
        sh4['W4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[26])
        sh4['X4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[27])
        sh4['Y4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[28])
        sh4['Z4'] = float(JiSuanYouLiangLv(11, qinan, qinv)[29])
        # # 全部七项  坐位体前屈    # # 女生坐位体前屈各等级的率
        sh4['V6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[25])
        sh4['W6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[26])
        sh4['X6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[27])
        sh4['Y6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[28])
        sh4['Z6'] = float(JiSuanYouLiangLv(15, qinan, qinv)[29])
        # # 全部七项  仰卧起坐    # # 女生   各等级的率
        sh4['V8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[25])
        sh4['W8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[26])
        sh4['X8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[27])
        sh4['Y8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[28])
        sh4['Z8'] = float(JiSuanYouLiangLv(19, qinan, qinv)[29])
        # # 全部七项  立定跳远    # # 女生   各等级的率
        sh4['V10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[25])
        sh4['W10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[26])
        sh4['X10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[27])
        sh4['Y10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[28])
        sh4['Z10'] = float(JiSuanYouLiangLv(27, qinan, qinv)[29])
        # #
        # # 全部七项  肺活量    # # 女生   各等级的率
        sh4['V12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[25])
        sh4['W12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[26])
        sh4['X12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[27])
        sh4['Y12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[28])
        sh4['Z12'] = float(JiSuanYouLiangLv(35, qinan, qinv)[29])
        # # 全部七项  BMI    # # 女生   各等级的率
        sh4['V14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[25])
        sh4['W14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[26])
        sh4['X14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[27])
        sh4['Y14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[28])
        sh4['Z14'] = float(JiSuanYouLiangLv(41, qinan, qinv)[29])
        # # 全部七项  总分    # # 女生   各等级的率
        sh4['V16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[25])
        sh4['W16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[26])
        sh4['X16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[27])
        sh4['Y16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[28])
        sh4['Z16'] = float(JiSuanYouLiangLv(45, qinan, qinv)[29])
        # #
        sh4['A2'] = "50米"
        sh4['A3'] = "人数"
        sh4['A4'] = "跳绳"
        sh4['A5'] = "人数"
        sh4['A6'] = "坐位体前屈"
        sh4['A7'] = "人数"
        sh4['A8'] = "仰卧起坐"
        sh4['A9'] = "人数"
        sh4['A10'] = "立定跳远"
        sh4['A11'] = "人数"
        sh4['A12'] = "肺活量"
        sh4['A13'] = "人数"
        sh4['A14'] = "BMI"
        sh4['A15'] = "人数"
        sh4['A16'] = "总分七项\n（100分）"
        sh4['A17'] = "人数"
        # #
        sh4['B1'] = "应测人数"
        sh4['C1'] = "实测人数"
        sh4['D1'] = "平均分"
        sh4['E1'] = "名次"
        sh4['F1'] = "优秀率"
        sh4['G1'] = "良好率"
        sh4['H1'] = "优良率"
        sh4['I1'] = "及格率"
        sh4['J1'] = "不及格率"
        sh4['K1'] = "男生\n应测人数"
        sh4['L1'] = "男生\n实测人数"
        sh4['M1'] = "男生\n平均分"
        sh4['N1'] = "男生\n优秀率"
        sh4['O1'] = "男生\n良好率"
        sh4['P1'] = "男生\n优良率"
        sh4['Q1'] = "男生\n及格率"
        sh4['R1'] = "男生\n不及格率"
        sh4['S1'] = "女生\n应测人数"
        sh4['T1'] = "女生\n实测人数"
        sh4['U1'] = "女生\n平均分"
        sh4['V1'] = "女生\n优秀率"
        sh4['W1'] = "女生\n良好率"
        sh4['X1'] = "女生\n优良率"
        sh4['Y1'] = "女生\n及格率"
        sh4['Z1'] = "女生\n不及格率"

        # 同时固定前4行（A1~A4）和前三列（A~C），冻结位置设为 D5
        sh4.freeze_panes = 'B2'

        # 全部内容居中  四项成绩
        for row in sh4.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # 添加边框
                cell.border = all_border

        # 调整宽度
        kuanhao = ['A', 'B', 'C', 'J', 'K', 'L', 'R', 'S', 'T', 'Z']
        for i in kuanhao:
            sh4.column_dimensions[i].width = 12
        #
        # 定义百分数格式（保留两位小数）
        baifenshu = numbers.FORMAT_PERCENTAGE_00  # 格式为“85.00%”
        # 区域1：男生/女生各项目的率（如优秀率、良好率等）
        xuhao = [2, 4, 6, 8, 10, 12,14,16]
        # 区域1：优秀率、良好率
        for i in xuhao:
            # 四项成绩
            for row in sh4.iter_rows(min_row=i, max_row=i, min_col=5, max_col=10):
                for cell in row:
                    cell.number_format = baifenshu
        # 区域2：男生/女生各项目的率（如优秀率、良好率等）
        for i in xuhao:
            for row in sh4.iter_rows(min_row=i, max_row=i, min_col=14, max_col=18):
                for cell in row:
                    cell.number_format = baifenshu
        # 区域3：男生/女生各项目的率（如优秀率、良好率等）
        for i in xuhao:
            for row in sh4.iter_rows(min_row=i, max_row=i, min_col=22, max_col=26):
                for cell in row:
                    cell.number_format = baifenshu

        # 步骤1：创建一个字体对象，设置字体大小为 20（其他属性保持默认，如字体、加粗等不变）
        target_font = Font(size=13)  # size 参数指定字体大小，这里设为 20
        # 步骤2：遍历工作表中所有有内容的单元格（高效遍历）
        # 若想遍历整个工作表（包括空单元格），可直接用 sh1.iter_rows() 无参数
        for row in sh4.iter_rows(min_row=1, max_row=sh4.max_row, min_col=1, max_col=sh4.max_column):
            for cell in row:
                # 步骤3：给单元格应用预设的字体对象
                cell.font = target_font
        # ================添加柱状图=================
        sinianji_qixiang_zhuzhuangtu(sh4)
        # 循环设置 B20~D25 居中 + 边框
        for row in sh4['B20:D25']:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  # 水平+垂直居中
                cell.border = all_border  # 添加边框

        # ==================添加训练策略=======================
        qi_xunlianfangan_plan(sh4)
        # 保存修改后的Excel文件
        wb.save(new_file_path)
        wb.close()  # 关闭文件，释放资源

        print(f"四年级文件路径：{new_file_path}")
        return new_file_path


    except Exception as e:
        error_msg = str(e)
        # if "row" in error_msg and "column" in error_msg:
        #     raise Exception(
        #         "四年级数据处理失败：Excel表格内容错误，请检查【成绩明细】工作表中的单元格数据是否异常！<br>"
        #         f"错误详情：{error_msg}"
        #     )
        if "index" in error_msg:
            raise Exception(
                "四年级数据处理失败：Excel行数/列数不匹配，请检查表格是否少行、少列、格式被修改！"
            )

        else:
            raise Exception(
                f"四年级数据处理失败：Excel表格格式或数据异常，请检查文件是否正确！<br>"
                f"错误信息：{error_msg}<br><br>"
                "常见错误代码对照：<br>"
                "错误信息1. '>' not supported between instances of 'str' and 'int' → 跳绳数据有误或有空单元格<br>"
                "错误信息2. can only concatenate str (not 'int') to str →→→→→→ 50米数据有误或有空单元格<br>"
                "错误信息3. cannot convert float NaN to integer→→→→→→→→数据内有多余序号或数据外有数字或字符<br>"
                "错误信息4. could not convert string to float: '字符'→→→→→→→数据内有非法字符<br>"
                "错误信息5. 'int' object is not subscriptable →→→→→→→→→→ 坐位体前屈数据有空单元格或超出范围<br>"
                "错误信息6. invalid literal for int() with base 10: 'nan' →→→→→→→→→→ 仰卧起坐数据有空单元格<br>"
                "错误信息7. local variable 'e' referenced before assignment →→→ 跳绳、仰卧起坐数据有小数情况<br>"
                "错误信息8. unsupported operand type(s) for +: 'int' and 'str'→→→ 坐位体前屈、肺活量数据有误<br>"
                "错误信息9. 其他情况请联系管理员,及时添加更多提示信息"

            )


