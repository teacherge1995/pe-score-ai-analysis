from flask import Flask, request, send_file, render_template_string, jsonify, redirect, url_for, session
import os
import pandas as pd
from datetime import datetime
from werkzeug.utils import secure_filename
import time
import threading
import queue
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, numbers
import uuid
from z一键清空 import clean_target_folders

# 导入原有年级处理函数  按索引5排序汇总表
from bYiNianJi_BiaoDan import yinianji_chuliwenjian
from cErNianJi_BiaoDan import ernianji_chuliwenjian
from dSanNianJi_BiaoDan import sannianji_chuliwenjian
from eSiNianJi_BiaoDan import sinianji_chuliwenjian
from fWuNianJi_BiaoDan import wunianji_chuliwenjian
from gLiuNianJi_BiaoDan import liunianji_chuliwenjian

# 导入各年级批量排序处理文件（和app.py同目录）
from byinianji_shiwuming import yinianji_sort_by_index5
from cernianji_shiwuming import ernianji_sort_by_index5
from dsannianji_shiwuming import sannianji_sort_by_index5
from esinianji_shiwuming import sinianji_sort_by_index5
from fwunianji_shiwuming import wunianji_sort_by_index5
from gliunianji_shiwuming import liunianji_sort_by_index5

# 导入模板生成模块
from b一年级下载模板 import generate_1_grade_template
from c二年级下载模板 import generate_2_grade_template
from d三年级下载模板 import generate_3_grade_template
from e四年级下载模板 import generate_4_grade_template
from f五年级下载模板 import generate_5_grade_template
from g六年级下载模板 import generate_6_grade_template

# ========== 基础配置 ==========
app = Flask(__name__)
app.secret_key = 'sportsystem_secure_key_2025'  # 登录会话密钥
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
DOWNLOAD_FOLDER = os.path.join(BASE_DIR, 'downloads')
TEMPLATE_FOLDER = os.path.join(BASE_DIR, 'templates')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
os.makedirs(TEMPLATE_FOLDER, exist_ok=True)

# ========== 固定账号密码 ==========
ADMIN_USER = "sportsystem"
ADMIN_PWD = "sportv2.0"

# ========== 全局状态：原有单文件成绩处理 ==========
file_queue = queue.Queue()
current_processing = ""
current_file_start_time = None
task_start_time = None
processed_files = []
current_grade = ""

# ========== 全局状态：批量按索引5排序（新增） ==========
batch_sort_queue = queue.Queue()
batch_current_processing = ""
batch_current_file_start = None
batch_task_start = None
batch_processed_files = []
batch_current_grade = ""


# ========== 登录拦截装饰器 ==========
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


# ========== 登录页面 ==========
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        login_html = '''
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <title>系统登录 - 体育成绩批处理智能分析系统</title>
            <style>
                * {margin:0;padding:0;box-sizing:border-box;font-family:微软雅黑;}
                body {background:#f5f7fa;display:flex;justify-content:center;align-items:center;height:100vh;}
                .login-box {width:400px;background:white;padding:40px;border-radius:12px;box-shadow:0 0 20px rgba(0,0,0,0.1);}
                .login-box h2 {text-align:center;margin-bottom:30px;color:#333;}
                .form-item {margin-bottom:20px;}
                .form-item label {display:block;margin-bottom:8px;font-weight:bold;color:#555;}
                .form-item input {width:100%;padding:12px;border:1px solid #ddd;border-radius:6px;font-size:16px;}
                .login-btn {width:100%;padding:12px;background:#007bff;color:white;border:none;border-radius:6px;font-size:16px;cursor:pointer;}
                .login-btn:hover {background:#0056b3;}
                .error {color:red;text-align:center;margin-bottom:15px;}
            </style>
        </head>
        <body>
            <div class="login-box">
                <h2>体育成绩批处理智能分析系统</h2>
                <form method="post">
                    <div class="form-item">
                        <label>账号</label>
                        <input type="text" name="username" placeholder="请输入账号" required>
                    </div>
                    <div class="form-item">
                        <label>密码</label>
                        <input type="password" name="password" placeholder="请输入密码" required>
                    </div>
                    <button type="submit" class="login-btn">登录</button>
                </form>
            </div>
        </body>
        </html>
        '''
        return render_template_string(login_html)
    else:
        username = request.form.get('username')
        password = request.form.get('password')
        if username == ADMIN_USER and password == ADMIN_PWD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            return render_template_string('''
            <!DOCTYPE html>
            <html>
            <body style="display:flex;justify-content:center;align-items:center;height:100vh;">
                <div style="color:red;font-size:18px;">账号或密码错误！<a href="/login">返回登录</a></div>
            </body>
            </html>
            ''')


# ========== 退出登录 ==========
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ========== 原有各年级成绩处理函数 ==========
def process_grade_1(df, origin_name):
    return yinianji_chuliwenjian(df, origin_name)


def process_grade_2(df, origin_name):
    return ernianji_chuliwenjian(df, origin_name)


def process_grade_3(df, origin_name):
    return sannianji_chuliwenjian(df, origin_name)


def process_grade_4(df, origin_name):
    return sinianji_chuliwenjian(df, origin_name)


def process_grade_5(df, origin_name):
    return wunianji_chuliwenjian(df, origin_name)


def process_grade_6(df, origin_name):
    return liunianji_chuliwenjian(df, origin_name)


# ========== 批量前15名排序核心处理函数（修复变量作用域） ==========
def batch_process_by_grade(grade, file_path_list, origin_name_list):
    all_rows = []
    all_rows_sorted = all_rows  # 提前定义变量，避免作用域错误
    # 读取所有上传的Excel文件并合并
    for fp in file_path_list:
        try:
            if fp.endswith('.xlsx'):
                df = pd.read_excel(fp, engine='openpyxl', header=None)
            else:
                df = pd.read_excel(fp, engine='xlrd', header=None)
            rows = df.fillna('').values.tolist()
            all_rows.extend(rows)
        except Exception as e:
            print(f"读取文件{fp}失败：{str(e)}")
            continue
    # 调用对应年级的排序函数
    try:
        if grade == '1':
            all_rows_sorted = yinianji_sort_by_index5(all_rows)
        elif grade == '2':
            all_rows_sorted = ernianji_sort_by_index5(all_rows)
        elif grade == '3':
            all_rows_sorted = sannianji_sort_by_index5(all_rows)
        elif grade == '4':
            all_rows_sorted = sinianji_sort_by_index5(all_rows)
        elif grade == '5':
            all_rows_sorted = wunianji_sort_by_index5(all_rows)
        elif grade == '6':
            all_rows_sorted = liunianji_sort_by_index5(all_rows)
    except Exception as e:
        print(f"{grade}年级调用排序函数失败：{str(e)}")
        all_rows_sorted = all_rows
    # 保存排序后的汇总文件
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    out_name = f"{grade}年级_前15名排序汇总_{ts}.xlsx"
    out_path = os.path.join(DOWNLOAD_FOLDER, out_name)
    try:
        biaotou_list = ['序号', '性别', '项目', '排序', '班级', '姓名', '成绩']
        biaotou = pd.DataFrame([biaotou_list])
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            biaotou.to_excel(writer, sheet_name='成绩明细', index=False, header=False, startrow=0)
            pd.DataFrame(all_rows_sorted).to_excel(writer,
                                                   sheet_name='成绩明细', index=False, header=None, startrow=1,
                                                   engine='openpyxl')
        wb = load_workbook(out_path)
        sh1 = wb['成绩明细']
        zuida_liehao = sh1.max_column + 1  # 获取工作表最大列号，自适应所有列，避免遗漏
        # 获取工作表实际有数据/样式的最大行号
        zuida_hanghao = sh1.max_row + 1
        # ==================================添加边框=======================================
        all_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        # 1. 定义================================居中对齐样式================================（只定义1次，重复使用，提升效率）
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # 双层循环：遍历所有行和所有列的单元格
        for row in sh1.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
                cell.border = all_border
            # -------- 5.3 可选：调整行高（让表头更美观） --------
        sh1.row_dimensions[1].height = 25  # 第一行高度
        for row_num in range(2, zuida_hanghao):
            sh1.row_dimensions[row_num].height = 20
        # 同时固定前4行（A1~A4）和前三列（A~C），冻结位置设为 D5
        sh1.freeze_panes = 'A2'
        # 加宽单元格
        sh1.column_dimensions['C'].width = 13

        # 步骤1：创建一个字体对象，设置字体大小为 20（其他属性保持默认，如字体、加粗等不变）
        target_font = Font(size=13)  # size 参数指定字体大小，这里设为 20
        # 步骤2：遍历工作表中所有有内容的单元格（高效遍历）
        # 若想遍历整个工作表（包括空单元格），可直接用 sh1.iter_rows() 无参数
        for row in sh1.iter_rows(min_row=1, max_row=sh1.max_row, min_col=1, max_col=sh1.max_column):
            for cell in row:
                # 步骤3：给单元格应用预设的字体对象
                cell.font = target_font
        # 2. 第一行字体加粗（核心逻辑）
        bold_font = Font(bold=True)  # 定义加粗字体样式
        for cell in sh1[1]:  # sh1[1] 代表工作表的**第一行**所有单元格（openpyxl行号从1开始）
            cell.font = bold_font  # 为第一行每个单元格应用加粗样式

        # ========================男女分界线===专用粗边框===========================（核心：style = 'medium'        实现边框加粗）
        # bold_side = Side(style='medium', color='black')  # medium=中粗实线，也可用'thick'=粗实线
        bold_borde = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='medium'), bottom=Side(style='thin'))
        hangxuhao = []
        i = 0  # 初始值（你代码里隐含的起始数，可根据需要修改）
        while i < zuida_hanghao - 17:  # 满足小于121时继续循环
            i = i + 15  # 每次加15
            hangxuhao.append(i + 2)  # 加2后存入列表

        # hangxuhao1 =[17,32,47,62,77,92,107]
        # 2. 单独处理第23行：遍历所有列，应用粗边框
        for col_num in range(1, zuida_liehao):
            for i in hangxuhao:
                # 定位第23行第col_num列的单元格
                current_cell = sh1.cell(row=i, column=col_num)
                # 应用粗边框样式
                current_cell.border = bold_borde
        # 保存修改后的Excel文件
        wb.save(out_path)
        wb.close()  # 关闭文件，释放资源

    except Exception as e:
        print(f"保存{grade}年级排序结果失败：{str(e)}")
        raise Exception(f"文件保存失败：{str(e)}")
    return out_path, out_name


# ========== 原有成绩处理后台线程 ==========
def process_thread():
    global current_processing, current_file_start_time, processed_files, current_grade
    while True:
        if not file_queue.empty():
            upload_path, origin_name, grade = file_queue.get()
            current_processing = origin_name
            current_grade = grade
            current_file_start_time = datetime.now()
            try:
                # 读取Excel文件
                if upload_path.lower().endswith('.xlsx'):
                    df = pd.read_excel(upload_path, engine='openpyxl')
                elif upload_path.lower().endswith('.xls'):
                    df = pd.read_excel(upload_path, engine='xlrd')
                else:
                    raise Exception("仅支持.xlsx和.xls格式文件")
                # 调用对应年级处理函数
                save_path = None
                save_filename = None
                if grade == "1":
                    save_path = process_grade_1(df, origin_name)
                    save_filename = os.path.basename(save_path)
                elif grade == "2":
                    save_path = process_grade_2(df, origin_name)
                    save_filename = os.path.basename(save_path)
                elif grade == "3":
                    save_path = process_grade_3(df, origin_name)
                    save_filename = os.path.basename(save_path)
                elif grade == "4":
                    save_path = process_grade_4(df, origin_name)
                    save_filename = os.path.basename(save_path)
                elif grade == "5":
                    save_path = process_grade_5(df, origin_name)
                    save_filename = os.path.basename(save_path)
                elif grade == "6":
                    save_path = process_grade_6(df, origin_name)
                    save_filename = os.path.basename(save_path)
                else:
                    grade_func_map = {"10": process_grade_6}
                    process_func = grade_func_map.get(grade, lambda x: x)
                    df_processed = process_func(df)
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S%f')
                    safe_name = secure_filename(origin_name)
                    save_filename = f"{grade}年级体育成绩单_{timestamp}_{safe_name}"
                    save_path = os.path.join(DOWNLOAD_FOLDER, save_filename)
                    df_processed.to_excel(save_path, index=False, engine='openpyxl')
                # 记录成功信息（保留每个文件耗时）
                cost_time = round((datetime.now() - current_file_start_time).total_seconds(), 2)
                processed_files.append({
                    "name": origin_name, "url": f"/download/{save_filename}",
                    "cost_time": cost_time, "grade": grade, "error": ""
                })
            except Exception as e:
                # 记录失败信息（保留每个文件耗时）
                cost_time = round((datetime.now() - current_file_start_time).total_seconds(), 2)
                processed_files.append({
                    "name": origin_name, "url": "", "cost_time": cost_time,
                    "grade": grade, "error": f"处理失败：{str(e)}"
                })
            finally:
                current_processing = ""
                current_grade = ""
                current_file_start_time = None
                file_queue.task_done()
        time.sleep(0.1)


# ========== 批量排序后台线程（新增） ==========
def batch_sort_thread():
    global batch_current_processing, batch_current_file_start, batch_processed_files, batch_current_grade
    while True:
        if not batch_sort_queue.empty():
            file_path_list, origin_name_list, grade = batch_sort_queue.get()
            batch_current_grade = grade
            batch_task_start = datetime.now()
            try:
                out_path, out_name = batch_process_by_grade(grade, file_path_list, origin_name_list)
                cost = round((datetime.now() - batch_task_start).total_seconds(), 2)
                batch_processed_files.append({
                    "grade": grade, "file_count": len(file_path_list),
                    "url": f"/download/{out_name}", "cost_time": cost, "error": ""
                })
            except Exception as e:
                cost = round((datetime.now() - batch_task_start).total_seconds(), 2)
                batch_processed_files.append({
                    "grade": grade, "file_count": len(file_path_list),
                    "url": "", "cost_time": cost, "error": str(e)
                })
            finally:
                batch_current_processing = ""
                batch_current_grade = ""
                batch_sort_queue.task_done()
        time.sleep(0.1)


# 启动两个后台守护线程
threading.Thread(target=process_thread, daemon=True).start()
threading.Thread(target=batch_sort_thread, daemon=True).start()


# ========== 模板生成接口 ==========
@app.route('/generate_template', methods=['POST'])
@login_required
def generate_template():
    try:
        data = request.json
        school_name = data.get('school_name', '')
        header_name = data.get('header_name', '')
        grade = data.get('grade', '')
        selected_items = data.get('selected_items', [])
        boy_count = int(data.get('boy_count', 0))
        girl_count = int(data.get('girl_count', 0))
        file_name = data.get('file_name', '体育成绩模板')

        # 生成模板文件
        ts = datetime.now().strftime('%Y%m%d%H%M%S')
        template_filename = f"{file_name}.xlsx"
        template_path = os.path.join(TEMPLATE_FOLDER, template_filename)

        if grade == '1':
            generate_1_grade_template(school_name, header_name, selected_items, boy_count, girl_count, template_path,
                                      file_name)
        elif grade == '2':
            generate_2_grade_template(school_name, header_name, selected_items, boy_count, girl_count, template_path,
                                      file_name)
        elif grade == '3':
            generate_3_grade_template(school_name, header_name, selected_items, boy_count, girl_count, template_path,
                                      file_name)
        elif grade == '4':
            generate_4_grade_template(school_name, header_name, selected_items, boy_count, girl_count, template_path,
                                      file_name)
        elif grade == '5':
            generate_5_grade_template(school_name, header_name, selected_items, boy_count, girl_count, template_path,
                                      file_name)
        elif grade == '6':
            generate_6_grade_template(school_name, header_name, selected_items, boy_count, girl_count, template_path,
                                      file_name)
        else:
            return jsonify({"code": 400, "msg": "无效的年级选择"}), 400

        return jsonify({
            "code": 200,
            "msg": "模板生成成功",
            "download_url": f"/download_template/{template_filename}"
        })
    except Exception as e:
        return jsonify({"code": 500, "msg": f"模板生成失败：{str(e)}"}), 500


# ========== 模板下载接口 ==========
@app.route('/download_template/<filename>')
@login_required
def download_template(filename):
    template_path = os.path.join(TEMPLATE_FOLDER, filename)
    if not os.path.exists(template_path):
        return "模板文件不存在", 404
    return send_file(
        template_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ========== 前端首页 ==========
@app.route('/')
@login_required
def index():
    # 重置所有全局状态
    global current_processing, current_file_start_time, task_start_time, processed_files, current_grade
    global batch_current_processing, batch_current_file_start, batch_task_start, batch_processed_files, batch_current_grade
    # 重置原有处理状态
    current_processing = ""
    current_file_start_time = None
    task_start_time = None
    processed_files = []
    current_grade = ""
    # 重置批量排序状态
    batch_current_processing = ""
    batch_current_file_start = None
    batch_task_start = None
    batch_processed_files = []
    batch_current_grade = ""
    # 清空所有队列
    while not file_queue.empty():
        try:
            file_queue.get_nowait()
        except queue.Empty:
            break
    while not batch_sort_queue.empty():
        try:
            batch_sort_queue.get_nowait()
        except queue.Empty:
            break

    # 前端页面HTML（已添加模板下载功能 + 页面切换 + 年级对应项目限制）
    html_template = '''
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <title>体育成绩批处理智能分析系统</title>
        <style>
            * {margin: 0; padding: 0; box-sizing: border-box; font-family: Arial, 微软雅黑;}
            .container {width: 900px; margin: 50px auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.05);}
            h1 {text-align: center; color: #333; margin-bottom: 30px;}
            .nav {text-align: center; margin-bottom: 20px;}
            .nav a {padding: 8px 16px; margin: 0 5px; background: #6c757d; color: white; border-radius: 4px; text-decoration: none;}
            .nav a.active {background: #007bff;}
            .logout {text-align:center; margin-bottom:15px;}
            .logout a {color:red; text-decoration:none;}
            .template-section {margin-bottom: 20px; padding: 15px; background: #f0f8ff; border-radius: 6px;}
            .template-btn {padding: 10px 25px; background: #17a2b8; color: white; border: none; border-radius: 4px; cursor: pointer; transition: background 0.2s; font-size: 16px;}
            .template-btn:hover {background: #138496;}
            .modal {display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.5); z-index: 1000; justify-content: center; align-items: center;}
            .modal-content {background: white; padding: 25px; border-radius: 8px; width: 500px; max-width: 90%;}
            .modal-step {display: none;}
            .modal-step.active {display: block;}
            .form-group {margin: 15px 0;}
            label {display: block; margin-bottom: 5px; font-weight: bold;}
            input[type="text"], input[type="number"] {width: 100%; padding: 8px; border: 1px solid #ddd; border-radius: 4px;}
            .checkbox-group {display: flex; flex-wrap: wrap; gap: 10px; margin: 10px 0;}
            .checkbox-item {display: flex; align-items: center; width: 30%;}
            .btn-group {display: flex; justify-content: space-between; margin-top: 20px;}
            .btn {padding: 8px 20px; border: none; border-radius: 4px; cursor: pointer;}
            .btn-next {background: #007bff; color: white;}
            .btn-prev {background: #6c757d; color: white;}
            .btn-submit {background: #28a745; color: white;}
            .file-input {width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 4px; margin-bottom: 20px;}
            .grade-buttons {display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 20px;}
            .grade-btn {padding: 10px 20px; background: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; transition: background 0.2s;}
            .grade-btn:hover {background: #0056b3;}
            .batch-btn {padding: 10px 20px; background: #28a745; color: white; border: none; border-radius: 4px; cursor: pointer; transition: background 0.2s;}
            .batch-btn:hover {background: #1e7e34;}
            .section-title {margin: 15px 0 8px; font-size: 16px; color: #333; font-weight: bold;}
            .file-list {padding: 10px; background: #f8f9fa; border-radius: 4px; margin-bottom: 20px;}
            .file-item {margin: 5px 0; color: #555;}
            .status {padding: 10px; border-radius: 4px; font-weight: bold; margin-bottom: 10px;}
            .status-wait {background: #fff3cd; color: #856404;}
            .status-processing {background: #d1ecf1; color: #0c5460;}
            .status-done {background: #d4edda; color: #155724;}
            .status-error {background: #f8d7da; color: #721c24;}
            .time-info {color: #666; font-size: 14px; margin-bottom: 20px;}
            .download-area {display: none; padding: 15px; background: #f8f9fa; border-radius: 4px; margin-bottom: 10px;}
            .batch-download-area {display: none; padding: 15px; background: #f8f9fa; border-radius: 4px;}
            .download-item {margin: 8px 0;}
            .download-item a {color: #007bff; text-decoration: none;}
            .download-item a:hover {text-decoration: underline;}
            .error-item {color: #dc3545; margin: 8px 0;}
            .cost-time {color: #666; font-size: 12px; margin-left: 10px;}
            .template-download-link {margin-top: 20px; padding: 15px; background: #d4edda; border-radius: 6px; display: none;}
            .template-download-link a {color: #155724; font-weight: bold; text-decoration: none;}
            /* 禁用的项目样式 */
            .checkbox-item.disabled {opacity: 0.4; pointer-events: none;}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="logout">
                <a href="/logout">退出登录</a>
            </div>
            <div class="nav">
                <a href="/" class="active">首页</a>
                <a href="/notes">使用须知</a>
            </div>
            <h1>体育成绩批处理智能分析系统</h1>

            <!-- 模板下载区域 -->
            <div class="template-section">
                <button class="template-btn" id="templateBtn">下载模板</button>
                <div class="template-download-link" id="templateDownloadLink">
                    模板生成成功！<a href="" id="templateLink" target="_blank">点击下载</a>
                </div>
            </div>

            <input type="file" class="file-input" id="fileInput" multiple accept=".xlsx,.xls" onchange="showSelectedFiles(this)">

            <!-- 批量班级成绩处理 -->
            <div class="section-title">批量班级成绩处理（选择各班成绩单表格）</div>
            <div class="grade-buttons">
                <button class="grade-btn" onclick="startProcess('1')">一年级</button>
                <button class="grade-btn" onclick="startProcess('2')">二年级</button>
                <button class="grade-btn" onclick="startProcess('3')">三年级</button>
                <button class="grade-btn" onclick="startProcess('4')">四年级</button>
                <button class="grade-btn" onclick="startProcess('5')">五年级</button>
                <button class="grade-btn" onclick="startProcess('6')">六年级</button>
            </div>

            <!-- 年级前15名排名 -->
            <div class="section-title">年级前15名排名（选择各班成绩处理后表格）</div>
            <div class="grade-buttons">
                <button class="batch-btn" onclick="startBatchSort('1')">一年级</button>
                <button class="batch-btn" onclick="startBatchSort('2')">二年级</button>
                <button class="batch-btn" onclick="startBatchSort('3')">三年级</button>
                <button class="batch-btn" onclick="startBatchSort('4')">四年级</button>
                <button class="batch-btn" onclick="startBatchSort('5')">五年级</button>
                <button class="batch-btn" onclick="startBatchSort('6')">六年级</button>
            </div>

            <!-- 已选文件 -->
            <div class="file-list">
                <h4>已选文件：</h4>
                <div id="selectedFiles">暂无文件</div>
            </div>

            <!-- 状态和耗时 -->
            <div class="status status-wait" id="processStatus">请选择文件，点击对应按钮开始处理</div>
            <div class="time-info" id="timeInfo"></div>

            <!-- 原有处理下载区（展示每个文件耗时） -->
            <div class="download-area" id="downloadArea">
                <h3>成绩处理完成，可下载文件：</h3>
                <div id="downloadList"></div>
            </div>

            <!-- 批量排序下载区 -->
            <div class="batch-download-area" id="batchDownloadArea">
                <h3 style="color:#28a745">批量排序完成，可下载汇总表：</h3>
                <div id="batchDownloadList"></div>
            </div>
        </div>

        <!-- 模板生成模态框 -->
        <div class="modal" id="templateModal">
            <div class="modal-content">
                <!-- 步骤1：所在学校 -->
                <div class="modal-step active" id="step1">
                    <h3>步骤1/7：填写所在学校</h3>
                    <div class="form-group">
                        <label for="schoolName">所在学校：</label>
                        <input type="text" id="schoolName" placeholder="请输入学校名称" required>
                    </div>
                    <div class="btn-group">
                        <button class="btn btn-next" onclick="nextStep()">下一步</button>
                    </div>
                </div>

                <!-- 步骤2：表头名称 -->
                <div class="modal-step" id="step2">
                    <h3>步骤2/7：填写表头名称 </h3>
                    <div class="form-group">
                        <label for="headerName">表头名称： 如：*月班级联赛</label>
                        <input type="text" id="headerName" placeholder="如：*月班级联赛" required>
                    </div>
                    <div class="btn-group">
                        <button class="btn btn-prev" onclick="prevStep()">上一步</button>
                        <button class="btn btn-next" onclick="nextStep()">下一步</button>
                    </div>
                </div>

                <!-- 步骤3：选择年级 -->
                <div class="modal-step" id="step3">
                    <h3>步骤3/7：选择下载几年级模板</h3>
                    <div class="form-group">
                        <label>选择年级：</label>
                        <div class="grade-buttons">
                            <button class="grade-btn" onclick="selectGrade('1')">一年级</button>
                            <button class="grade-btn" onclick="selectGrade('2')">二年级</button>
                            <button class="grade-btn" onclick="selectGrade('3')">三年级</button>
                            <button class="grade-btn" onclick="selectGrade('4')">四年级</button>
                            <button class="grade-btn" onclick="selectGrade('5')">五年级</button>
                            <button class="grade-btn" onclick="selectGrade('6')">六年级</button>
                        </div>
                        <input type="hidden" id="selectedGrade" value="">
                    </div>
                    <div class="btn-group">
                        <button class="btn btn-prev" onclick="prevStep()">上一步</button>
                        <button class="btn btn-next" onclick="nextStep()" id="gradeNextBtn" disabled>下一步</button>
                    </div>
                </div>

                <!-- 步骤4：选择未测试的项目 -->
                <div class="modal-step" id="step4">
                    <h3>步骤4/7：填写未测试的项目（可多选）</h3>
                    <div class="form-group">
                        <label>可选项目：</label>
                        <div class="checkbox-group">
                            <div class="checkbox-item" data-grade="1,2,3,4,5,6"><input type="checkbox" id="item1" value="50米"> <label for="item1">50米</label></div>
                            <div class="checkbox-item" data-grade="1,2,3,4,5,6"><input type="checkbox" id="item2" value="跳绳"> <label for="item2">跳绳</label></div>
                            <div class="checkbox-item" data-grade="1,2,3,4,5,6"><input type="checkbox" id="item3" value="坐位体前屈"> <label for="item3">坐位体前屈</label></div>
                            <div class="checkbox-item" data-grade="2,3,4,5,6"><input type="checkbox" id="item4" value="仰卧起坐"> <label for="item4">仰卧起坐</label></div>
                            <div class="checkbox-item" data-grade="5,6"><input type="checkbox" id="item5" value="50*8"> <label for="item5">50*8</label></div>
                            <div class="checkbox-item" data-grade="3,4"><input type="checkbox" id="item6" value="立定跳远"> <label for="item6">立定跳远</label></div>
                            <div class="checkbox-item" data-grade="5,6"><input type="checkbox" id="item7" value="直臂悬垂"> <label for="item7">直臂悬垂</label></div>
                            <div class="checkbox-item" data-grade="1,2,3,4,5,6"><input type="checkbox" id="item8" value="身高"> <label for="item8">身高</label></div>
                            <div class="checkbox-item" data-grade="1,2,3,4,5,6"><input type="checkbox" id="item9" value="体重"> <label for="item9">体重</label></div>
                            <div class="checkbox-item" data-grade="1,2,3,4,5,6"><input type="checkbox" id="item10" value="肺活量"> <label for="item10">肺活量</label></div>
                        </div>
                    </div>
                    <div class="btn-group">
                        <button class="btn btn-prev" onclick="prevStep()">上一步</button>
                        <button class="btn btn-next" onclick="nextStep()">下一步</button>
                    </div>
                </div>

                <!-- 步骤5：男生人数 -->
                <div class="modal-step" id="step5">
                    <h3>步骤5/7：填写男生人数</h3>
                    <div class="form-group">
                        <label for="boyCount">男生人数（0-28）：</label>
                        <input type="number" id="boyCount" min="0" max="28" value="0" required>
                    </div>
                    <div class="btn-group">
                        <button class="btn btn-prev" onclick="prevStep()">上一步</button>
                        <button class="btn btn-next" onclick="nextStep()">下一步</button>
                    </div>
                </div>

                <!-- 步骤6：女生人数 -->
                <div class="modal-step" id="step6">
                    <h3>步骤6/7：填写女生人数</h3>
                    <div class="form-group">
                        <label for="girlCount">女生人数（0-28）：</label>
                        <input type="number" id="girlCount" min="0" max="28" value="0" required>
                    </div>
                    <div class="btn-group">
                        <button class="btn btn-prev" onclick="prevStep()">上一步</button>
                        <button class="btn btn-next" onclick="nextStep()">下一步</button>
                    </div>
                </div>

                <!-- 步骤7：文件名称 -->
                <div class="modal-step" id="step7">
                    <h3>步骤7/7：输入文件名称 如：1.1</h3>
                    <div class="form-group">
                        <label for="fileName">文件名称：</label>
                        <input type="text" id="fileName" placeholder="请输入模板文件名称" required>
                    </div>
                    <div class="btn-group">
                        <button class="btn btn-prev" onclick="prevStep()">上一步</button>
                        <button class="btn btn-submit" onclick="submitTemplate()">生成模板</button>
                    </div>
                </div>
            </div>
        </div>

        <script>
            let selectedFiles = [];
            let pollInterval = null;
            let currentStep = 1;
            let selectedGrade = '';
            let templateData = {};

            // 显示已选文件
            function showSelectedFiles(input) {
                selectedFiles = Array.from(input.files);
                const fileListDiv = document.getElementById('selectedFiles');
                if (selectedFiles.length === 0) {
                    fileListDiv.innerHTML = '暂无文件';
                    return;
                }
                let fileHtml = '';
                selectedFiles.forEach((file, index) => {
                    fileHtml += `<div class="file-item">${index + 1}. ${file.name}</div>`;
                });
                fileListDiv.innerHTML = fileHtml;
            }

            // 状态轮询（按操作类型展示，无优先级，保留每个文件耗时）
            function pollStatus() {
                if (pollInterval) clearInterval(pollInterval);
                pollInterval = setInterval(() => {
                    fetch('/get_all_status')
                    .then(res => res.json())
                    .then(data => {
                        const statusDiv = document.getElementById('processStatus');
                        const timeDiv = document.getElementById('timeInfo');
                        const downloadArea = document.getElementById('downloadArea');
                        const batchArea = document.getElementById('batchDownloadArea');
                        const downloadList = document.getElementById('downloadList');
                        const batchList = document.getElementById('batchDownloadList');
                        // 初始化重置
                        statusDiv.className = 'status status-wait';
                        downloadArea.style.display = 'none';
                        batchArea.style.display = 'none';
                        timeDiv.textContent = '';
                        downloadList.innerHTML = '';
                        batchList.innerHTML = '';
                        // 按后端返回的操作类型精准展示
                        switch (data.op_type) {
                            case "normal":
                                statusDiv.className = 'status status-processing';
                                statusDiv.textContent = `正在处理${data.normal_grade}年级：${data.normal_current}`;
                                timeDiv.textContent = `当前文件耗时：${data.normal_cur_cost}s | 任务总耗时：${data.normal_total_cost}s`;
                                break;
                            case "normal_done":
                                statusDiv.className = 'status status-done';
                                statusDiv.textContent = `${data.normal_grade}年级普通处理完成！`;
                                timeDiv.textContent = `任务总耗时：${data.normal_total_cost}s`;
                                // 渲染普通处理列表，**展示每个文件的单独耗时**
                                data.normal_files.forEach((it, i) => {
                                    if (it.error) {
                                        downloadList.innerHTML += `<div class="error-item">${i+1}. ${it.name} → ${it.error} <span class="cost-time">（耗时${it.cost_time}s）</span></div>`;
                                    } else {
                                        downloadList.innerHTML += `<div class="download-item"><a href="${it.url}" target="_blank">${i+1}. ${it.name}（${it.grade}年级体育成绩单）</a> <span class="cost-time">（耗时${it.cost_time}s）</span></div>`;
                                    }
                                });
                                downloadArea.style.display = 'block';
                                clearInterval(pollInterval);
                                break;
                            case "batch":
                                statusDiv.className = 'status status-processing';
                                statusDiv.textContent = `正在${data.batch_grade}年级批量排序：共${data.batch_file_count}个文件`;
                                timeDiv.textContent = `任务已耗时：${data.batch_cost}s`;
                                break;
                            case "batch_done":
                                statusDiv.className = 'status status-done';
                                statusDiv.textContent = `${data.batch_grade}年级批量排序完成！`;
                                timeDiv.textContent = `任务总耗时：${data.batch_cost}s`;
                                data.batch_files.forEach(item => {
                                    if (item.error) {
                                        batchList.innerHTML += `<div class="error-item">批量处理失败：${item.error} <span class="cost-time">（耗时${item.cost_time}s）</span></div>`;
                                    } else {
                                        batchList.innerHTML += `<div class="download-item"><a href="${item.url}" target="_blank">${item.grade}年级 | 共${item.file_count}个文件 | 按年级前十五名排序汇总表</a> <span class="cost-time">（总耗时${item.cost_time}s）</span></div>`;
                                    }
                                });
                                batchArea.style.display = 'block';
                                clearInterval(pollInterval);
                                break;
                            default:
                                statusDiv.textContent = '请选择文件，点击对应按钮开始处理';
                                break;
                        }
                    }).catch(err => {
                        statusDiv.className = 'status status-error';
                        statusDiv.textContent = '状态获取失败，请刷新页面重试';
                        console.error('状态查询错误：', err);
                    });
                }, 500);
            }

            // 原有成绩处理按钮点击事件
            function startProcess(grade) {
                if (selectedFiles.length === 0) { 
                    alert('请先选择要处理的Excel文件！'); 
                    return; 
                }
                document.getElementById('fileInput').value = ''; // 清空文件选择框
                const statusDiv = document.getElementById('processStatus');
                statusDiv.className = 'status status-processing';
                statusDiv.textContent = `正在上传${grade}年级文件，准备处理...`;
                const fd = new FormData();
                selectedFiles.forEach(f => fd.append('files', f));
                fetch(`/upload_files/${grade}`, {method: 'POST', body: fd})
                .then(res => res.json())
                .then(() => pollStatus())
                .catch(err => {
                    statusDiv.className = 'status status-error';
                    statusDiv.textContent = '文件上传失败，请重试';
                });
            }

            // 批量排序按钮点击事件
            function startBatchSort(grade) {
                if (selectedFiles.length === 0) { 
                    alert('请先选择要批量处理的Excel文件！'); 
                    return; 
                }
                document.getElementById('fileInput').value = ''; // 清空文件选择框
                const statusDiv = document.getElementById('processStatus');
                statusDiv.className = 'status status-processing';
                statusDiv.textContent = `正在上传${grade}年级批量文件，准备排序...`;
                const fd = new FormData();
                selectedFiles.forEach(f => fd.append('files', f));
                fetch(`/upload_batch_sort/${grade}`, {method: 'POST', body: fd})
                .then(res => res.json())
                .then(() => pollStatus())
                .catch(err => {
                    statusDiv.className = 'status status-error';
                    statusDiv.textContent = '批量文件上传失败，请重试';
                });
            }

            // 模板生成相关函数
            document.getElementById('templateBtn').addEventListener('click', function() {
                document.getElementById('templateModal').style.display = 'flex';
                currentStep = 1;
                showStep(currentStep);
                // 重置模板数据和项目选择
                templateData = {};
                selectedGrade = '';
                document.getElementById('selectedGrade').value = '';
                document.getElementById('gradeNextBtn').disabled = true;
                document.querySelectorAll('#step3 .grade-btn').forEach(btn => {
                    btn.style.background = '#007bff';
                });
                document.getElementById('templateDownloadLink').style.display = 'none';
                // 重置所有项目选择
                document.querySelectorAll('.checkbox-group input').forEach(input => {
                    input.checked = false;
                });
                document.querySelectorAll('.checkbox-item').forEach(item => {
                    item.classList.remove('disabled');
                });
            });

            function showStep(step) {
                // 隐藏所有步骤
                document.querySelectorAll('.modal-step').forEach(el => {
                    el.classList.remove('active');
                });
                // 显示当前步骤
                document.getElementById(`step${step}`).classList.add('active');
                currentStep = step;

                // 进入步骤4时，根据选中年级过滤项目
                if(step === 4){
                    filterItemsByGrade(selectedGrade);
                }
            }

            // 根据年级过滤可选项目
            function filterItemsByGrade(grade){
                if(!grade) return;
                document.querySelectorAll('.checkbox-item').forEach(item => {
                    const allowedGrades = item.getAttribute('data-grade').split(',');
                    if(allowedGrades.includes(grade)){
                        item.classList.remove('disabled');
                    }else{
                        item.classList.add('disabled');
                        // 取消禁用项目的勾选
                        item.querySelector('input').checked = false;
                    }
                });
            }

            function nextStep() {
                // 保存当前步骤数据
                switch(currentStep) {
                    case 1:
                        templateData.school_name = document.getElementById('schoolName').value.trim();
                        if (!templateData.school_name) {
                            alert('请输入学校名称！');
                            return;
                        }
                        break;
                    case 2:
                        templateData.header_name = document.getElementById('headerName').value.trim();
                        if (!templateData.header_name) {
                            alert('请输入表头名称！');
                            return;
                        }
                        break;
                    case 3:
                        if (!selectedGrade) {
                            alert('请选择年级！');
                            return;
                        }
                        templateData.grade = selectedGrade;
                        break;
                    case 4:
                        // 获取选中的项目
                        const selectedItems = [];
                        document.querySelectorAll('.checkbox-group input:checked').forEach(el => {
                            selectedItems.push(el.value);
                        });
                        templateData.selected_items = selectedItems;
                        break;
                    case 5:
                        const boyCount = parseInt(document.getElementById('boyCount').value);
                        if (isNaN(boyCount) || boyCount < 0 || boyCount > 28) {
                            alert('男生人数必须是0-28之间的数字！');
                            return;
                        }
                        templateData.boy_count = boyCount;
                        break;
                    case 6:
                        const girlCount = parseInt(document.getElementById('girlCount').value);
                        if (isNaN(girlCount) || girlCount < 0 || girlCount > 28) {
                            alert('女生人数必须是0-28之间的数字！');
                            return;
                        }
                        templateData.girl_count = girlCount;
                        break;
                }

                if (currentStep < 7) {
                    showStep(currentStep + 1);
                }
            }

            function prevStep() {
                if (currentStep > 1) {
                    showStep(currentStep - 1);
                }
            }

            function selectGrade(grade) {
                selectedGrade = grade;
                document.getElementById('selectedGrade').value = grade;
                document.getElementById('gradeNextBtn').disabled = false;
                // 高亮选中的年级按钮
                document.querySelectorAll('#step3 .grade-btn').forEach(btn => {
                    btn.style.background = '#007bff';
                });
                event.target.style.background = '#0056b3';
            }

            function submitTemplate() {
                templateData.file_name = document.getElementById('fileName').value.trim();
                if (!templateData.file_name) {
                    alert('请输入文件名称！');
                    return;
                }

                // 提交生成模板请求
                fetch('/generate_template', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify(templateData)
                })
                .then(res => res.json())
                .then(data => {
                    if (data.code === 200) {
                        // 显示下载链接
                        document.getElementById('templateModal').style.display = 'none';
                        document.getElementById('templateLink').href = data.download_url;
                        document.getElementById('templateDownloadLink').style.display = 'block';
                    } else {
                        alert(data.msg);
                    }
                })
                .catch(err => {
                    alert('模板生成失败：' + err.message);
                });
            }

            // 关闭模态框
            window.addEventListener('click', function(event) {
                if (event.target === document.getElementById('templateModal')) {
                    document.getElementById('templateModal').style.display = 'none';
                }
            });
        </script>
    </body>
    </html>
    '''
    return render_template_string(html_template)


# ========== 新增：使用本系统须知页面 ==========
@app.route('/notes')
@login_required
def notes():
    html = '''
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <title>使用系统须知 - 体育成绩处理系统</title>
        <style>
            * {margin: 0; padding: 0; box-sizing: border-box; font-family: Arial, 微软雅黑;}
            .container {width: 900px; margin: 50px auto; padding: 20px; border: 1px solid #e0e0e0; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.05);}
            h1 {text-align: center; color: #333; margin-bottom: 30px;}
            .nav {text-align: center; margin-bottom: 20px;}
            .nav a {padding: 8px 16px; margin: 0 5px; background: #6c757d; color: white; border-radius: 4px; text-decoration: none;}
            .nav a.active {background: #007bff;}
            .logout {text-align:center; margin-bottom:15px;}
            .logout a {color:red; text-decoration:none;}
            .content {line-height: 1.8; font-size: 16px; color: #333; padding: 10px 20px;}
            .content h3 {margin: 15px 0 10px; color: #007bff;}
            .content p {margin: 8px 0; padding-left: 10px;}
            .content ul {padding-left: 30px;}
            .content li {margin: 5px 0;}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="logout">
                <a href="/logout">退出登录</a>
            </div>
            <div class="nav">
                <a href="/">首页</a>
                <a href="/notes" class="active">使用须知</a>
            </div>
            <h1>使用系统须知</h1>
            <div class="content">



                <h3>一、系统说明</h3>
                <ul>
                    <li>本系统是历下区统一体育测试标准的教学辅助工具，专为体育教师设计，可实现体育成绩的批量处理、智能分析与教学策略生成，解决传统人工统计效率低、分析难的痛点。</li>
                    <li>系统核心功能：</li>
                    <li>1. 标准化成绩录入模板下载</li>
                    <li>2. 批量数据处理与自动成绩核算</li>
                    <li>3. 班级成绩排名与优良率统计</li>
                    <li>4. 可视化数据图表生成</li>
                    <li>5. 基于成绩分析的教学训练策略建议</li>
                </ul>
                <h3>二、模板下载填写说明</h3>
                <ul>
                    <li>（一）学校名称：填写学校全称</li>
                    <li>（二）表头名称：填写测试或赛事名称，如：**小学*月份班级联赛</li>
                    <li>（三）选择年级：下载对应年级的专用成绩录入模板</li>
                    <li>（四）未测试项目：测试时如果没有测全，可以选择该年级没有测试项目，系统会自动填写满分的数据</li>
                    <li>（五）男生人数：按班级实际人数填写</li>
                    <li>（六）女生人数：按班级实际人数填写</li>
                    <li>（七）文件名称：按“年级.班级”格式填写，如：一年级一班 1.1（排名必须使用此格式）</li>
                </ul>
                <h3>三、批量出成绩使用说明</h3>
                <ul>
                    <li>上传已填写完成的成绩模板文件（文件名如：1.1），支持同年级多文件批量上传。</li>
                    <li>系统自动生成：学生各项成绩与总成绩报表、班级优良率报表。</li>
                    <li>同步生成可视化柱状图，并根据优良率自动推荐教学训练提升方案。</li>
                </ul>

                    <h3>四、自动出前十名使用说明</h3>
                <ul>
                    <li>上传同年级所有班级成绩文件（文件名如：1年级体育成绩单_1.1），系统自动统一排名，可直接下载结果。</li>
                </ul>

                    <h3>五、成绩填写规范</h3>
                <ul>
                    <li>按照提示单位填写成绩</li>
                    <li>有序号的行不得留空单元格（“特殊男”“特殊女”除外）</li>
                    <li>不得修改表格原有格式，不得删除行</li>
                    <li>男生成绩从A5单元格开始填写，女生成绩从A37单元格开始填写</li>
                </ul>

                    <h3>六、成绩处理规则</h3>
                <ul>
                    <li>总分计算：自动去掉10%的学生成绩，人数按四舍五入计算</li>
                    <li>特殊学生：性别标注为“特殊男”“特殊女”，不计入总分统计</li>
                </ul>



                <h3>七、各项输入范围</h3>
                <ul>
                    <h4>一年级 </h4> 
                    <li>50米 （7-20s）跳绳 （0-500） 坐位体前屈 （-30-30CM）  肺活量 （0-5000ML）</li>

                    <h4>二年级 </h4>
                    <li>50米 （7-20s）跳绳 （0-500） 坐位体前屈 （-30-40CM）  仰卧起坐 （0-80）  肺活量 （0-6000ML）</li>

                    <h4>三年级 </h4>
                    <li>50米 （7-20s）跳绳 （0-500） 坐位体前屈 （-30-40CM）  仰卧起坐 （0-80）</li>
                    <li>立定跳远 （0-220CM）  肺活量 （0-6000ML）</li>

                    <h4>四年级 </h4>
                    <li>50米 （6-20s）跳绳 （0-500） 坐位体前屈 （-30-40CM）  仰卧起坐 （0-80）</li>
                    <li>立定跳远 （0-230CM）  肺活量 （0-6000ML）</li>

                    <h4>五年级 </h4>
                    <li>50米 （6-20s）跳绳 （0-500） 坐位体前屈 （-30-40CM）  仰卧起坐 （0-80）</li>
                    <li>50*8 （1'00-4'00） 直臂悬垂 （0-220s） 肺活量 （0-7000ML）</li>

                    <h4>六年级 </h4>
                    <li>50米 （6-20s）跳绳 （0-500） 坐位体前屈 （-30-40CM）  仰卧起坐 （0-80）</li>
                    <li>50*8 （1'00-4'00） 直臂悬垂 （0-220s） 肺活量 （0-7000ML）</li>


                </ul>

                <h3>八、文件格式要求</h3>
                <ul>
                    <li>仅支持 .xlsx 格式的 Excel 文件</li>
                    <li>文件请勿加密、请勿设置复杂格式</li>

                </ul>

                <h3>九、使用流程说明</h3>
                <ul>
                    <li>先下载对应年级模板，按模板填写成绩</li>
                    <li>首页选择文件 → 点击对应年级按钮进行处理</li>
                    <li>如需年级排名，上传处理后的文件进行批量排序</li>
                </ul>

                <h3>十、常见问题</h3>
                <ul>
                    <li>处理失败：检查文件格式、数据是否准确</li>
                    <li>无反应：刷新页面重新上传，检查网络</li>
                    <li>下载异常：更换浏览器（推荐 Chrome / Edge）</li>
                </ul>

                <h3>十一、安全提示</h3>
                <ul>
                    <li>本系统仅在本地运行，数据不会上传外网</li>
                    <li>处理完成后及时下载文件，避免数据丢失</li>
                    <li>不要上传与成绩无关的文件</li>
                </ul>  
            </div>
        </div>
    </body>
    </html>
    '''
    return render_template_string(html)


# ========== 原有文件上传接口（修改：清空批量状态） ==========
@app.route('/upload_files/<grade>', methods=['POST'])
@login_required
def upload_files(grade):
    global task_start_time, processed_files
    # 清空批量排序的所有状态和队列
    global batch_current_processing, batch_task_start, batch_processed_files, batch_current_grade, batch_sort_queue
    batch_current_processing = ""
    batch_task_start = None
    batch_processed_files = []
    batch_current_grade = ""
    while not batch_sort_queue.empty():
        try:
            batch_sort_queue.get_nowait()
        except queue.Empty:
            break
    # 重置原有处理状态
    processed_files = []
    task_start_time = datetime.now()
    # 处理文件上传
    files = request.files.getlist('files')
    if not files:
        return jsonify({"msg": "未接收到任何文件", "code": 400}), 400
    success_count = 0
    for file in files:
        try:
            if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
                continue
            origin_name = os.path.basename(file.filename)
            safe_name = secure_filename(origin_name)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S%f')
            upload_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{safe_name}")
            file.save(upload_path)
            file_queue.put((upload_path, origin_name, grade))
            success_count += 1
        except Exception as e:
            print(f"文件上传失败：{origin_name} - {str(e)}")
            continue
    return jsonify({
        "msg": f"{grade}年级文件上传完成，共{success_count}个文件加入处理队列",
        "code": 200, "success_count": success_count
    })


# ========== 批量排序上传接口（修改：清空原有状态） ==========
@app.route('/upload_batch_sort/<grade>', methods=['POST'])
@login_required
def upload_batch_sort(grade):
    global batch_task_start, batch_processed_files
    # 清空原有成绩处理的所有状态和队列
    global current_processing, current_file_start_time, task_start_time, processed_files, current_grade, file_queue
    current_processing = ""
    current_file_start_time = None
    task_start_time = None
    processed_files = []
    current_grade = ""
    while not file_queue.empty():
        try:
            file_queue.get_nowait()
        except queue.Empty:
            break
    # 重置批量排序状态
    batch_processed_files = []
    batch_task_start = datetime.now()
    files = request.files.getlist('files')
    paths = []
    names = []
    for f in files:
        if not f.filename.lower().endswith(('.xlsx', '.xls')):
            continue
        origin = os.path.basename(f.filename)
        safe = secure_filename(origin)
        ts = datetime.now().strftime('%Y%m%d%H%M%S%f')
        p = os.path.join(UPLOAD_FOLDER, f'batch_{ts}_{safe}')
        f.save(p)
        paths.append(p)
        names.append(origin)
    if paths:
        batch_sort_queue.put((paths, names, grade))
    return jsonify({"code": 200, "count": len(paths)})


# ========== 统一状态查询接口（修复500错误+新增操作类型） ==========
@app.route('/get_all_status')
@login_required
def get_all_status():
    global current_processing, current_file_start_time, task_start_time, processed_files, file_queue, current_grade
    global batch_current_processing, batch_task_start, batch_processed_files, batch_sort_queue, batch_current_grade
    # 原有处理状态计算（保留当前文件耗时、总耗时）
    n_cur_cost = round((datetime.now() - current_file_start_time).total_seconds(), 2) if (
            current_file_start_time and current_processing) else 0
    n_total_cost = round((datetime.now() - task_start_time).total_seconds(), 2) if task_start_time else 0
    n_running = (current_processing != "")
    n_done = (file_queue.qsize() == 0 and not n_running and len(processed_files) > 0)
    # 批量排序状态计算（修复len()错误）
    b_total_cost = round((datetime.now() - batch_task_start).total_seconds(), 2) if batch_task_start else 0
    b_running = (batch_current_grade != "")
    b_done = (batch_sort_queue.qsize() == 0 and not b_running and len(batch_processed_files) > 0)
    b_cnt = batch_processed_files[-1]["file_count"] if b_done and batch_processed_files else 0
    # 操作类型标识（核心）
    op_type = "none"
    if n_running:
        op_type = "normal"
    elif n_done:
        op_type = "normal_done"
    elif b_running:
        op_type = "batch"
    elif b_done:
        op_type = "batch_done"
    # 返回状态
    return jsonify({
        "op_type": op_type,
        # 原有处理（保留所有耗时字段）
        "normal_running": n_running, "normal_done": n_done,
        "normal_current": current_processing, "normal_grade": current_grade,
        "normal_cur_cost": n_cur_cost, "normal_total_cost": n_total_cost,
        "normal_files": processed_files,
        # 批量排序
        "batch_running": b_running, "batch_done": b_done,
        "batch_grade": batch_current_grade, "batch_file_count": b_cnt,
        "batch_cost": b_total_cost, "batch_files": batch_processed_files
    })


# ========== 原有状态查询接口（保留，兼容旧逻辑） ==========
@app.route('/get_process_status')
@login_required
def get_process_status():
    global current_processing, current_file_start_time, task_start_time, processed_files, file_queue, current_grade
    current_file_cost = 0.0
    if current_file_start_time and current_processing:
        current_file_cost = round((datetime.now() - current_file_start_time).total_seconds(), 2)
    total_cost = 0.0
    if task_start_time:
        total_cost = round((datetime.now() - task_start_time).total_seconds(), 2)
    is_done = (file_queue.qsize() == 0) and (current_processing == "") and (len(processed_files) > 0)
    return jsonify({
        "current": current_processing, "current_grade": current_grade,
        "current_file_cost": current_file_cost, "pending_count": file_queue.qsize(),
        "total_cost": total_cost, "done": is_done, "files": processed_files
    })


# ========== 文件下载接口（兼容两种操作） ==========
@app.route('/download/<filename>')
@login_required
def download_file(filename):
    # 先查原有处理文件，再查批量排序文件
    file_info = next((f for f in processed_files if f["url"].endswith(filename)), None)
    if not file_info:
        file_info = next((f for f in batch_processed_files if f["url"].endswith(filename)), None)
    if not file_info:
        return "文件不存在或未处理完成", 404
    file_path = os.path.join(DOWNLOAD_FOLDER, filename)
    if not os.path.exists(file_path):
        return "文件已被删除或处理失败", 404
    # 生成友好下载名
    if "前15名排序汇总" in filename:
        download_name = filename
    else:
        download_name = f"{file_info['grade']}年级体育成绩单_{file_info['name']}"
    return send_file(
        file_path, as_attachment=True, download_name=download_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# 方式1：默认清空当前脚本所在目录下的四个文件夹
clean_target_folders()
# ========== 程序入口 ==========
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)